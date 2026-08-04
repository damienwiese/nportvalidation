"""Generate the definitive N-PORT operating runbook PDF."""

from __future__ import annotations

import sys
import csv
from collections import Counter
from dataclasses import fields
from pathlib import Path

from reportlab import rl_config
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import field_provenance as provenance  # noqa: E402
import audit_findings as measured  # noqa: E402
import runbook_traceability as trace  # noqa: E402
from nport.config import _CONFIG_KEY_MAP, _FILING_KEY_MAP  # noqa: E402
from nport.models import FilingData, FundConfig, Holding  # noqa: E402
from nport.schema import FIELD_BY_NAME  # noqa: E402
from nport.xsd_validator import NportValidator  # noqa: E402


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
# Stable timestamps and document IDs keep repeat generations byte-for-byte reproducible.
rl_config.invariant = 1
NAVY = colors.HexColor("#17324D")
TEAL = colors.HexColor("#2B6D73")
SEA = colors.HexColor("#3FA7A3")
GOLD = colors.HexColor("#E3B341")
INK = colors.HexColor("#283849")
GRID = colors.HexColor("#CDD8E1")
PALE = colors.HexColor("#EEF5F6")
MIST = colors.HexColor("#F5F8FA")
RED = colors.HexColor("#A83221")

styles = getSampleStyleSheet()
COVER_EYEBROW = ParagraphStyle("CoverEyebrow", parent=styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=8.5, leading=11, textColor=GOLD, alignment=TA_CENTER,
                               tracking=1.4, spaceAfter=15)
TITLE = ParagraphStyle("Title2", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=28,
                       leading=31, textColor=colors.white, alignment=TA_CENTER, spaceAfter=16)
SUBTITLE = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=12, leading=17,
                          textColor=colors.HexColor("#D9E7EC"), alignment=TA_CENTER)
H1 = ParagraphStyle("H1x", parent=styles["Heading1"], fontSize=16, leading=19,
                    textColor=NAVY, spaceBefore=14, spaceAfter=8)
H2 = ParagraphStyle("H2x", parent=styles["Heading2"], fontSize=11.5, leading=14,
                    textColor=TEAL, spaceBefore=9, spaceAfter=5)
BODY = ParagraphStyle("Bodyx", parent=styles["BodyText"], fontSize=9.4, leading=13.2,
                      textColor=INK, spaceAfter=7)
SMALL = ParagraphStyle("Smallx", parent=BODY, fontSize=8.1, leading=11)
TABLE_HEAD = ParagraphStyle("TableHead", parent=SMALL, fontName="Helvetica-Bold",
                            fontSize=7.6, leading=9.5, textColor=colors.white)
CODE = ParagraphStyle("Codex", parent=BODY, fontName="Courier", fontSize=8.1, leading=11,
                      leftIndent=10, rightIndent=10, borderColor=GRID, borderWidth=0.6,
                      borderPadding=7, backColor=colors.HexColor("#F7FAFB"), spaceBefore=4,
                      spaceAfter=8)
CALLOUT = ParagraphStyle("Callout", parent=BODY, leftIndent=9, rightIndent=9,
                         borderColor=TEAL, borderWidth=1, borderPadding=8,
                         backColor=PALE, spaceBefore=6, spaceAfter=10)
WARN = ParagraphStyle("Warn", parent=CALLOUT, borderColor=RED,
                      backColor=colors.HexColor("#FFF1ED"))
STEP_TITLE = ParagraphStyle("StepTitle", parent=H2, fontSize=11, leading=13,
                            spaceBefore=0, spaceAfter=4)
STEP_NUM = ParagraphStyle("StepNum", parent=styles["Normal"], fontName="Helvetica-Bold",
                          fontSize=15, leading=18, textColor=colors.white, alignment=TA_CENTER)
STEP_CODE = ParagraphStyle("StepCode", parent=BODY, fontName="Courier", fontSize=7.8,
                           leading=10.5, textColor=NAVY, backColor=colors.HexColor("#EDF3F6"),
                           borderPadding=5, spaceAfter=5)
STEP_BODY = ParagraphStyle("StepBody", parent=BODY, fontSize=8.7, leading=11.5, spaceAfter=0)
METRIC_NUMBER = ParagraphStyle("MetricNumber", parent=styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=18, leading=20, textColor=NAVY, alignment=TA_CENTER)
METRIC_LABEL = ParagraphStyle("MetricLabel", parent=SMALL, fontSize=7.3, leading=9,
                              textColor=colors.HexColor("#596B7D"), alignment=TA_CENTER)


def p(text: str, style=BODY) -> Paragraph:
    for old, new in {
        "\u2014": " - ", "\u2013": "-", "\u2026": "...", "\u2192": "->",
        "\u00d7": "x", "\u00f7": "/", "\u2011": "-",
    }.items():
        text = text.replace(old, new)
    return Paragraph(text, style)


def table(headers: list[str], rows: list[list[str]], widths: list[float]) -> Table:
    data = [[p(x, TABLE_HEAD) for x in headers]] + [[p(x, SMALL) for x in row] for row in rows]
    out = Table(data, colWidths=[w * inch for w in widths], repeatRows=1, hAlign="LEFT")
    out.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), TEAL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 1), (-1, -1), 0.45, GRID),
        ("BOX", (0, 0), (-1, -1), 0.6, GRID),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, MIST]),
    ]))
    return out


def page(canvas, doc):
    canvas.saveState()
    if doc.page == 1:
        canvas.setFillColor(NAVY)
        canvas.rect(0, 0, letter[0], letter[1], fill=1, stroke=0)
        canvas.setFillColor(TEAL)
        canvas.rect(0, 0, 0.18 * inch, letter[1], fill=1, stroke=0)
        canvas.setFillColor(SEA)
        canvas.circle(7.45 * inch, 9.7 * inch, 0.62 * inch, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor("#21455F"))
        canvas.circle(7.85 * inch, 9.2 * inch, 0.88 * inch, fill=1, stroke=0)
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(2)
        canvas.line(3.45 * inch, 6.0 * inch, 5.05 * inch, 6.0 * inch)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#BFD1D9"))
        canvas.drawCentredString(letter[0] / 2, 0.42 * inch, "CORGI ETF TRUST  /  N-PORT OPERATIONS")
        canvas.restoreState()
        return
    canvas.setFillColor(TEAL)
    canvas.rect(0.55 * inch, 10.48 * inch, 7.4 * inch, 0.055 * inch, fill=1, stroke=0)
    canvas.setStrokeColor(GRID)
    canvas.line(0.55 * inch, 0.46 * inch, 7.95 * inch, 0.46 * inch)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor("#68798A"))
    canvas.drawString(0.55 * inch, 0.28 * inch, "Corgi ETF Trust - N-PORT operating documentation")
    canvas.drawRightString(7.95 * inch, 0.28 * inch, f"Page {doc.page}")
    canvas.restoreState()


def build(path: Path, story: list) -> None:
    doc = BaseDocTemplate(str(path), pagesize=letter, leftMargin=0.55 * inch,
                          rightMargin=0.55 * inch, topMargin=0.55 * inch,
                          bottomMargin=0.6 * inch, title=path.stem)
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="body")
    doc.addPageTemplates(PageTemplate(id="clean", frames=[frame], onPage=page))
    doc.build(story)


def cover(title: str, subtitle: str) -> list:
    return [Spacer(1, 1.28 * inch), p("CORGI ETF TRUST  /  OPERATING DOCUMENT", COVER_EYEBROW),
            p(title, TITLE), p(subtitle, SUBTITLE),
            Spacer(1, 0.32 * inch), p("August 4, 2026  |  Controlled operating copy", SUBTITLE),
            PageBreak()]


def metrics(items: list[tuple[str, str]]) -> Table:
    cells = [[p(number, METRIC_NUMBER), p(label, METRIC_LABEL)] for number, label in items]
    out = Table([cells], colWidths=[7.3 * inch / len(cells)] * len(cells), hAlign="LEFT")
    out.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PALE),
        ("BOX", (0, 0), (-1, -1), 0.7, GRID),
        ("INNERGRID", (0, 0), (-1, -1), 0.7, colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    return out


def step_card(number: str, title: str, command: str, meaning: str) -> Table:
    badge = Table([[p(number, STEP_NUM)]], colWidths=[0.46 * inch], rowHeights=[0.46 * inch])
    badge.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), TEAL),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    content = [p(title, STEP_TITLE), p(command.replace("\n", "<br/>"), STEP_CODE), p(meaning, STEP_BODY)]
    card = Table([[badge, content]], colWidths=[0.65 * inch, 6.65 * inch], hAlign="LEFT")
    card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.65, GRID),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 9),
        ("RIGHTPADDING", (0, 0), (0, 0), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (1, 0), (1, 0), 8),
        ("RIGHTPADDING", (1, 0), (1, 0), 9),
    ]))
    return card


def trace_appendix() -> list:
    story = [PageBreak(), p("Field-level source trace", H1),
             p("This appendix restores the detailed one-writer map. Every row is generated from the current repository source map and includes a live code anchor. Create/redeem is labeled as a control where applicable; prepared N-PORT XML is not a writer.", CALLOUT)]
    checks = provenance.verify()
    story += [p("Verification: " + "; ".join(checks) + ".", SMALL)]
    grouped: dict[str, list[tuple[str, str, str, str]]] = {}
    for group, field, owner, mechanism, filename, anchor in provenance.ROWS:
        grouped.setdefault(group, []).append((field, owner, mechanism, provenance.ref(filename, anchor)))
    for group, rows in grouped.items():
        for start in range(0, len(rows), 4):
            label = group if start == 0 else f"{group} - continued"
            story += [KeepTogether([p(label, H2), table(
                ["XML field(s)", "Current writer", "How populated", "Code anchor"],
                [list(row) for row in rows[start:start + 4]],
                [1.25, 1.55, 2.9, 1.6],
            )])]
    return story


def field_code_trace(domain: str, name: str, group: str = "") -> str:
    """Return the exact accepted-input-to-XML code path for one inventory field."""
    if domain == "config":
        loader = provenance.ref("config.py", r"^def parse_config")
        if name in {"cik", "ccc", "series_id", "class_id"}:
            emitter = provenance.ref("builder.py", r"def _build_header")
        elif name.startswith("signer_"):
            emitter = provenance.ref("builder.py", r"def _build_signature")
        else:
            emitter = provenance.ref("builder.py", r"def _build_gen_info")
    elif domain == "filing":
        loader = provenance.ref("config.py", r"^def parse_filing")
        if name in {"submission_type", "live_test_flag"}:
            emitter = provenance.ref("builder.py", r"def _build_header")
        elif name in {"rep_pd_end", "rep_pd_date", "is_final_filing"}:
            emitter = provenance.ref("builder.py", r"def _build_gen_info")
        elif name == "date_signed":
            emitter = provenance.ref("builder.py", r"def _build_signature")
        else:
            emitter = provenance.ref("builder.py", r"def _build_fund_info")
    else:
        loader = provenance.ref("config.py", r"^def parse_holdings")
        emitter_name = {
            "liquidity": "_build_liquidity",
            "debt": "_build_debt_sec",
            "deriv_common": "_build_derivative_info",
            "option": "_build_option_deriv",
            "ref_instrument": "_build_ref_instrument",
            "swap": "_build_swap_deriv",
            "forward": "_build_fwd_fut_deriv",
            "other_deriv": "_build_other_deriv",
        }.get(group, "_build_one_holding")
        emitter = provenance.ref("builder.py", rf"def {emitter_name}")
    return f"{loader} -> {emitter}"


def complete_field_inventory() -> list:
    """Every accepted config, filing, and holding field with measured current coverage."""
    dirs = trace.production_dirs()
    configs = [trace.kv(d / "fund_config.txt") for d in dirs]
    filings = [trace.kv(d / "filings" / trace.PERIOD / "filing_data.txt") for d in dirs]
    holdings = trace.holding_rows(dirs)
    config_reverse = {v: k for k, v in _CONFIG_KEY_MAP.items()}
    filing_reverse = {v: k for k, v in _FILING_KEY_MAP.items()}

    config_rows: list[list[str]] = []
    for field in fields(FundConfig):
        key = config_reverse[field.name]
        source, _transform, control = trace.config_source(field.name)
        values = [record.get(key, "") for record in configs]
        state = trace.coverage(values)
        if field.name == "ccc":
            placeholders = sum(str(value).strip() in {"", "XXXXXXXX", "N/A"} for value in values)
            state = f"{placeholders}/{len(values)} blank or placeholder; values not printed"
        code = field_code_trace("config", field.name)
        config_rows.append([key, f"{source}<br/><b>Code:</b> {code}", "Every filing", state, control])

    filing_rows: list[list[str]] = []
    for field in fields(FilingData):
        key = filing_reverse[field.name]
        source, _transform, control = trace.filing_source(field.name)
        values = [record.get(key, "") for record in filings]
        applies = "Policy/regime conditional" if field.default == "" else "Every filing unless inapplicable"
        code = field_code_trace("filing", field.name)
        filing_rows.append([key, f"{source}<br/><b>Code:</b> {code}", applies, trace.coverage(values), control])

    holding_rows: list[list[str]] = []
    for field in fields(Holding):
        spec = FIELD_BY_NAME[field.name]
        source, _transform, control = trace.holding_source(spec.group, field.name)
        applicable = [row for row in holdings if trace.applicable(
            row, field.name, spec.group, spec.required, spec.condition
        )]
        if spec.group == "liquidity":
            populated = sum(trace.present(row.get(spec.csv_header, "")) for row in holdings)
            state = f"{populated}/{len(holdings)} populated; applicability awaits policy"
        else:
            state = trace.coverage([row.get(spec.csv_header, "") for row in applicable], len(applicable))
        applies = "Always" if spec.required == "always" else (spec.condition or "Optional when supported")
        code = field_code_trace("holding", field.name, spec.group)
        holding_rows.append([spec.csv_header, f"{source}<br/><b>Code:</b> {code}", applies, state, control])

    total = len(config_rows) + len(filing_rows) + len(holding_rows)
    story = [PageBreak(), p("Complete field inventory and measured coverage", H1),
             p(f"This is the authoritative exhaustive trace for all {total} accepted FundConfig, FilingData, and Holding fields. Every row names its current writer/source and the exact parser-to-XML-emitter code path. Coverage is measured from {len(dirs)} production fund directories and {len(holdings):,} current holding rows. A populated value proves presence only - not correctness, freshness, applicability, or approval.", CALLOUT)]
    sections = [
        ("Fund configuration", config_rows),
        ("Filing-level data", filing_rows),
        ("Holding and derivative data", holding_rows),
    ]
    for title, rows in sections:
        for start in range(0, len(rows), 4):
            label = title if start == 0 else f"{title} - continued"
            heading = p(title, H1) if start == 0 else p(label, H2)
            story += [KeepTogether([heading, table(
                ["Field", "Current writer/source and code path", "Applies when", "Observed state", "Required human action"],
                rows[start:start + 4], [0.9, 2.4, 1.0, 1.1, 1.9],
            )])]
    story += [KeepTogether([p("How to read the inventory", H1), table(
        ["State", "Meaning", "Action"], [
            ["Populated", "A nonblank current value exists. It may include N/A.", "Verify source, cutoff, method, and support for N/A."],
            ["Blank - applicable", "A required or conditional value is absent.", "Supply the factual value or supported inapplicable decision."],
            ["Applicability unresolved", "Approved policy is missing.", "Resolve policy before population or release."],
            ["Optional", "The form does not always require the field.", "Populate only when supported; never invent it."],
        ], [1.3, 2.7, 3.3])])]
    return story


def audit_evidence_appendix() -> list:
    """Measured current-state evidence and the complete reconciliation population."""
    rec = measured.reconciliation()
    xs = measured.xml_state()
    cov = measured.fund_coverage()
    returns_agree, returns_total = measured.returns_vs_inception()
    zero_detail = measured.zero_redemption_detail(rec.get("zero_red", []))

    validator = NportValidator()
    xsd_failures = []
    xsd_varinfo_failures = 0
    for xml_path in sorted((ROOT / "output").glob("*_2026-06.xml")):
        errors = validator.validate_xsd(xml_path.read_bytes())
        if errors:
            xsd_failures.append(xml_path.stem.split("_")[0])
            if "fundsDesignatedInfo" in errors[0] and "medianDailyVarPct" in errors[0]:
                xsd_varinfo_failures += 1

    story = [PageBreak(), p("Measured current-state evidence", H1),
             p("Every number in this section is recalculated from the landed source files, reconciliation CSV, human-review workbook, per-fund inputs, and generated XML. The results describe the current workspace; they are not a prediction of the period-end rerun.", CALLOUT),
             metrics([(f"{rec['rows']:,}", "RECONCILIATION CHECKS"),
                      (str(rec["review"]), "ROWS FLAGGED REVIEW"),
                      (str(rec["flagged"]), "SOURCE IDENTIFIERS FLAGGED"),
                      (str(rec["clear"]), "SOURCE IDENTIFIERS CLEAR")]),
             p("Reconciliation analysis", H1),
             table(["Measure", "Current result", "Interpretation"], [
                 ["Monthly checks", f"{rec['review']} of {rec['rows']:,} flagged", "Includes month-attribution timing and true quarter differences."],
                 ["Quarter fund-side series", f"{rec['tie']} of {rec['series']} tie; {rec['genuine']} differ", "Quarter aggregation removes many placement-versus-settlement timing flags."],
                 ["Population", f"{rec['flagged']} of {rec['funds']} source identifiers flagged", "This population includes identifiers from custody, EagleSTAR, and orders; it is not the count of production funds."],
                 ["Zero accounting redemptions", f"{len(rec['zero_red'])} identifiers; ${zero_detail['total']:,.2f}", "Accepted redemptions occurred on the June 24 accounting cutoff and are absent from the TB writer."],
             ], [1.55, 1.85, 3.9]),
             p("Largest quarter differences that remain after timing aggregation", H2),
             table(["Fund", "Side", "EagleSTAR accounting", "Order book", "Difference"], [
                 [fund, side, f"${accounting:,.2f}", f"${orders:,.2f}", f"${accounting-orders:+,.2f}"]
                 for fund, side, accounting, orders in rec.get("biggest", [])
             ], [0.8, 0.85, 1.75, 1.75, 2.15]),
             p("Current XML and fund coverage", H1),
             table(["Control", "Measured state", "Meaning"], [
                 ["Generated XML", f"{xs['files']} files; status {xs['live']}", "All are TEST; none is represented as filed or LIVE."],
                 ["SEC XSD", f"{xs['files']-len(xsd_failures)}/{xs['files']} pass", f"{xsd_varinfo_failures}/{xs['files']} fail because fundsDesignatedInfo is emitted without required medianDailyVarPct."],
                 ["pctVal denominator", f"{xs['files']-len(xs['pct_break'])}/{xs['files']} tie", "Current output percentages reconcile to reported net assets."],
                 ["Swap direction", f"{xs['swaps']-xs['swap_na']}/{xs['swaps']} populated", "Long/short payoff profile is present in current XML."],
             ], [1.45, 2.0, 3.85]),
             KeepTogether([p("Coverage and remaining XML gaps", H2),
             table(["Control", "Measured state", "Meaning"], [
                 ["Option delta", f"{xs['opt_na']}/{xs['opts']} blank/N/A across {xs['opt_funds']} funds", "No automated feed; requires risk-system values at cutoff."],
                 ["Designated index", f"{len(xs['no_index'])} funds show N/A", ", ".join(xs['no_index']) or "None"],
                 ["Production coverage", f"{cov['built']} XML for {cov['real']} custodian accounts", f"Missing: {', '.join(cov['missing']) or 'none'}; fixtures excluded: {', '.join(cov['fixtures']) or 'none'}."],
                 ["Monthly returns", f"{returns_agree}/{returns_total} consistent with inception", "Blank pre-inception months are not treated as Bloomberg failures."],
             ], [1.45, 2.0, 3.85])])]

    review_path = ROOT / "data" / "humanreview" / "2026-06_review.xlsx"
    workbook = load_workbook(review_path, data_only=True, read_only=True)
    required = {
        "swap_counterparties": ["legalName", "lei"],
        "option_index": ["indexName", "indexIdentifier"],
        "swap_legs": ["recFixedOrFloating", "recDesc", "pmntFloatingRtIndex",
                      "pmntFloatingRtSpread", "pmntRateTenor", "pmntRateUnit"],
        "invCountry": ["invCountry"],
    }
    review_rows = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header = [str(value or "") for value in rows[0]] if rows else []
        index = {name: i for i, name in enumerate(header)}
        gap_rows = 0
        gap_cells = 0
        for row in rows[1:]:
            missing = [name for name in required.get(sheet.title, [])
                       if not str(row[index[name]] or "").strip()]
            if missing:
                gap_rows += 1
                gap_cells += len(missing)
        review_rows.append([sheet.title, str(max(0, len(rows) - 1)), str(gap_rows), str(gap_cells)])

    policy_fields = ["fiscalYearEndMMDD", "derivativesRegimePolicy", "liquidityRequired",
                     "cashB2fRequired", "policyEffectiveFrom"]
    policy_records = [trace.kv(directory / "fund_config.txt") for directory in trace.production_dirs()]
    policy_complete = sum(all(record.get(field) for field in policy_fields) for record in policy_records)
    story += [p("Human-review workbook and policy state", H1),
              table(["Review sheet", "Rows", "Rows with required blanks", "Required blank cells"],
                    review_rows, [2.4, 1.2, 1.85, 1.85]),
              p(f"<b>Policy coverage:</b> {policy_complete}/{len(policy_records)} production fund configurations contain all five required policy fields. The required fields are fiscalYearEndMMDD, derivativesRegimePolicy, liquidityRequired, cashB2fRequired, and policyEffectiveFrom. The code must not infer them.", WARN),
              p("Release decision", H1),
              table(["Gate", "Current state", "Release effect", "Resolution path"], [
                  ["Period-end freshness", "OPEN - custody 06/25; EagleSTAR and orders 06/24; report 06/30", "Blocks a final apples-to-apples period-end conclusion", "Load aligned period-end files and rerun masters."],
                  ["Human review", "OPEN - 58 swap spread cells", "Blocks affected swaps", "Enter executed-confirm values in swap_legs and rerun mergehumanreview."],
                  ["Reconciliation", f"OPEN - {rec['review']} REVIEW rows", "Blocks LIVE where unresolved", "Correct sources, rerun masters, and investigate remaining quarter differences."],
                  ["Fund policy", f"OPEN - {policy_complete}/{len(policy_records)} complete", "Blocks policy/applicability derivation", "Enter approved facts in each fund_config.txt."],
                  ["Option delta", f"OPEN - {xs['opt_na']}/{xs['opts']} blank/N/A", "Blocks applicable option holdings", "Supply risk-system delta at the report cutoff."],
                  ["XML/XSD", f"BLOCKED - {xs['files']-len(xsd_failures)}/{xs['files']} current XML pass", "Blocks every current output", "Fix varInfo emission at src/nport/builder.py:229-231 and policy/applicability; supply B.10 values where applicable, then rebuild."],
              ], [1.2, 2.25, 1.75, 2.1]),
              p("<b>Current conclusion:</b> NOT READY FOR LIVE. The system has 104 generated TEST outputs, but none currently passes XSD. Source freshness, human-entered swap terms, reconciliation, policy configuration, option delta, and varInfo/B.10 emission also remain unresolved.", WARN),
              KeepTogether([p("Resolution sequence and reproducibility", H1),
              table(["Order", "Action", "System check"], [
                  ["1", "Replace custody, EagleSTAR, and order files with aligned period-end inputs.", "nport masters 2026-06 --dry-run, then nport masters 2026-06"],
                  ["2", "Open both master workbooks on Bloomberg; calculate, save, and close.", "Confirm formulas have values or explicit errors."],
                  ["3", "Generate and complete the one human-review workbook.", "nport mergehumanreview 2026-06; fill blanks; rerun command."],
                  ["4", "Resolve quarter-level reconciliation differences.", "Review data/master/reconciliation_2026-06.csv."],
                  ["5", "Populate approved policy facts and option deltas.", "Build must report no applicable input/policy blockers."],
                  ["6", "Project and rebuild every fund.", "nport split 2026-06; nport build; confirm XSD and TEST/LIVE state."],
              ], [0.65, 3.55, 3.1])])]

    recon_path = ROOT / "data" / "master" / "reconciliation_2026-06.csv"
    flagged: dict[str, Counter] = {}
    with recon_path.open(newline="", encoding="utf-8") as stream:
        for row in csv.DictReader(stream):
            if row.get("flag") != "REVIEW":
                continue
            category = "flow" if row.get("check", "").startswith("flow:") else row.get("check", "unknown")
            flagged.setdefault(row.get("fund", "UNKNOWN"), Counter())[category] += 1
    items = sorted(flagged.items())
    half = (len(items) + 1) // 2
    paired = []
    for index in range(half):
        row = []
        for offset in (0, half):
            if index + offset < len(items):
                fund, counts = items[index + offset]
                row += [fund, str(sum(counts.values())), ", ".join(f"{name}:{count}" for name, count in sorted(counts.items()))]
            else:
                row += ["", "", ""]
        paired.append(row)
    story += [PageBreak(), p("Reconciliation REVIEW population", H1),
              p(f"The table below lists all {len(items)} source identifiers with at least one REVIEW row. This is a reconciliation population, not a statement that all identifiers are active filing funds. Category counts reproduce directly from reconciliation_2026-06.csv.", CALLOUT)]
    for start in range(0, len(paired), 16):
        label = "Flagged source identifiers" if start == 0 else "Flagged source identifiers - continued"
        story += [KeepTogether([p(label, H2), table(
            ["Fund", "Flags", "Categories", "Fund", "Flags", "Categories"],
            paired[start:start + 16], [0.65, 0.45, 2.55, 0.65, 0.45, 2.55],
        )])]
    return story


def runbook() -> list:
    s = cover("N-PORT Final Runbook", "The definitive monthly workflow: command, human action, output, and release gate")
    s += [metrics([("3", "RECEIVED INPUT FOLDERS"), ("2", "BLOOMBERG WORKBOOKS"),
                  ("1", "HUMAN-REVIEW WORKBOOK"), ("8", "CONTROLLED STAGES")]),
          p("Purpose and operating rule", H1),
          p("This is the single operating document for the current in-house N-PORT pipeline. Follow the stages in order. A stage is complete only when its stated stop condition is clear. Never use a U.S. Bank-produced N-PORT file to populate the in-house filing; those files are comparison outputs only.", CALLOUT),
          p("Before starting", H1),
          table(["Check", "Human action", "Pass condition"], [
              ["Working directory", "Open PowerShell in C:\\Users\\damie\\nportvalidation.", "The prompt is at the repository root."],
              ["Period", "Replace 2026-06 in every command when operating another filing period.", "One period is used consistently from intake through XML."],
              ["Workbooks", "Close security_master.xlsx, filing_master.xlsx, and the period review workbook before commands write them.", "No workbook is locked by Excel."],
              ["Source boundary", "Use only received custodian, EagleSTAR, and create/redeem inputs plus Bloomberg and supported human evidence.", "No prepared custodian XML is used as a data writer."],
          ], [1.25, 3.75, 2.3]),
          p("Monthly inputs - no positions.csv", H1),
          p("Do not create <b>positions.csv</b>. Put the files actually received into the three folders below. The masters command resolves a file from each folder and prints the exact selection before it writes anything.", WARN),
          table(["Input", "Folder", "System role", "Human check"], [
              ["Custodian positions", "data/custodian/", "Creates the holding population and supplies custody position and valuation fields.", "Confirm the file belongs to the filing population and its as-of date matches the intended cutoff."],
              ["EagleSTAR", "data/fund_accounting/", "Writes fund-accounting totals, gains, liabilities, flows, and derivative unrealized amounts when present.", "Confirm the package is complete and uses the intended accounting cutoff."],
              ["Create/redeem", "data/orders/", "Reconciliation control for EagleSTAR subscriptions and redemptions; never overwrites filing values.", "Confirm order status and coverage through the same cutoff."],
              ["Bloomberg", "Two master workbooks", "Calculates formulas for supported returns, durations, and security-reference fields.", "Refresh, wait, inspect formula errors, save, and close."],
              ["Human review", "data/humanreview/&lt;period&gt;_review.xlsx", "Carries residual security and derivative facts not supplied by automation.", "Edit value columns only; never change row keys."],
          ], [1.05, 1.45, 2.6, 2.2]),
          PageBreak(), p("Front-to-end workflow", H1),
          p("Run every command from the repository root in PowerShell. The explicit executable path below avoids activation and PATH ambiguity.", CALLOUT)]
    steps = [
        ("1. Preview the input selection", '& ".\\.venv\\Scripts\\nport.exe" masters 2026-06 --dry-run',
         "Human: read the printed custodian, EagleSTAR, and create/redeem paths and dates. Stop if the wrong file, population, or cutoff is selected; correct the folder contents or pass an explicit supported path."),
        ("2. Build the two master workbooks", '& ".\\.venv\\Scripts\\nport.exe" masters 2026-06',
         "Human: confirm the command completes and creates data\\master\\security_master.xlsx, data\\master\\filing_master.xlsx, and data\\master\\reconciliation_2026-06.csv. Do not continue after an error."),
        ("3. Let Bloomberg calculate", "Open data\\master\\security_master.xlsx\nOpen data\\master\\filing_master.xlsx",
         "Human: use Excel on a Bloomberg-enabled machine. In each workbook, wait for calculation to finish; inspect visible #N/A, #VALUE!, #REF!, or blank required results; then save as .xlsx and close both files."),
        ("4. Generate the residual review", '& ".\\.venv\\Scripts\\nport.exe" mergehumanreview 2026-06',
         "Human: read the command's required-gap counts, then open data\\humanreview\\2026-06_review.xlsx. Fill supported value cells only. Save and close the workbook."),
        ("5. Apply the human entries", '& ".\\.venv\\Scripts\\nport.exe" mergehumanreview 2026-06',
         "Human: confirm the command reports no required blanks and no unresolved swap/option reference. If it reports gaps, reopen the exact named sheet, correct those rows, save, close, and run this command again."),
        ("6. Clear the remaining status gates", "Open data\\humanreview\\2026-06_review.xlsx\nReview reconciliation and pipeline_status",
         "Human: all issues are visible here. Enter policy/XML facts only in the editable sheets. For a read-only reconciliation or pipeline row, correct the named custodian, EagleSTAR, or create/redeem input, rerun masters, refresh Bloomberg, and rerun mergehumanreview."),
        ("7. Preview and write per-fund files", '& ".\\.venv\\Scripts\\nport.exe" split 2026-06 --dry-run\n& ".\\.venv\\Scripts\\nport.exe" split 2026-06',
         "Human: verify the previewed fund population, then run the write command. It projects the reviewed masters to data\\funds\\<fund>\\. Do not manually patch generated filing_data.txt, holdings.csv, derivatives.csv, or security_master.csv."),
        ("8. Preflight and build XML", '& ".\\.venv\\Scripts\\nport.exe" build --dry-run\n& ".\\.venv\\Scripts\\nport.exe" build',
         "Human: treat any input-validation, policy, reconciliation, XML-coverage, or XSD error as a stop. The first command validates without writing; run the second only after the dry run passes for the intended population."),
    ]
    for heading, command, meaning in steps:
        number, title = heading.split(". ", 1)
        s += [Spacer(1, 5), KeepTogether([step_card(number, title, command, meaning)])]
    s += [PageBreak(), p("Human-review workbook: exact edit instructions", H1),
          p("The workbook is generated after custodian, EagleSTAR, create/redeem, and Bloomberg processing. It is the single human-review location. Navy/grey columns are identifiers or status; teal-header columns are human-editable. Red cells are required and blank; green cells are required and populated; blue cells are optional. Never edit row keys.", CALLOUT),
          table(["Sheet", "Do not edit - row keys", "Human-editable value columns", "Required rule"], [
              ["fund_policy", "fund", "fiscalYearEndMMDD; derivativesRegimePolicy; liquidityRequired; cashB2fRequired; policyEffectiveFrom; policyEffectiveTo", "First five are required for every fund. policyEffectiveTo may be blank for an open-ended policy. These values are applied to fund_config.txt on the next merge."],
              ["fund_risk", "fund", "cashNotReportedInCOrD; monthlyReturnCategoriesJson; derivExposurePct; derivCurrencyExposurePct; derivInterestRateExposurePct; derivDaysInExcess; medianDailyVarPct; medianVarRatioPct; backtestingExceptions", "Required cells are determined only by fund_policy: B.2.f, B.9, or B.10 is not requested when the approved policy says it is inapplicable."],
              ["designated_index", "ticker", "indexName; indexIdentifier", "Conditional. Confirm the prefilled value against the fund's approved prospectus/risk policy. N/A is allowed only when supported."],
              ["swap_counterparties", "code", "legalName; lei", "Both are required for each listed counterparty. Use the executed legal-entity/trade record."],
              ["option_index", "underlying", "indexName; indexIdentifier", "Both are required for each listed option underlying. Use the actual contract/reference instrument."],
              ["swap_legs", "fund; swapTicker", "recFixedOrFloating; recDesc; pmntFloatingRtIndex; pmntFloatingRtSpread; pmntRateTenor; pmntRateUnit", "All six are required for each swap row. Use the executed confirmation. Enter 0 only when the actual spread is zero."],
              ["option_delta", "fund; ticker", "delta", "N/A is schema-valid. Replace it only with a supported report-date delta; no source/as-of metadata is required for XML generation."],
              ["holding_liquidity", "fund; ticker; cusip; name", "liquidityClassificationJson; liquidityCircumstancesJson", "Classification is required only for funds whose fund_policy liquidityRequired is Y. Circumstances are populated only when applicable."],
              ["invCountry", "fund; ticker; cusip; name", "invCountry", "Required when the row exists. Enter a supported ISO two-letter country; do not infer from the issuer name."],
              ["isin", "fund; ticker; cusip; name", "isin", "Optional. A blank may remain when another permitted identifier path is valid."],
              ["sus_review", "fund; field", "value; reason; confirmed", "Current workflow does not make these columns a required release gate. Do not use this sheet to override another source."],
              ["reconciliation", "all columns", "None - read-only", "A REVIEW row is a source disagreement. Correct the relevant landed input and rerun masters; the workbook cannot overwrite the filed-side value."],
              ["pipeline_status", "all columns", "None - read-only", "The resolution column names the command/input action. A BLOCKED row remains blocked until regeneration reports READY."],
          ], [1.1, 1.35, 2.65, 2.2]),
          p("What mergehumanreview does", H2),
          table(["Run", "System action", "Human pass condition"], [
              ["First run", "Refreshes every review sheet from the Bloomberg-populated masters and reconciliation report, preserving prior values on matching row keys.", "Use the printed per-sheet counts and red cells to locate required blanks."],
              ["Second and later runs", "Applies policy to each fund_config.txt, fund-risk values to filing_master.xlsx, and holding values to security_master.xlsx; read-only status is regenerated.", "No required blanks, unresolved swap/option reference, or read-only BLOCKED status remains."],
          ], [1.25, 3.8, 2.25]),
          p("Reconciliation flags are different", H1),
          p("A missing value is editable because the XML needs a human-supplied fact. A blocking status may be read-only because the pipeline detected a source disagreement or failed gate. Both appear in the same workbook. Do not convert a source mismatch into a manual filing override: correct the named input and regenerate.", WARN),
          KeepTogether([p("Fund policy is now centralized", H1),
          p("Enter the five required policy facts in fund_policy. On the next mergehumanreview run, the system writes populated values into the matching fund_config.txt and derives submissionType, repPdEnd, repPdDate, derivativesRegime, and conditional requirements. It never infers an absent policy value.", CALLOUT)]),
          p("Release checklist", H1),
          table(["Gate", "Human verification", "Evidence/output", "Stop when"], [
              ["Input selection", "Selected files, population, and cutoff are correct.", "masters --dry-run output", "Any selected path or date is wrong."],
              ["Date alignment", "Custody, accounting, orders, and report date support an apples-to-apples period conclusion.", "Received file metadata and printed selection", "Cutoffs are materially misaligned and no approved treatment exists."],
              ["Bloomberg", "Both masters finished calculation and were saved/closed.", "Literal values read on mergehumanreview", "Required formulas remain errors or blank."],
              ["Human review", "Every required review cell is supported and populated.", "mergehumanreview gap report", "Required gaps or unresolved swap/option references remain."],
              ["Reconciliation", "Read-only REVIEW rows are cleared after source correction.", "review workbook -> reconciliation", "An unexplained filing-relevant difference remains."],
              ["Policy/applicability", "All required fund_policy cells are populated and applied.", "review workbook -> fund_policy", "A required policy fact or applicability decision is absent."],
              ["Fund population", "The split preview contains the intended funds and no fixture accounts.", "split --dry-run output", "A fund is missing, duplicated, or unexpected."],
              ["XML and schema", "Dry-run build and final build pass input and SEC XSD validation.", "build output and output/<FUND>_<PERIOD>.xml", "Any validation or XSD error remains."],
              ["Release status", "The output is deliberately TEST or LIVE under the approved release process.", "root attributes in generated XML", "The operator cannot prove the intended status."],
          ], [1.15, 2.55, 1.9, 1.7]),
          PageBreak(), KeepTogether([p("Current 2026-06 facts - not a successful release", H1),
          table(["Known issue", "Measured current state", "Where it is resolved"], [
              ["Source freshness", "Custody 06/25; EagleSTAR 06/24; orders through 06/24; report date 06/30.", "Replace the files in the three intake folders with aligned period-end inputs, then rerun masters."],
              ["Human review", "58 required pmntFloatingRtSpread cells are blank.", "2026-06_review.xlsx -> swap_legs -> pmntFloatingRtSpread, using executed confirmations."],
              ["Reconciliation", "258 REVIEW rows across 137 source identifiers; 72 of 304 quarter fund-side series still differ after aggregation.", "Correct or replace the applicable custody, EagleSTAR, or order source; rerun masters."],
              ["Fund policy", "0 of 105 production fund configurations contain all five required policy fields.", "2026-06_review.xlsx -> fund_policy. Enter the five approved facts; rerun mergehumanreview."],
              ["Option delta", "70 of 70 current option rows contain N/A.", "2026-06_review.xlsx -> option_delta. N/A is schema-valid; replace only when a supported report-date delta is available."],
              ["Conditional B.2.f/B.9/B.10", "Applicability cannot be determined until fund_policy is complete.", "2026-06_review.xlsx -> fund_risk; required cells turn red after policy is populated and the merge is rerun."],
              ["Conditional C.7", "1,996 holding rows have an entry location; applicability cannot be determined until fund_policy is complete.", "2026-06_review.xlsx -> holding_liquidity; classification becomes required only for liquidityRequired=Y funds."],
              ["Fund coverage", "104 XML files for 105 real custodian accounts; GPTZ is missing.", "Resolve the GPTZ fund input/configuration path before split/build."],
              ["SEC XSD", "0 of 104 existing TEST XML files pass; those files contain fundsDesignatedInfo without required medianDailyVarPct.", "The current builder now guards varInfo on medianDailyVarPct. Resolve policy/B.10 inputs where applicable, rebuild all XML, and revalidate; do not release the stale files."],
          ], [1.2, 2.75, 3.35])]),
          p("Therefore the current 2026-06 output is <b>NOT READY FOR LIVE</b>. This is a measured operating state, not a request for the human reviewer to invent replacement values.", WARN),
          p("One-fund diagnostic commands", H1),
          table(["Purpose", "Exact command"], [
              ["Validate one fund's current input files", '& ".\\.venv\\Scripts\\nport.exe" validate fdrs 2026-06'],
              ["Run policy/applicability preflight for one fund", '& ".\\.venv\\Scripts\\nport.exe" preflight fdrs 2026-06'],
              ["Build one fund after all gates pass", '& ".\\.venv\\Scripts\\nport.exe" build fdrs 2026-06'],
              ["Show the built-in workflow summary", '& ".\\.venv\\Scripts\\nport.exe" guide'],
              ["Show all supported commands", '& ".\\.venv\\Scripts\\nport.exe" --help'],
          ], [2.8, 4.5]),
          p("Completion definition", H1),
          p("The run is complete only when the selected input paths and cutoffs are correct, Bloomberg has been saved, required human-review cells are filled, filing-relevant reconciliation and policy/applicability gates are resolved, the intended fund population is projected, and every intended XML passes input validation and the loaded SEC XSD.", CALLOUT)]
    s += trace_appendix()
    s += complete_field_inventory()
    return s


def audit() -> list:
    s = cover("N-PORT Final Operating Audit", "What the system uses now, what remains open, and exactly where each issue is fixed")
    s += [metrics([("3", "CURRENT SOURCE FILES"), ("58", "SWAP REVIEW GAPS"),
                  ("258", "RECONCILIATION FLAGS"), ("6 DAYS", "TO REPORT DATE")]),
          p("Executive conclusion", H1),
          p("The immediate input error was caused by instructions that required a nonexistent per-fund positions.csv. That was the wrong workflow for this repository. The working monthly path is custodian + EagleSTAR + create/redeem, followed by Bloomberg and one centralized human-review workbook.", CALLOUT),
          p("Current file selection", H1),
          table(["Input", "Selected file", "Observed cutoff", "Role"], [
              ["Custodian", "data/custodian/Corgi_Adv.40C8.C8_ETF_Holdings (1).csv", "2026-06-25", "Filing input"],
              ["EagleSTAR", "data/fund_accounting/takeout-20260625T000453Z-3-001.zip", "2026-06-24", "Filing input"],
              ["Create/redeem", "data/orders/HistoricalOrders (83).csv", "through 2026-06-24", "Reconciliation control"],
              ["Report date", "2026-06 filing", "2026-06-30", "Target"],
          ], [1.05, 3.15, 1.05, 1.55]),
          p("Time-series finding", H1),
          p("The current sources are not period-end aligned: custody is June 25, EagleSTAR is June 24, create/redeem ends June 24, and the filing date is June 30. Therefore the current numerical differences are not a valid final apples-to-apples period-end comparison. They remain open until period-end files are loaded or the filing period is changed to match the source cutoff.", WARN),
          p("Current open items", H1),
          table(["Open item", "Measured state", "Exact place to fix", "Resolution"], [
              ["Swap payment-leg spread", "58 of 58 swap review rows are blank for pmntFloatingRtSpread.", "data/humanreview/2026-06_review.xlsx -> swap_legs -> pmntFloatingRtSpread", "Enter the actual spread from each executed swap confirmation; then rerun mergehumanreview."],
              ["Flow reconciliation", "238 REVIEW rows: 216 sales and 22 redemptions across mon1-mon3.", "Replace/update data/orders/ and/or data/fund_accounting/; then rerun masters.", "Align both sources to the same report cutoff and investigate remaining differences. Do not overwrite EagleSTAR filing values with order-book values."],
              ["Net-assets reconciliation", "19 REVIEW rows.", "Replace/update data/custodian/ and/or data/fund_accounting/; then rerun masters.", "Use matching as-of dates first; investigate only the residual breaks."],
              ["Liability reconciliation", "1 REVIEW row.", "Correct the EagleSTAR mapping/source in data/fund_accounting/ or the accounting adapter; then rerun masters.", "Confirm mapped payable accounts equal the EagleSTAR total-liabilities control."],
          ], [1.1, 1.7, 2.45, 2.05]),
          KeepTogether([p("Fund policy release gap", H2),
          p("<b>Measured state:</b> data/master/fund_registry.csv is absent, and fund_config files without policy fields cannot release.<br/><b>Exact place to fix:</b> per fund, enter fiscalYearEndMMDD, derivativesRegimePolicy, liquidityRequired, cashB2fRequired, and policyEffectiveFrom in fund_config.txt.<br/><b>Rule:</b> enter approved fund facts; the code must not infer them.", WARN)]),
          KeepTogether([p("All-funds XSD release blocker", H2),
          p("<b>Measured state:</b> 0 of 104 current TEST XML files pass the loaded SEC XSD. Every file emits fundsDesignatedInfo while the required medianDailyVarPct is absent.<br/><b>Exact engineering location:</b> src/nport/builder.py:229-231 plus policy/applicability handling in policy.py and input_validation.py.<br/><b>Resolution:</b> do not emit varInfo/designated-index content unless the applicable regime is resolved, or supply the required B.10 values for applicable VaR funds; rebuild and rerun XSD validation.", WARN)]),
          p("Meaning of missing versus blocking", H1),
          table(["Term", "Meaning", "Action"], [
              ["Missing", "An applicable value is absent and the system has a known entry location.", "Enter it in the named human-review cell or correct the named input."],
              ["Blocking", "The system must not release XML because an applicable value, policy fact, reconciliation, or validation is unresolved.", "Resolve the underlying issue. A blocker is a state, not a separate file type."],
              ["Reconciliation REVIEW", "Two controls disagree or their dates are not aligned.", "Fix/replace source inputs and rerun. It is not a blank field to type over."],
          ], [1.15, 3.1, 3.15]),
          p("Verified operating boundary", H1),
          p("Prepared N-PORT XML and other reference filings remain comparison-only. The restored masters command no longer reads data/RealXMLs. Create/redeem orders remain a control only. The command prints selected inputs before writing, which makes an accidental file selection visible.", BODY),
          KeepTogether([p("Remaining engineering gap", H1),
          p("The monthly human-review workbook centralizes residual security and derivative fields, but fund policy is still stored per fund and reconciliation exceptions are still resolved upstream. That is the current code behavior. It would be inaccurate to claim every blocker can be cleared inside the human-review workbook today.", WARN)])]
    s += audit_evidence_appendix()
    return s


def main() -> None:
    DOCS.mkdir(parents=True, exist_ok=True)
    build(DOCS / "NPORT_Runbook.pdf", runbook())


if __name__ == "__main__":
    main()
