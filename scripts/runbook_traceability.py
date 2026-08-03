"""Insert measured, field-level traceability tables into the N-PORT runbook."""

from __future__ import annotations

import csv
import html
from dataclasses import fields
from datetime import datetime
from pathlib import Path

import field_provenance as fp
from openpyxl import load_workbook
from nport.config import _CONFIG_KEY_MAP, _FILING_KEY_MAP, _HOLDINGS_KEY_MAP
from nport import eaglestar
from nport.ap_orders import parse_ap_orders
from nport.custodian import parse_custodian_csv
from nport.models import FilingData, FundConfig, Holding
from nport.schema import FIELD_BY_NAME


ROOT = Path(__file__).resolve().parents[1]
FUNDS = ROOT / "data" / "funds"
RUNBOOK = ROOT / "docs" / "nport_runbook.html"
PERIOD = "2026-06"
FIXTURES = {"bond_fund", "buffered_etf", "leveraged_etf"}


def current_lineage_section() -> str:
    """The old-runbook provenance view, regenerated from current code and files."""
    checks = fp.verify()
    custodian_path = sorted((ROOT / "data" / "custodian").glob("*.csv"))[0]
    custodian_rows = parse_custodian_csv(custodian_path)
    custodian_dates = sorted({r.date for r in custodian_rows})

    export = eaglestar.resolve_export(ROOT / "data" / "fund_accounting")
    cache = ROOT / "data" / "fund_accounting" / ".cache"
    accounting_dates = {
        kind: eaglestar._dates(cache, kind) for kind in ("pval", "tb", "nav")
    }

    order_path = sorted((ROOT / "data" / "orders").glob("*.csv"))[0]
    orders = parse_ap_orders(order_path)
    order_dates = [datetime.strptime(o.trade_date, "%m/%d/%Y").date() for o in orders]
    review_path = ROOT / "data" / "humanreview" / f"{PERIOD}_review.xlsx"
    review_book = load_workbook(review_path, data_only=True)
    review_counts = {ws.title: max(0, ws.max_row - 1) for ws in review_book.worksheets}

    facts = [
        ["U.S. Bank custodian holdings CSV", str(custodian_path.relative_to(ROOT)).replace("\\", "/"),
         f"{len(custodian_rows):,} rows; {len({r.account for r in custodian_rows})} accounts; source Date {', '.join(custodian_dates)}",
         "Active writer for the position population and custody fields"],
        ["Gmail/Google Takeout - U.S. Bank EagleSTAR attachments", str(export.relative_to(ROOT)).replace("\\", "/") if export else "Not found",
         ("; ".join(f"{kind.upper()} {len(ds)} snapshots, latest {ds[-1]}" for kind, ds in accounting_dates.items() if ds)
          or "No decoded snapshots"),
         "Active writer for accounting fields and derivative mark-to-market"],
        ["Create/redeem order export", str(order_path.relative_to(ROOT)).replace("\\", "/"),
         f"{len(orders):,} orders; {sum(o.status == 'ACCEPTED' for o in orders):,} accepted; latest trade date {max(order_dates).isoformat()}",
         "Reconciliation cross-check only in the current CLI; it does not write filing flows"],
        ["Bloomberg workbook formulas", "security_master.xlsx and filing_master.xlsx", "Calculated only after the workbooks are opened and saved on a Bloomberg terminal",
         "Active enrichment writer for security IDs, country, debt terms, returns, and B.3 inputs"],
        ["Human-review workbook", str(review_path.relative_to(ROOT)).replace("\\", "/"),
         "; ".join(f"{name} {count}" for name, count in review_counts.items()),
         "Current rows are seeds/prefills or preserved edits; the workbook has sourceNote but no required reviewer identity, review time, or evidence-reference columns"],
        ["Fund configuration and tool logic", "data/funds/*/fund_config.txt plus source code", "Static filing identity, constants, and derived values",
         "Active writer for the fields identified below"],
        ["U.S. Bank prepared N-PORT files", "Downloaded comparison packages", "No writer path into masters, per-fund inputs, or XML",
         "Comparison only - distinct from the U.S. Bank custody holdings CSV"],
    ]

    fact_html = "".join(row(item) for item in facts)
    provenance_rows: list[str] = []
    last = None
    for group, field, owner, mechanism, filename, anchor in fp.ROWS:
        if group != last:
            provenance_rows.append(f'<tr class="grp"><td colspan="4">{html.escape(group)}</td></tr>')
            last = group
        provenance_rows.append(row([field, owner, mechanism, fp.ref(filename, anchor)]))

    return (
        '<section><h2>10. Current data lineage - verified from code</h2>'
        '<div class="box bad"><b>Correction to the prior audit:</b> the U.S. Bank custody holdings CSV is a current production input. '
        'Only the prepared U.S. Bank N-PORT filing packages are comparison-only. These are different artifacts and different code paths.</div>'
        '<p>The table below reports the files and roles the current code actually uses. It does not describe the future-state source architecture.</p>'
        '<table><tr><th>Current source</th><th>Landed artifact</th><th>Observed coverage/date</th><th>Actual role</th></tr>'
        + fact_html + '</table>'
        '<div class="box warn"><b>Freshness fact:</b> the filing report date is 2026-06-30. The landed custodian file reports 2026-06-25, '
        'and the latest EagleSTAR and create/redeem records are 2026-06-24. All three need period-end replacement or a formally supported '
        'cutoff treatment before release.</div>'
        '<h3>One-writer field map</h3><p>This map is generated from <code>scripts/field_provenance.py</code>. '
        'Every row has a live code anchor. A cross-check is labeled as a cross-check and is not presented as a writer.</p>'
        '<table class="prov"><tr><th>Field(s)</th><th>Current writer/source</th><th>How the value is obtained</th><th>Code reference</th></tr>'
        + ''.join(provenance_rows) + '</table>'
        '<div class="box"><b>Verification:</b> ' + html.escape('; '.join(checks)) + '.</div></section>'
    )


def present(value: object) -> bool:
    # Presence is deliberately separate from correctness/applicability. "N/A" is
    # a supplied value for several Form N-PORT fields and still needs human review.
    return str(value or "").strip() != ""


def kv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if "=" in line and not line.lstrip().startswith("#"):
            key, value = line.split("=", 1)
            out[key.strip()] = value.strip()
    return out


def production_dirs() -> list[Path]:
    return sorted(p for p in FUNDS.iterdir() if p.is_dir() and p.name not in FIXTURES)


def holding_rows(dirs: list[Path]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for fund in dirs:
        filing_dir = fund / "filings" / PERIOD
        path = filing_dir / "holdings.csv"
        if path.exists():
            with path.open(encoding="utf-8-sig", newline="") as fh:
                fund_rows = list(csv.DictReader(fh))
            supplements: dict[str, dict[str, str]] = {}
            for supplemental_name in ("debt_securities.csv", "derivatives.csv"):
                supplemental_path = filing_dir / supplemental_name
                if supplemental_path.exists():
                    with supplemental_path.open(encoding="utf-8-sig", newline="") as fh:
                        for supplemental in csv.DictReader(fh):
                            supplements.setdefault(supplemental.get("holdingId", ""), {}).update(supplemental)
            for record in fund_rows:
                record.update({k: v for k, v in supplements.get(record.get("holdingId", ""), {}).items() if k != "holdingId"})
            rows.extend(fund_rows)
    return rows


def config_source(name: str) -> tuple[str, str, str]:
    if name == "ccc":
        return (
            "fund_config.txt (current); EDGAR filer administrator",
            "Injected only at execution; never copied from comparison files",
            "Filer administrator verifies the credential is active; Operations confirms no plaintext copy enters the release bundle",
        )
    if name.startswith("signer_"):
        return (
            "fund_config.txt (current static value)",
            "Direct mapping; evidence retained with filing approval",
            "Legal/Compliance confirms signer authority for the report date",
        )
    if name.startswith("series_") or name == "class_id":
        return (
            "fund_config.txt (current static value)",
            "Direct identifier mapping; stable-identity reconciliation",
            "Fund Administration verifies identifiers against approved internal records",
        )
    return (
        "fund_config.txt (current static value)",
        "Direct mapping; stable-identity reconciliation",
        "Fund Administration verifies the value and Compliance approves unresolved identity changes",
    )


def filing_source(name: str) -> tuple[str, str, str]:
    policy = {"submission_type", "rep_pd_end", "rep_pd_date", "derivatives_regime"}
    operator = {"live_test_flag", "is_final_filing", "date_signed"}
    totals = {"tot_assets", "tot_liabs", "net_assets", "assets_attr_misc_sec", "assets_invested"}
    accounting = {
        "amt_pay_one_yr_banks_borr", "amt_pay_one_yr_ctrld_comp", "amt_pay_one_yr_oth_affil",
        "amt_pay_one_yr_other", "amt_pay_aft_one_yr_banks_borr", "amt_pay_aft_one_yr_ctrld_comp",
        "amt_pay_aft_one_yr_oth_affil", "amt_pay_aft_one_yr_other", "delay_deliv",
        "stand_by_commit", "liquid_pref", "is_non_cash_collateral", "net_realized_gain_mon1",
        "net_unrealized_appr_mon1", "net_realized_gain_mon2", "net_unrealized_appr_mon2",
        "net_realized_gain_mon3", "net_unrealized_appr_mon3", "cash_not_reported_in_c_or_d",
        "monthly_return_categories_json",
    }
    flows = {n for n in _FILING_KEY_MAP.values() if n.startswith("mon")}
    returns = {"rtn1", "rtn2", "rtn3"}
    risk = {
        "cur_metrics_json", "credit_sprd_risk_ig_json", "credit_sprd_risk_nonig_json",
        "deriv_exposure_pct", "deriv_currency_exposure_pct", "deriv_interest_rate_exposure_pct",
        "deriv_days_in_excess", "median_daily_var_pct", "median_var_ratio_pct", "backtesting_exceptions",
    }
    designated = {"name_designated_index", "index_identifier"}
    if name in policy:
        if name == "derivatives_regime":
            return ("No current populated source; future approved fund registry", "policy.py derives it only when the registry exists", "Compliance/Risk must approve the registry row; current production registry is absent")
        return ("Tool-derived filing period/default; future approved fund registry", "filing_master.py derives the period; policy.py overrides from an approved registry when present", "Operations verifies period; Compliance approves the registry row before LIVE")
    if name in operator:
        return ("Filing Operations and authorized signer workflow", "Operator-controlled status/date with release gate", "Authorized reviewer confirms final status, signature date, and TEST-to-LIVE promotion")
    if name in totals:
        return ("Gmail export - U.S. Bank EagleSTAR Trial Balance; custodian fallback", "eaglestar.py overrides filing_master.py custody-derived fallback", "Fund Accounting verifies attachment cutoff, entity mapping, totals, and reconciliation")
    if name in accounting:
        if name in {"cash_not_reported_in_c_or_d", "monthly_return_categories_json"}:
            return ("No current feed", "Typed field exists; no Gmail, custodian, create/redeem, or Bloomberg writer", "Fund Accounting must provide and approve a supported calculation when applicable")
        return ("Gmail export - U.S. Bank EagleSTAR Trial Balance, or tool constant where identified", "eaglestar.py writes mapped accounting values; filing_master.py supplies explicit defaults", "Fund Accounting verifies which values came from the attachment versus a coded assumption")
    if name in flows:
        if name.endswith("reinvestment"):
            return ("No current feed; filing_master.py supplies N/A", "Create/redeem orders do not contain reinvested distributions", "Transfer Agent/Fund Accounting supplies evidence or approves not applicable")
        return ("Gmail export - EagleSTAR Trial Balance", "eaglestar.py writes subscriptions/redemptions; create/redeem order export is reconciliation-only", "Fund Accounting resolves the accounting-versus-order-book reconciliation")
    if name in returns:
        return ("Bloomberg formula in filing_master.xlsx", "BDP custom total-return formula for each reporting month", "Reviewer opens/saves on Bloomberg terminal and verifies dates and fund scope")
    if name in risk:
        if name in {"cur_metrics_json", "credit_sprd_risk_ig_json", "credit_sprd_risk_nonig_json"}:
            return ("Bloomberg formulas plus U.S. Bank custodian market value", "filing_master.py aggregates duration/rating fields after workbook calculation", "Risk reviewer verifies Bloomberg calculation and aggregation")
        return ("No current feed", "B.9/B.10 typed fields and preflight gates exist, but no writer populates them", "Derivatives Risk supplies and approves the applicable calculation")
    if name in designated:
        return ("Human-review workbook designated_index sheet (legacy 497K seed)", "humanreview.py seed values merge through filing_master.py", "Reviewer verifies the seed against the current filing; current workbook does not record reviewer identity")
    return ("Approved internal control record", "Direct filing-master mapping", "Assigned data owner verifies value and evidence")


def holding_source(group: str, name: str) -> tuple[str, str, str]:
    if group == "liquidity":
        return ("No current feed", "Policy-conditional typed C.7 field exists; no current writer populates it", "Liquidity Risk must supply and approve the data if the registry marks C.7 applicable")
    if group == "swap":
        if name in {"termination_dt", "notional_amt", "swap_cur_cd", "swap_flag", "ref_cusip"}:
            return ("U.S. Bank custodian CSV", "custodian.py parses the swap ticker and custody values", "Reviewer verifies parsed economics against the trade record")
        return ("Human-review workbook swap_legs (TRS-template seed; spread blank)", "master_sheet.py overlays workbook values onto custodian-created swap rows", "Derivatives Operations verifies every seed against the confirm; current workbook does not record reviewer identity")
    if group in {"option", "ref_instrument", "other_deriv", "forward"}:
        if name == "delta":
            return ("Manual value preserved in security_master.xlsx", "No automated feed; master_sheet.py preserves the workbook value", "Risk reviewer identifies the source system, as-of date, and approval")
        if name in {"put_or_call", "written_or_pur", "share_no", "exercise_price", "exercise_price_cur_cd", "exp_dt", "payoff_prof_deriv"}:
            return ("U.S. Bank custodian CSV", "custodian.py parses option/derivative terms from the custody position", "Derivatives Operations verifies parsed contract terms")
        if name in {"ref_index_name", "ref_index_identifier"}:
            return ("Human-review workbook option_index sheet (legacy-map seed)", "master_sheet.py applies the workbook reference-index mapping", "Reviewer verifies the seed from evidence; current workbook does not record reviewer identity")
        if name in {"ref_issuer_name", "ref_issue_title", "ref_isin", "ref_ticker"}:
            return ("Bloomberg formula in security_master.xlsx", "master_sheet.py BDP formulas enrich the reference security", "Reviewer saves on Bloomberg terminal and verifies the result")
        if name == "ref_cusip":
            return ("U.S. Bank custodian CSV", "custodian.py parses/copies the reference CUSIP", "Reviewer verifies the reference instrument")
        return ("Current workbook field; no verified automated writer for this condition", "Mapped by master_sheet.py when present", "Derivatives Operations documents the evidence or leaves it inapplicable")
    if group == "deriv_common":
        if name == "unrealized_appr":
            return ("Gmail export - U.S. Bank EagleSTAR PVal attachment", "eaglestar.py matches Primary Asset ID; custodian.py uses it as derivative value", "Fund Accounting verifies the 2026-06-24 PVal cutoff and match coverage")
        if name in {"counterparty_name", "counterparty_lei"}:
            return ("GLEIF-seeded human-review workbook for swaps; tool OCC constant for options", "master_sheet.py workbook overlay or custodian.py OCC constant", "Reviewer verifies seeded swap counterparties and the applicable route")
        return ("U.S. Bank custodian CSV plus tool classification", "custodian.py classifies the position as a derivative", "Derivatives Operations verifies the classification")
    if group == "debt":
        if name in {"maturity_dt", "coupon_kind", "annualized_rt"}:
            return ("Bloomberg formula in security_master.xlsx", "master_sheet.py BDP formulas populate C.9 terms", "Reviewer saves on Bloomberg terminal and verifies terms")
        return ("Tool constant in custodian.py", "Current code writes N for default/arrears/paid-in-kind flags", "Human reviewer must confirm the coded assumption is true for every applicable debt position")
    if name in {"inv_country", "isin", "ticker", "lei", "cusip", "asset_cat", "issuer_cat"}:
        if name in {"inv_country", "isin", "lei"}:
            return ("Bloomberg formula in security_master.xlsx", "master_sheet.py BDP formulas enrich the custody row", "Reviewer saves on Bloomberg terminal and verifies exceptions")
        if name in {"ticker", "cusip"}:
            return ("U.S. Bank custodian CSV", "custodian.py copies the custody identifier", "Reviewer verifies exceptions; never invent an identifier")
        return ("Tool classification from U.S. Bank custodian position", "custodian.py maps the holding type to N-PORT taxonomy", "Reviewer verifies exceptional classifications")
    if name in {"balance", "units", "cur_cd", "val_usd", "pct_val", "exchange_rt"}:
        if name == "pct_val":
            return ("Derived from valUSD and filed netAssets", "custodian.py recalculates after EagleSTAR overrides", "Fund Accounting verifies the denominator and total")
        return ("U.S. Bank custodian CSV", "custodian.py transforms the custody position; swap valUSD is later replaced by EagleSTAR PVal", "Operations verifies the 2026-06-25 cutoff and position reconciliation")
    return ("U.S. Bank custodian CSV or tool constant, as identified by custodian.py", "Custodian transform plus input validation", "Operations reviews the coded mapping and exceptions")


def applicable(row: dict[str, str], name: str, group: str, required: str, condition: str) -> bool:
    if required == "always":
        return True
    if required == "never":
        return present(row.get(_csv(name)))
    deriv = row.get("derivCat", "")
    if group == "conditional":
        if name == "issuer_conditional_desc": return row.get("issuerCat") == "OTHER"
        if name == "asset_conditional_desc": return row.get("assetCat") == "OTHER"
        if name == "exchange_rt": return row.get("curCd") not in {"", "USD"}
    if group == "debt": return row.get("assetCat") == "DBT" or present(row.get("maturityDt"))
    if group == "deriv_common": return present(deriv)
    if group == "option": return deriv in {"OPT", "SWO", "WAR"}
    if group == "ref_instrument":
        typ = row.get("refInstType", "")
        return ("indexBasket" in condition and typ == "indexBasket") or ("otherRefInst" in condition and typ == "otherRefInst")
    if group == "swap":
        if deriv != "SWP": return False
        if "recFixedOrFloating==Fixed" in condition: return row.get("recFixedOrFloating") == "Fixed"
        if "recFixedOrFloating==Floating" in condition: return row.get("recFixedOrFloating") == "Floating"
        if "pmntFixedOrFloating==Fixed" in condition: return row.get("pmntFixedOrFloating") == "Fixed"
        if "pmntFixedOrFloating==Floating" in condition: return row.get("pmntFixedOrFloating") == "Floating"
        return True
    if group == "forward": return deriv in {"FWD", "FUT"}
    if group == "other_deriv": return deriv == "OTH"
    return False  # C.7 applicability is not knowable until registry approval.


def _csv(name: str) -> str:
    return next((key for key, value in _HOLDINGS_KEY_MAP.items() if value == name), name)


def coverage(values: list[str], applicable_count: int | None = None) -> str:
    denom = len(values) if applicable_count is None else applicable_count
    num = sum(present(v) for v in values)
    return f"{num}/{denom} populated" if denom else "No currently applicable rows"


def row(cells: list[str]) -> str:
    return "<tr>" + "".join(f"<td>{html.escape(str(c))}</td>" for c in cells) + "</tr>"


def section(title: str, intro: str, rows: list[list[str]]) -> str:
    head = ["Input / XML field", "Current writer/source", "Transform and code", "Applicability", "Observed state", "Human control"]
    return (
        f'<section><h2>{html.escape(title)}</h2><p>{html.escape(intro)}</p><table class="trace">'
        + "<tr>" + "".join(f"<th>{h}</th>" for h in head) + "</tr>"
        + "".join(row(r) for r in rows) + "</table></section>"
    )


def main() -> None:
    dirs = production_dirs()
    configs = [kv(d / "fund_config.txt") for d in dirs]
    filings = [kv(d / "filings" / PERIOD / "filing_data.txt") for d in dirs]
    holdings = holding_rows(dirs)

    config_reverse = {v: k for k, v in _CONFIG_KEY_MAP.items()}
    filing_reverse = {v: k for k, v in _FILING_KEY_MAP.items()}

    config_rows: list[list[str]] = []
    for f in fields(FundConfig):
        key = config_reverse[f.name]
        source, transform, review = config_source(f.name)
        values = [x.get(key, "") for x in configs]
        state = coverage(values)
        if f.name == "ccc":
            placeholders = sum(str(v).strip() in {"", "XXXXXXXX", "N/A"} for v in values)
            state = f"{placeholders}/{len(values)} blank or placeholder; credential values intentionally not printed"
        config_rows.append([f"{key} -> {f.name}", source, f"{transform}; config.py parse_config", "Every filing", state, review])

    filing_rows: list[list[str]] = []
    for f in fields(FilingData):
        key = filing_reverse[f.name]
        source, transform, review = filing_source(f.name)
        values = [x.get(key, "") for x in filings]
        optional = f.default == ""
        applicability = "Policy/regime conditional" if optional else "Every filing unless Form instructions make the item inapplicable"
        filing_rows.append([f"{key} -> {f.name}", source, f"{transform}; filing_master.py / builder.py", applicability, coverage(values), review])

    holding_rows_out: list[list[str]] = []
    for f in fields(Holding):
        spec = FIELD_BY_NAME[f.name]
        key = spec.csv_header
        source, transform, review = holding_source(spec.group, f.name)
        applicable_rows = [r for r in holdings if applicable(r, f.name, spec.group, spec.required, spec.condition)]
        values = [r.get(key, "") for r in applicable_rows]
        if spec.group == "liquidity":
            state = f"{sum(present(r.get(key, '')) for r in holdings)}/{len(holdings)} populated; required denominator unresolved until registry approval"
        else:
            state = coverage(values, len(applicable_rows))
        condition = "Always" if spec.required == "always" else (spec.condition or "Optional when supported")
        holding_rows_out.append([f"{key} -> {f.name}", source, f"{transform}; custodian.py / humanreview.py / builder.py", condition, state, review])

    appendix = "".join([
        section("11. Fund configuration field traceability", "Every static filing identity field is listed below. Coverage is measured across the 105 production fund directories; credential contents are never printed.", config_rows),
        section("12. Filing-level field traceability", "Every FilingData field is listed below. A populated cell proves presence only; the named human owner must still verify period, scope, method, and evidence.", filing_rows),
        section("13. Holding and derivative field traceability", f"Every Holding field is listed below. Coverage is measured across {len(holdings):,} current holding rows using each field's known condition. Policy-dependent C.7 applicability remains unresolved until the fund registry is approved.", holding_rows_out),
        '<section><h2>14. How to interpret traceability</h2><div class="box"><b>Presence is not approval.</b> "Populated" means a nonblank value exists in the current input. It can include an N/A entry where the source supplied one. It does not prove accuracy, freshness, support for N/A, or regulatory applicability.</div><table><tr><th>State</th><th>Meaning</th><th>Human action</th></tr><tr><td>Populated</td><td>Automation found a nonblank value.</td><td>Owner verifies source, date, methodology, N/A support, and reconciliation.</td></tr><tr><td>Blank for an applicable row</td><td>A required or conditionally required input is absent.</td><td>Owner supplies internal evidence or records a supported not-applicable decision; no guessing.</td></tr><tr><td>Applicability unresolved</td><td>The approved policy record does not yet exist.</td><td>Compliance/Risk decides applicability before data population or release.</td></tr><tr><td>Optional</td><td>Form/schema does not always require the field.</td><td>Populate only when supported; never fabricate identifiers or terms.</td></tr></table><p><b>Code enforcement:</b> parsing is in <code>config.py</code>; policy derivation is in <code>policy.py</code>; source and coverage gates are in <code>preflight.py</code>; conditional validation is in <code>input_validation.py</code>; XML construction is in <code>builder.py</code>; SEC validation is in <code>xsd_validator.py</code>.</p></section>',
    ])

    text = RUNBOOK.read_text(encoding="utf-8")
    registry_gap = '<tr><td>Approved registry</td><td class="block">Missing; 0 release-eligible funds</td><td>Compliance / Legal</td></tr>'
    reconciliation_gap = '<tr><td>Reconciliation</td><td class="block">258 REVIEW rows: 238 flow, 19 netAssets, 1 liability; 137 source fund identifiers</td><td>Fund Accounting / Operations / Capital Markets</td></tr>'
    if reconciliation_gap not in text:
        text = text.replace(registry_gap, reconciliation_gap + registry_gap)
    provenance_start = text.index("<!-- PROVENANCE:BEGIN -->") + len("<!-- PROVENANCE:BEGIN -->")
    provenance_end = text.index("<!-- PROVENANCE:END -->")
    text = text[:provenance_start] + "\n" + current_lineage_section() + "\n" + text[provenance_end:]
    start = text.index("<!-- TRACEABILITY:BEGIN -->") + len("<!-- TRACEABILITY:BEGIN -->")
    end = text.index("<!-- TRACEABILITY:END -->")
    RUNBOOK.write_text(text[:start] + "\n" + appendix + "\n" + text[end:], encoding="utf-8")


if __name__ == "__main__":
    main()
