"""Fund-period input workflow and fail-closed provenance controls."""

import csv
from decimal import Decimal
from pathlib import Path
import shutil

from openpyxl import load_workbook

from nport.config import _CONFIG_KEY_MAP, _FILING_KEY_MAP, parse_filing
from nport.fund_review import evaluate_inputs, finalize_inputs, prepare_inputs


def _sheet_rows(workbook: Path, sheet: str):
    wb = load_workbook(workbook, data_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, ("" if value is None else str(value) for value in row)))
            for row in ws.iter_rows(min_row=2, values_only=True)]


def test_prepare_inputs_creates_simple_workbook_and_bloomberg_formulas(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")

    workbook = prepare_inputs(fund_dir, "2025-12")
    wb = load_workbook(workbook, data_only=True)

    assert wb.sheetnames == [
        "Summary", "Bloomberg", "Sources", "GapGuide", "FundFields",
        "HoldingFields", "ReconciliationInputs", "Reconciliation",
    ]
    assert workbook.name == "filing_inputs.xlsx"
    assert "Approvals" not in wb.sheetnames
    assert [cell.value for cell in wb["FundFields"][1]] == [
        "gapId", "targetFile", "recordKey", "fieldName", "currentValue",
        "proposedValue", "sourceId", "status", "comment",
    ]
    assert [cell.value for cell in wb["Sources"][1]] == [
        "sourceId", "dataset", "sourceType", "sourceSystem", "sourcePath",
        "sourceAsOf", "sha256", "recordCount", "comment",
    ]
    assert [cell.value for cell in wb["GapGuide"][1]] == [
        "gapId", "category", "gap", "inputSheet", "rowToFind",
        "columnsToComplete", "sourceRow", "completionRule", "evidenceToRegister",
    ]
    assert [cell.value for cell in wb["ReconciliationInputs"][1]] == [
        "checkId", "category", "actualBasis", "controlValue", "tolerance",
        "sourceId", "status", "comment",
    ]
    bloomberg = load_workbook(workbook, data_only=False)["Bloomberg"]
    assert bloomberg["F2"].value.startswith("=BDP(")
    formula_wb = load_workbook(workbook, data_only=False)
    fund_ws = formula_wb["FundFields"]
    fund_headers = {cell.value: cell.column for cell in fund_ws[1]}
    return_row = next(
        row for row in range(2, fund_ws.max_row + 1)
        if fund_ws.cell(row, fund_headers["fieldName"]).value == "rtn1"
    )
    assert fund_ws.cell(return_row, fund_headers["status"]).value.startswith("=IFERROR(")
    assert len(list(wb["GapGuide"].iter_rows(min_row=2, values_only=True))) == 13
    guide = {row["gapId"]: row for row in _sheet_rows(workbook, "GapGuide")}
    assert "FundFields[gapId starts G-002]" in guide["G-002"]["rowToFind"]
    assert guide["G-013"]["inputSheet"] == "No workbook entry"
    totals = [row for row in _sheet_rows(workbook, "FundFields")
              if row["fieldName"] == "totAssets"]
    assert totals[0]["currentValue"] == ""
    assert totals[0]["proposedValue"] == ""
    assert totals[0]["status"] == "MISSING"
    assert not any(row["fieldName"] == "policySourceRef"
                   for row in _sheet_rows(workbook, "FundFields"))


def test_prohibited_us_bank_source_is_reported_as_blocker(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "US Bank Positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["Sources"]
    headers = {cell.value: cell.column for cell in ws[1]}
    for field, value in {
        "sourceSystem": "U.S. Bank custody", "sourceAsOf": "2025-12-31",
    }.items():
        ws.cell(2, headers[field], value)
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert "INPUT_SOURCE_PROHIBITED" in {item.code for item in result["blockers"]}


def test_prepare_normalizes_pdf_line_break_inside_positions_path(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "internal_positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    copied_from_pdf = f"{positions.parent}\n\\{positions.name}"

    workbook = prepare_inputs(fund_dir, "2025-12", positions=copied_from_pdf)

    assert workbook.is_file()
    assert _sheet_rows(workbook, "Sources")[0]["sourcePath"] == str(positions.resolve())


def test_missing_reconciliation_input_sheet_fails_closed(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    workbook = prepare_inputs(fund_dir, "2025-12")
    wb = load_workbook(workbook)
    del wb["ReconciliationInputs"]
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    findings = [item for item in result["blockers"] if item.code == "G-012_INPUT"]
    assert findings
    assert "create ReconciliationInputs" in findings[0].technical


def test_supplied_inputs_finalize_without_required_source_metadata(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "internal_positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)
    filing = parse_filing(fdrs_dir / "filings" / "2025-12" / "filing_data.txt")

    wb = load_workbook(workbook)

    policy_values = {
        "fiscalYearEndMMDD": "12-31", "derivativesRegimePolicy": "NONE",
        "liquidityRequired": "N", "cashB2fRequired": "N",
        "policyEffectiveFrom": "2020-01-01",
        "policySourceRef": "Internal policy memorandum 2025-01",
    }
    fields = wb["FundFields"]
    field_headers = {cell.value: cell.column for cell in fields[1]}
    for rownum in range(2, fields.max_row + 1):
        target = fields.cell(rownum, field_headers["targetFile"]).value
        name = fields.cell(rownum, field_headers["fieldName"]).value
        status = fields.cell(rownum, field_headers["status"]).value
        if status in {"SYSTEM_DERIVED", "SYSTEM_CONTROL"}:
            continue
        value = ""
        if target == "fund_config.txt":
            value = policy_values.get(name, fields.cell(rownum, field_headers["currentValue"]).value or "")
        elif name in _FILING_KEY_MAP:
            value = getattr(filing, _FILING_KEY_MAP[name])
        if not value:
            fields.cell(rownum, field_headers["status"], "NOT_APPLICABLE")
            fields.cell(rownum, field_headers["comment"], "Not applicable for this synthetic test fund")
        else:
            fields.cell(rownum, field_headers["proposedValue"], str(value))
            fields.cell(rownum, field_headers["status"], "PROVIDED")
        fields.cell(rownum, field_headers["sourceId"], "")

    holdings = wb["HoldingFields"]
    holding_headers = {cell.value: cell.column for cell in holdings[1]}
    for rownum in range(2, holdings.max_row + 1):
        current = holdings.cell(rownum, holding_headers["currentValue"]).value or ""
        holdings.cell(rownum, holding_headers["status"], "PROVIDED" if current else "NOT_APPLICABLE")
        holdings.cell(rownum, holding_headers["sourceId"], "")
        if not current:
            holdings.cell(rownum, holding_headers["comment"], "Not applicable in source record")

    with open(positions, newline="", encoding="utf-8-sig") as handle:
        position_total = sum(
            (Decimal(row["valUSD"]) for row in csv.DictReader(handle)), Decimal("0")
        )
    recon = wb["ReconciliationInputs"]
    recon_headers = {cell.value: cell.column for cell in recon[1]}
    for rownum in range(2, recon.max_row + 1):
        check_id = recon.cell(rownum, recon_headers["checkId"]).value
        if check_id == "POSITIONS_TO_GL":
            value = position_total
            status = "PROVIDED"
            comment = "Independent GL control balance"
        else:
            field = check_id.split(":", 1)[1]
            raw = getattr(filing, _FILING_KEY_MAP[field])
            value = "" if str(raw).upper() == "N/A" else raw
            status = "NOT_APPLICABLE" if str(raw).upper() == "N/A" else "PROVIDED"
            comment = "Filed flow is not applicable" if status == "NOT_APPLICABLE" else "Independent flow control"
        recon.cell(rownum, recon_headers["controlValue"], str(value))
        recon.cell(rownum, recon_headers["tolerance"], "0.01" if status == "PROVIDED" else "")
        recon.cell(rownum, recon_headers["sourceId"], "")
        recon.cell(rownum, recon_headers["status"], status)
        recon.cell(rownum, recon_headers["comment"], comment)

    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert result["blockers"] == []
    checks = {(row["check"], row["status"]) for row in result["reconciliation"]}
    assert ("Holdings total to GL control", "PASS") in checks
    assert ("Filed mon1Sales to independent flow control", "PASS") in checks
    bundle = finalize_inputs(fund_dir, "2025-12", output_root=tmp_path / "builds", run_id="test-run")
    assert {path.name for path in bundle.iterdir()} == {
        "fund_config.txt", "filing_data.txt", "holdings.csv", "source_manifest.csv",
        "field_provenance.csv", "reconciliation.csv", "input_receipt.json",
    }
    provenance = (bundle / "field_provenance.csv").read_text(encoding="utf-8")
    assert "fund_config.txt,FUND,cik" in provenance
    assert "filing_data.txt,FUND,totAssets" in provenance
    manifest = (bundle / "source_manifest.csv").read_text(encoding="utf-8")
    assert "internal_positions" in manifest


def test_legacy_human_review_is_migrated_without_signoff_columns(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    filing_dir = fund_dir / "filings" / "2025-12"
    filing_dir.mkdir(parents=True)
    legacy = filing_dir / "human_review.xlsx"
    first = prepare_inputs(fund_dir, "2025-12")
    first.replace(legacy)
    wb = load_workbook(legacy)
    ws = wb["FundFields"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["proposedValue"], "0000000000")
    ws.cell(2, headers["sourceId"], "ENTITY")
    ws.cell(2, headers["status"], "APPROVED")
    wb.save(legacy)

    migrated = prepare_inputs(fund_dir, "2025-12")
    assert migrated.name == "filing_inputs.xlsx"
    assert legacy.is_file()
    rows = _sheet_rows(migrated, "FundFields")
    assert rows[0]["proposedValue"] == "0000000000"
    assert rows[0]["status"] == "MISSING"
