"""Contract tests for the versioned multi-fund input workbook."""

import csv
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from nport.input_workbook import (
    CONFIG_HEADERS,
    CONFIG_SHEET_SUFFIX,
    FUND_DATA_HEADERS,
    FUND_DATA_SHEET_SUFFIX,
    INPUT_WORKBOOK_FORMAT_VERSION,
    ORDERS_SHEET,
    POSITION_BLOOMBERG_FIELDS,
    POSITION_HEADERS,
    POSITIONS_SHEET_SUFFIX,
    _expected_sheet_names,
    _sheet_name,
    _unresolved_bloomberg_value,
    create_input_workbook_template,
    prepare_from_input_workbook,
)


def _fund_dir(tmp_path: Path, fdrs_dir: Path) -> Path:
    fund_dir = tmp_path / "funds" / "fdrs"
    fund_dir.mkdir(parents=True)
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    return fund_dir


def _headers(sheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1]}


def _fund_sheets(fund: str) -> tuple[str, str, str]:
    return (
        _sheet_name(fund, CONFIG_SHEET_SUFFIX),
        _sheet_name(fund, FUND_DATA_SHEET_SUFFIX),
        _sheet_name(fund, POSITIONS_SHEET_SUFFIX),
    )


def _first_fixture_holding(fdrs_dir: Path) -> dict[str, str]:
    with (fdrs_dir / "filings" / "2025-12" / "holdings.csv").open(
        newline="", encoding="utf-8"
    ) as handle:
        return next(csv.DictReader(handle))


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("#N/A Requesting Data...", True),
        ("#VALUE!", True),
        ("Requesting Data", True),
        ("N/A", False),
        ("US58733R1023", False),
    ],
)
def test_bloomberg_placeholder_detection(value, expected):
    assert _unresolved_bloomberg_value(value) is expected


def test_template_contains_only_input_sheets(tmp_path, fdrs_dir):
    fund_dir = _fund_dir(tmp_path, fdrs_dir)
    output = tmp_path / "nport-inputs-v4_2026-07_nport-v1.13.xlsx"

    create_input_workbook_template(fund_dir.parent, ["2026-07"], output)

    workbook = load_workbook(output, data_only=False)
    try:
        config_name, fund_data_name, positions_name = _fund_sheets("FDRS")
        assert workbook.sheetnames == _expected_sheet_names(["FDRS"])
        assert tuple(cell.value for cell in workbook[config_name][1]) == CONFIG_HEADERS
        assert tuple(cell.value for cell in workbook[fund_data_name][1]) == FUND_DATA_HEADERS
        assert tuple(cell.value for cell in workbook[positions_name][1]) == POSITION_HEADERS
        metadata = {item.name: item.value for item in workbook.custom_doc_props}
        assert metadata["nportFormatVersion"] == INPUT_WORKBOOK_FORMAT_VERSION
        assert metadata["nportSchemaVersion"] == "1.13"
        assert metadata["nportPeriods"] == "2026-07"
        assert CONFIG_HEADERS == ("fieldName", "value")
        assert FUND_DATA_HEADERS == ("Period", "fieldName", "value")
        forbidden_headers = {
            "currentValue", "proposedValue", "sourceId", "status", "comment",
            "controlValue", "tolerance",
        }
        for sheet in workbook.worksheets:
            assert forbidden_headers.isdisjoint(cell.value for cell in sheet[1])
        fund_data = workbook[fund_data_name]
        fund_headers = _headers(fund_data)
        input_fields = {
            fund_data.cell(row, fund_headers["fieldName"]).value
            for row in range(2, fund_data.max_row + 1)
        }
        assert not {
            "submissionType", "liveTestFlag", "repPdEnd", "repPdDate",
            "derivativesRegime",
        } & input_fields
        return_row = next(
            row for row in range(2, fund_data.max_row + 1)
            if fund_data.cell(row, fund_headers["fieldName"]).value == "rtn1"
        )
        assert "CUST_TRR_RETURN_HOLDING_PER" in fund_data.cell(
            return_row, fund_headers["value"]
        ).value
        positions = workbook[positions_name]
        position_headers = _headers(positions)
        for field in POSITION_BLOOMBERG_FIELDS:
            assert "BDP(" in positions.cell(2, position_headers[field]).value
    finally:
        workbook.close()


def test_input_workbook_prepares_normal_review_workbook(tmp_path, fdrs_dir):
    fund_dir = _fund_dir(tmp_path, fdrs_dir)
    input_path = tmp_path / "nport-inputs-v4_2026-07_nport-v1.13.xlsx"
    create_input_workbook_template(fund_dir.parent, ["2026-07"], input_path)

    workbook = load_workbook(input_path)
    positions = workbook[_fund_sheets("FDRS")[2]]
    position_headers = _headers(positions)
    holding = _first_fixture_holding(fdrs_dir)
    positions.cell(2, position_headers["holdingId"], "MELI")
    for field, value in holding.items():
        positions.cell(2, position_headers[field], value)
    positions.cell(2, position_headers["liquidityClassificationJson"], "N/A")

    orders = workbook[ORDERS_SHEET]
    order_headers = _headers(orders)
    orders.cell(2, order_headers["Side"], "CREATE")
    orders.cell(2, order_headers["Trade Date"], "7/15/2026")
    orders.cell(2, order_headers["Notional"], "12.34")
    orders.cell(2, order_headers["Status"], "ACCEPTED")

    fund_data = workbook[_fund_sheets("FDRS")[1]]
    data_headers = _headers(fund_data)
    for row_number in range(2, fund_data.max_row + 1):
        if fund_data.cell(row_number, data_headers["fieldName"]).value == "totAssets":
            fund_data.cell(row_number, data_headers["value"], "100.00")
            break
    workbook.save(input_path)
    workbook.close()

    review_path = prepare_from_input_workbook(
        fund_dir, "FDRS", "2026-07", input_path,
    )

    assert review_path == fund_dir / "filings" / "2026-07" / "filing_inputs.xlsx"
    review = load_workbook(review_path, data_only=False)
    try:
        holding_sheet = review["HoldingFields"]
        holding_headers = _headers(holding_sheet)
        holding_rows = {
            (
                holding_sheet.cell(row, holding_headers["recordKey"]).value,
                holding_sheet.cell(row, holding_headers["fieldName"]).value,
            ): row
            for row in range(2, holding_sheet.max_row + 1)
        }
        liquidity_row = holding_rows[("MELI", "liquidityClassificationJson")]
        assert holding_sheet.cell(
            liquidity_row, holding_headers["proposedValue"]
        ).value == "N/A"
        assert holding_sheet.cell(
            liquidity_row, holding_headers["status"]
        ).value == "PROVIDED"

        fund_sheet = review["FundFields"]
        fund_headers = _headers(fund_sheet)
        total_row = next(
            row for row in range(2, fund_sheet.max_row + 1)
            if fund_sheet.cell(row, fund_headers["fieldName"]).value == "totAssets"
        )
        assert fund_sheet.cell(total_row, fund_headers["proposedValue"]).value == "100.00"
        assert fund_sheet.cell(total_row, fund_headers["status"]).value == "PROVIDED"
        sales_row = next(
            row for row in range(2, fund_sheet.max_row + 1)
            if fund_sheet.cell(row, fund_headers["fieldName"]).value == "mon3Sales"
        )
        assert fund_sheet.cell(sales_row, fund_headers["currentValue"]).value == "12.34"
        assert fund_sheet.cell(sales_row, fund_headers["status"]).value == "PROVIDED"
        assert fund_sheet.cell(sales_row, fund_headers["sourceId"]).value == "ORDERS"
        source_ids = [cell.value for cell in review["Sources"]["A"]]
        assert "INPUT_WORKBOOK" in source_ids
    finally:
        review.close()

    normalized = list((tmp_path / "_normalized" / "nport-v1.13").glob("*/fdrs/2026-07/positions.csv"))
    assert len(normalized) == 1
    with normalized[0].open(newline="", encoding="utf-8") as handle:
        normalized_holding = next(csv.DictReader(handle))
    assert normalized_holding["holdingId"] == "MELI"
    assert normalized_holding["name"] == "MercadoLibre Inc"


def test_input_workbook_rejects_formula_input(tmp_path, fdrs_dir):
    fund_dir = _fund_dir(tmp_path, fdrs_dir)
    input_path = tmp_path / "nport-inputs-v4_2026-07_nport-v1.13.xlsx"
    create_input_workbook_template(fund_dir.parent, ["2026-07"], input_path)
    workbook = load_workbook(input_path)
    positions = workbook[_fund_sheets("FDRS")[2]]
    headers = _headers(positions)
    positions.cell(2, headers["holdingId"], "FORMULA")
    positions.cell(2, headers["name"], "=1+1")
    workbook.save(input_path)
    workbook.close()

    with pytest.raises(ValueError, match="formulas are not accepted"):
        prepare_from_input_workbook(fund_dir, "FDRS", "2026-07", input_path)


def test_input_workbook_rejects_deleted_contract_row(tmp_path, fdrs_dir):
    fund_dir = _fund_dir(tmp_path, fdrs_dir)
    input_path = tmp_path / "nport-inputs-v4_2026-07_nport-v1.13.xlsx"
    create_input_workbook_template(fund_dir.parent, ["2026-07"], input_path)
    workbook = load_workbook(input_path)
    sheet = workbook[_fund_sheets("FDRS")[1]]
    headers = _headers(sheet)
    target = next(
        row for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, headers["fieldName"]).value == "backtestingExceptions"
    )
    sheet.delete_rows(target)
    workbook.save(input_path)
    workbook.close()

    with pytest.raises(ValueError, match="FDRS_FundData: invalid field contract.*backtestingExceptions"):
        prepare_from_input_workbook(fund_dir, "FDRS", "2026-07", input_path)


def test_unresolved_generated_bloomberg_formula_remains_in_review_path(
    tmp_path, fdrs_dir,
):
    fund_dir = _fund_dir(tmp_path, fdrs_dir)
    input_path = tmp_path / "nport-inputs-v4_2026-07_nport-v1.13.xlsx"
    create_input_workbook_template(fund_dir.parent, ["2026-07"], input_path)
    workbook = load_workbook(input_path)
    sheet = workbook[_fund_sheets("FDRS")[2]]
    headers = _headers(sheet)
    for field, value in {
        "holdingId": "FORMULA-HOLDING",
        "name": "Formula Holding",
        "title": "Formula Holding",
        "cusip": "000000000",
        "ticker": "FORM",
        "assetCat": "EC",
    }.items():
        sheet.cell(2, headers[field], value)
    workbook.save(input_path)
    workbook.close()

    review_path = prepare_from_input_workbook(
        fund_dir, "FDRS", "2026-07", input_path,
    )

    with next(
        (tmp_path / "_normalized" / "nport-v1.13").glob(
            "*/fdrs/2026-07/positions.csv"
        )
    ).open(newline="", encoding="utf-8") as handle:
        normalized = next(csv.DictReader(handle))
    assert normalized["isin"] == ""

    review = load_workbook(review_path, data_only=False)
    try:
        holding_sheet = review["HoldingFields"]
        holding_headers = _headers(holding_sheet)
        isin_row = next(
            row for row in range(2, holding_sheet.max_row + 1)
            if holding_sheet.cell(row, holding_headers["recordKey"]).value == "FORMULA-HOLDING"
            and holding_sheet.cell(row, holding_headers["fieldName"]).value == "isin"
        )
        assert str(holding_sheet.cell(
            isin_row, holding_headers["proposedValue"],
        ).value).startswith("='Bloomberg'!")
        assert str(holding_sheet.cell(
            isin_row, holding_headers["status"],
        ).value).startswith("=IFERROR(")
    finally:
        review.close()
