"""Fund-period input workflow and fail-closed provenance controls."""

import csv
import shutil
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

import nport.pipeline as pipeline
from nport.cli import main
from nport.config import FILING_KEY_MAP, parse_filing
from nport.pipeline import (
    _holding_objects,
    evaluate_inputs,
    finalize_inputs,
    prepare_inputs,
)


def _sheet_rows(workbook: Path, sheet: str):
    wb = load_workbook(workbook, data_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, ("" if value is None else str(value) for value in row)))
            for row in ws.iter_rows(min_row=2, values_only=True)]


def _ready_result(tmp_path, factories):
    workbook = tmp_path / "filing_inputs.xlsx"
    workbook.write_bytes(b"test workbook identity")
    return {
        "workbook": workbook,
        "config": factories.config(),
        "filing": factories.filing(),
        "holdings": [factories.equity()],
        "sources": [],
        "usedSources": set(),
        "blockers": [],
        "provenance": [],
        "reconciliation": [],
        "reconciliationInputs": [],
    }


def test_prepare_inputs_creates_simple_workbook_and_bloomberg_formulas(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")

    workbook = prepare_inputs(fund_dir, "2025-12")
    wb = load_workbook(workbook, data_only=True)

    assert wb.sheetnames == [
        "Summary", "Bloomberg", "Sources", "GapGuide", "FundFields",
        "HoldingFields", "ReconciliationInputs", "Reconciliation",
    ]
    assert workbook.name == "filing_inputs.xlsx"
    assert "Approvals" not in wb.sheetnames
    assert [cell.value for cell in wb["FundFields"][1]] == [
        "gapId", "targetFile", "recordKey", "fieldName", "currentValue",
        "proposedValue", "sourceId", "status", "comment",
    ]
    assert [cell.value for cell in wb["Sources"][1]] == [
        "sourceId", "dataset", "sourceType", "sourceSystem", "sourcePath",
        "sourceAsOf", "sha256", "recordCount", "comment",
    ]
    assert [cell.value for cell in wb["GapGuide"][1]] == [
        "gapId", "category", "gap", "inputSheet", "rowToFind",
        "columnsToComplete", "sourceRow", "completionRule", "evidenceToRegister",
    ]
    assert [cell.value for cell in wb["ReconciliationInputs"][1]] == [
        "checkId", "category", "actualBasis", "controlValue", "tolerance",
        "sourceId", "status", "comment",
    ]
    bloomberg = load_workbook(workbook, data_only=False)["Bloomberg"]
    assert bloomberg["F2"].value.startswith("=BDP(")
    formula_wb = load_workbook(workbook, data_only=False)
    fund_ws = formula_wb["FundFields"]
    fund_headers = {cell.value: cell.column for cell in fund_ws[1]}
    return_row = next(
        row for row in range(2, fund_ws.max_row + 1)
        if fund_ws.cell(row, fund_headers["fieldName"]).value == "rtn1"
    )
    assert fund_ws.cell(return_row, fund_headers["status"]).value.startswith("=IFERROR(")
    assert len(list(wb["GapGuide"].iter_rows(min_row=2, values_only=True))) == 13
    guide = {row["gapId"]: row for row in _sheet_rows(workbook, "GapGuide")}
    assert "FundFields[gapId starts G-002]" in guide["G-002"]["rowToFind"]
    assert guide["G-013"]["inputSheet"] == "No workbook entry"
    totals = [row for row in _sheet_rows(workbook, "FundFields")
              if row["fieldName"] == "totAssets"]
    assert totals[0]["currentValue"] == ""
    assert totals[0]["proposedValue"] == ""
    assert totals[0]["status"] == "MISSING"
    assert not any(row["fieldName"] == "policySourceRef"
                   for row in _sheet_rows(workbook, "FundFields"))


def test_prohibited_custodian_source_is_reported_as_blocker(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "Custodian Positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["Sources"]
    headers = {cell.value: cell.column for cell in ws[1]}
    for field, value in {
        "sourceSystem": "External custodian", "sourceAsOf": "2025-12-31",
    }.items():
        ws.cell(2, headers[field], value)
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert "INPUT_SOURCE_PROHIBITED" in {item.code for item in result["blockers"]}


def test_missing_reconciliation_input_sheet_fails_closed(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    workbook = prepare_inputs(fund_dir, "2025-12")
    wb = load_workbook(workbook)
    del wb["ReconciliationInputs"]
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    findings = [item for item in result["blockers"] if item.code == "G-012_INPUT"]
    assert findings
    assert "create ReconciliationInputs" in findings[0].technical


def test_supplied_inputs_finalize_without_required_source_metadata(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "internal_positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)
    filing = parse_filing(fdrs_dir / "filings" / "2025-12" / "filing_data.txt")

    wb = load_workbook(workbook)

    policy_values = {
        "fiscalYearEndMMDD": "12-31", "derivativesRegimePolicy": "NONE",
        "liquidityRequired": "N", "cashB2fRequired": "N",
        "policyEffectiveFrom": "2020-01-01",
        "policySourceRef": "Internal policy memorandum 2025-01",
    }
    fields = wb["FundFields"]
    field_headers = {cell.value: cell.column for cell in fields[1]}
    for rownum in range(2, fields.max_row + 1):
        target = fields.cell(rownum, field_headers["targetFile"]).value
        name = fields.cell(rownum, field_headers["fieldName"]).value
        status = fields.cell(rownum, field_headers["status"]).value
        if status in {"SYSTEM_DERIVED", "SYSTEM_CONTROL"}:
            continue
        value = ""
        if target == "fund_config.txt":
            value = policy_values.get(name, fields.cell(rownum, field_headers["currentValue"]).value or "")
        elif name in FILING_KEY_MAP:
            value = getattr(filing, FILING_KEY_MAP[name])
        if not value:
            fields.cell(rownum, field_headers["status"], "NOT_APPLICABLE")
            fields.cell(rownum, field_headers["comment"], "Not applicable for this synthetic test fund")
        else:
            fields.cell(rownum, field_headers["proposedValue"], str(value))
            fields.cell(rownum, field_headers["status"], "PROVIDED")
        fields.cell(rownum, field_headers["sourceId"], "")

    holdings = wb["HoldingFields"]
    holding_headers = {cell.value: cell.column for cell in holdings[1]}
    for rownum in range(2, holdings.max_row + 1):
        current = holdings.cell(rownum, holding_headers["currentValue"]).value or ""
        holdings.cell(rownum, holding_headers["status"], "PROVIDED" if current else "NOT_APPLICABLE")
        holdings.cell(rownum, holding_headers["sourceId"], "")
        if not current:
            holdings.cell(rownum, holding_headers["comment"], "Not applicable in source record")

    with open(positions, newline="", encoding="utf-8-sig") as handle:
        position_total = sum(
            (Decimal(row["valUSD"]) for row in csv.DictReader(handle)), Decimal("0")
        )
    recon = wb["ReconciliationInputs"]
    recon_headers = {cell.value: cell.column for cell in recon[1]}
    for rownum in range(2, recon.max_row + 1):
        check_id = recon.cell(rownum, recon_headers["checkId"]).value
        if check_id == "POSITIONS_TO_GL":
            value = position_total
            status = "PROVIDED"
            comment = "Independent GL control balance"
        else:
            field = check_id.split(":", 1)[1]
            raw = getattr(filing, FILING_KEY_MAP[field])
            value = "" if str(raw).upper() == "N/A" else raw
            status = "NOT_APPLICABLE" if str(raw).upper() == "N/A" else "PROVIDED"
            comment = "Filed flow is not applicable" if status == "NOT_APPLICABLE" else "Independent flow control"
        recon.cell(rownum, recon_headers["controlValue"], str(value))
        recon.cell(rownum, recon_headers["tolerance"], "0.01" if status == "PROVIDED" else "")
        recon.cell(rownum, recon_headers["sourceId"], "")
        recon.cell(rownum, recon_headers["status"], status)
        recon.cell(rownum, recon_headers["comment"], comment)

    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert result["blockers"] == []
    checks = {(row["check"], row["status"]) for row in result["reconciliation"]}
    assert ("Holdings total to GL control", "PASS") in checks
    assert ("Filed mon1Sales to independent flow control", "PASS") in checks
    bundle = finalize_inputs(fund_dir, "2025-12", bundle_root=tmp_path / "builds", run_id="test-run")
    assert {path.name for path in bundle.iterdir()} == {
        "fund_config.txt", "filing_data.txt", "holdings.csv", "source_manifest.csv",
        "field_provenance.csv", "reconciliation.csv", "input_receipt.json",
    }
    provenance = (bundle / "field_provenance.csv").read_text(encoding="utf-8")
    assert "fund_config.txt,FUND,cik" in provenance
    assert "filing_data.txt,FUND,totAssets" in provenance
    manifest = (bundle / "source_manifest.csv").read_text(encoding="utf-8")
    assert "internal_positions" in manifest

    main(["build", "FDRS", "2025-12", "--fund-dir", str(fund_dir), "--dry-run"])
    release_root = tmp_path / "output"
    main([
        "build", "FDRS", "2025-12", "--fund-dir", str(fund_dir),
        "--bundle-root", str(tmp_path / "cli-builds"),
        "--release-root", str(release_root),
    ])
    releases = list((release_root / "fdrs" / "2025-12" / "nport-v1.13").glob("*/*.xml"))
    assert len(releases) == 1
    xml = releases[0]
    assert xml.stem.startswith("FDRS_2025-12_nport-v1.13_")
    assert xml.is_file()
    assert xml.with_suffix(".manifest.json").is_file()


def test_refresh_drops_orphan_holding_review_rows(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["HoldingFields"]
    ws.append([
        "STALE", "holdings.csv", "REMOVED-POSITION", "couponKind", "Fixed",
        "Fixed", "POSITIONS", "PROVIDED", "stale row",
    ])
    wb.save(workbook)
    wb.close()

    prepare_inputs(fund_dir, "2025-12", positions=positions)
    assert "REMOVED-POSITION" not in {
        row["recordKey"] for row in _sheet_rows(workbook, "HoldingFields")
    }


def test_duplicate_workbook_key_is_a_blocker(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["FundFields"]
    ws.append([cell.value for cell in ws[2]])
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert "INPUT_DUPLICATE_ROW" in {finding.code for finding in result["blockers"]}


def test_unknown_status_is_rejected(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["FundFields"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["status"], "UNKNOWN")
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert "INPUT_STATUS" in {finding.code for finding in result["blockers"]}


def test_changed_position_source_hash_is_a_blocker(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    prepare_inputs(fund_dir, "2025-12", positions=positions)
    positions.write_text(positions.read_text(encoding="utf-8") + "\n", encoding="utf-8")

    result = evaluate_inputs(fund_dir, "2025-12")
    assert "INPUT_SOURCE_HASH" in {finding.code for finding in result["blockers"]}


@pytest.mark.parametrize(("field", "value", "code"), [
    ("sourceAsOf", "12/31/2025", "INPUT_SOURCE_ASOF"),
    ("recordCount", "54.5", "INPUT_SOURCE_COUNT"),
    ("sha256", "not-a-hash", "INPUT_SOURCE_SHA256"),
])
def test_source_metadata_types_are_validated(
    tmp_path, fdrs_dir, field, value, code,
):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    shutil.copyfile(fdrs_dir / "filings" / "2025-12" / "holdings.csv", positions)
    workbook = prepare_inputs(fund_dir, "2025-12", positions=positions)

    wb = load_workbook(workbook)
    ws = wb["Sources"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers[field], value)
    wb.save(workbook)

    result = evaluate_inputs(fund_dir, "2025-12")
    assert code in {finding.code for finding in result["blockers"]}


def test_duplicate_derived_holding_key_requires_explicit_id(tmp_path, fdrs_dir):
    fund_dir = tmp_path / "fdrs"
    fund_dir.mkdir()
    shutil.copyfile(fdrs_dir / "fund_config.txt", fund_dir / "fund_config.txt")
    positions = tmp_path / "positions.csv"
    positions.write_text("ticker\nDUP\nDUP\n", encoding="utf-8")

    with pytest.raises(ValueError, match="add a unique holdingId"):
        prepare_inputs(fund_dir, "2025-12", positions=positions)


def test_not_applicable_review_clears_source_value(tmp_path):
    positions = tmp_path / "positions.csv"
    positions.write_text("holdingId,couponKind\nH-1,Fixed\n", encoding="utf-8")
    rows = [{
        "targetFile": "holdings.csv", "recordKey": "H-1",
        "fieldName": "couponKind", "currentValue": "Fixed",
        "proposedValue": "", "sourceId": "POSITIONS",
        "status": "NOT_APPLICABLE", "comment": "not a debt security",
    }]

    holdings, _ = _holding_objects(positions, rows, "POSITIONS")
    assert holdings[0].coupon_kind == ""


def test_bundle_publication_is_atomic_on_write_failure(
    tmp_path, factories, monkeypatch,
):
    fund_dir = tmp_path / "atomic"
    fund_dir.mkdir()
    monkeypatch.setattr(pipeline, "evaluate_inputs", lambda *_: _ready_result(tmp_path, factories))

    def fail_write(*_):
        raise OSError("simulated disk failure")

    monkeypatch.setattr(pipeline, "_write_holdings", fail_write)
    with pytest.raises(OSError, match="simulated disk failure"):
        finalize_inputs(fund_dir, "2026-07", bundle_root=tmp_path / "builds", run_id="run-1")

    version_dir = tmp_path / "builds" / "atomic" / "2026-07" / "nport-v1.13"
    assert not version_dir.exists() or list(version_dir.iterdir()) == []


def test_bundle_version_is_immutable(tmp_path, factories, monkeypatch):
    fund_dir = tmp_path / "immutable"
    fund_dir.mkdir()
    monkeypatch.setattr(pipeline, "evaluate_inputs", lambda *_: _ready_result(tmp_path, factories))
    bundle = finalize_inputs(
        fund_dir, "2026-07", bundle_root=tmp_path / "builds", run_id="run-1",
    )
    sentinel = (bundle / "input_receipt.json").read_bytes()

    with pytest.raises(FileExistsError, match="refusing to overwrite"):
        finalize_inputs(
            fund_dir, "2026-07", bundle_root=tmp_path / "builds", run_id="run-1",
        )
    assert (bundle / "input_receipt.json").read_bytes() == sentinel


def test_bundle_run_id_cannot_escape_version_directory(
    tmp_path, factories, monkeypatch,
):
    fund_dir = tmp_path / "safe-name"
    fund_dir.mkdir()
    monkeypatch.setattr(pipeline, "evaluate_inputs", lambda *_: _ready_result(tmp_path, factories))
    with pytest.raises(ValueError, match="run_id must contain only"):
        finalize_inputs(
            fund_dir, "2026-07", bundle_root=tmp_path / "builds", run_id="../escape",
        )
