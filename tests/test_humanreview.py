"""Human-review workbook — deterministic unit tests (no Bloomberg/network)."""
from nport import humanreview
from nport.humanreview import build_review_workbook, read_review


def test_reference_sheets_seeded_from_literals(tmp_path):
    """A fresh workbook seeds the reference sheets from the legacy maps."""
    p = tmp_path / "2026-06_review.xlsx"
    build_review_workbook(p, generated={})
    rev = read_review(p)
    # designated index, swap counterparties, option index all present from seed
    assert rev.swap_counterparties()["CANT"] == ("Cantor Fitzgerald & Co.", "5493004J7H4GCPG6OB62")
    assert rev.option_index()["SPY"] == ("S&P 500 Index", "SPX")
    di = rev.designated_index()
    assert di["FDRS"] == ("S&P 500 Total Return Index", "SPXT")
    assert di["CHYG"][1] == "CFIIH52C"


def test_empty_file_reads_as_empty_review(tmp_path):
    rev = read_review(tmp_path / "nope.xlsx")
    assert rev.designated_index() == {} and rev.swap_counterparties() == {}


def test_per_item_gap_sheet_and_gap_count(tmp_path):
    """A generated invCountry row with no value is a GAP; with a value it isn't."""
    p = tmp_path / "r.xlsx"
    gaps = build_review_workbook(p, generated={
        "invCountry": [
            {"fund": "AV", "ticker": "ASML", "cusip": "N/A", "name": "ASML", "invCountry": "",
             "sourceNote": "Bloomberg returned no country"},
            {"fund": "AV", "ticker": "AAPL", "cusip": "037833100", "name": "Apple", "invCountry": "US",
             "sourceNote": ""},
        ],
    })
    assert gaps["invCountry"] == 1                       # ASML blank → gap; AAPL filled → not
    rev = read_review(p)
    assert rev.inv_country() == {("AV", "037833100"): "US"}   # only the filled one


def test_operator_edits_preserved_on_rebuild(tmp_path):
    """Rebuilding keeps a value the operator filled into a previously-blank gap."""
    p = tmp_path / "r.xlsx"
    gen = {"invCountry": [
        {"fund": "AV", "ticker": "ASML", "cusip": "N/A", "name": "ASML", "invCountry": "", "sourceNote": ""},
    ]}
    build_review_workbook(p, generated=gen)
    # Operator fills it by editing the xlsx — simulate by writing the value then rebuilding
    import openpyxl
    wb = openpyxl.load_workbook(p)
    ws = wb["invCountry"]
    hdr = [c.value for c in ws[1]]
    ws.cell(row=2, column=hdr.index("invCountry") + 1, value="NL")
    wb.save(p)
    # Rebuild with the SAME generated gap → operator's "NL" must survive, gap cleared
    gaps = build_review_workbook(p, generated=gen)
    assert gaps["invCountry"] == 0
    assert read_review(p).inv_country() == {("AV", "N/A"): "NL"}


def test_swap_legs_roundtrip(tmp_path):
    p = tmp_path / "r.xlsx"
    row = {"fund": "CMAG", "swapTicker": "02079K305-TRS-05/31/27-L-CANT",
           "recFixedOrFloating": "Other", "recDesc": "Total return of ALPHABET INC.",
           "pmntFloatingRtIndex": "USD-SOFR", "pmntFloatingRtSpread": "",
           "pmntRateTenor": "Month", "pmntRateUnit": "3", "sourceNote": "seed: TRS template"}
    build_review_workbook(p, generated={"swap_legs": [row]})
    legs = read_review(p).swap_legs()
    key = ("CMAG", "02079K305-TRS-05/31/27-L-CANT")
    assert legs[key]["recDesc"] == "Total return of ALPHABET INC."
    assert legs[key]["pmntFloatingRtIndex"] == "USD-SOFR"


def test_seed_maps_match_legacy_values(tmp_path):
    """The relocated seeds equal the values that used to live in custodian/filing_master."""
    assert humanreview.SEED_SWAP_COUNTERPARTIES["MREX"][1] == "5493006BWPDUCYG6EQ34"
    assert humanreview.SEED_OPTION_INDEX["QQQ"] == ("NASDAQ 100 Index", "NDX")
    # XPAV is an SPX fund per the legacy benchmark map
    assert humanreview.SEED_DESIGNATED_INDEX["XPAV"] == ("S&P 500 Index", "SPX")


def test_central_fund_and_holding_inputs_roundtrip(tmp_path):
    p = tmp_path / "r.xlsx"
    generated = {
        "fund_policy": [{
            "fund": "FDRS", "fiscalYearEndMMDD": "06-30",
            "derivativesRegimePolicy": "VAR_RELATIVE", "liquidityRequired": "Y",
            "cashB2fRequired": "Y", "policyEffectiveFrom": "2026-01-01",
            "policyEffectiveTo": "", "sourceNote": "",
        }],
        "fund_risk": [{"fund": "FDRS", "cashNotReportedInCOrD": "",
                       "medianDailyVarPct": "", "medianVarRatioPct": "",
                       "backtestingExceptions": "", "sourceNote": ""}],
        "option_delta": [{"fund": "FDRS", "ticker": "SPY-OPT", "delta": "0.42",
                          "sourceNote": ""}],
        "holding_liquidity": [{"fund": "FDRS", "ticker": "SPY", "cusip": "78462F103",
                               "name": "SPDR", "liquidityClassificationJson": "",
                               "liquidityCircumstancesJson": "", "sourceNote": ""}],
    }
    gaps = build_review_workbook(p, generated)
    assert gaps["fund_policy"] == 0
    assert gaps["fund_risk"] == 1
    assert gaps["holding_liquidity"] == 1
    review = read_review(p)
    assert review.option_delta()[("FDRS", "SPY-OPT")] == "0.42"
    assert review.fund_policy()["FDRS"]["derivativesRegimePolicy"] == "VAR_RELATIVE"


def test_apply_fund_policy_to_config_preserves_unrelated_lines(tmp_path):
    config = tmp_path / "funds" / "fdrs" / "fund_config.txt"
    config.parent.mkdir(parents=True)
    config.write_text("# Identity\ncik=123\nfiscalYearEndMMDD=12-31\n", encoding="utf-8")
    workbook = tmp_path / "r.xlsx"
    build_review_workbook(workbook, {"fund_policy": [{
        "fund": "FDRS", "fiscalYearEndMMDD": "06-30",
        "derivativesRegimePolicy": "NONE", "liquidityRequired": "N",
        "cashB2fRequired": "N", "policyEffectiveFrom": "2026-01-01",
        "policyEffectiveTo": "", "sourceNote": "",
    }]})
    assert humanreview.apply_fund_policy_to_configs(tmp_path / "funds", read_review(workbook)) == 1
    text = config.read_text(encoding="utf-8")
    assert "cik=123" in text and "fiscalYearEndMMDD=06-30" in text
    assert "derivativesRegimePolicy=NONE" in text


def test_reconciliation_rows_are_visible_but_not_editable_gaps(tmp_path):
    source = tmp_path / "reconciliation.csv"
    source.write_text(
        "check,fund,source_a,value_a,source_b,value_b,diff,flag\n"
        "flow,FDRS,EagleSTAR,1,Orders,2,-1,REVIEW\n"
        "nav,FDRS,Custody,3,EagleSTAR,3,0,PASS\n", encoding="utf-8",
    )
    rows = humanreview.generate_reconciliation_rows(source)
    assert len(rows) == 1 and rows[0]["flag"] == "REVIEW"
    workbook = tmp_path / "r.xlsx"
    gaps = build_review_workbook(workbook, {"reconciliation": rows})
    assert gaps["reconciliation"] == 0
