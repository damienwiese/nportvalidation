"""Fund-period inputs, provenance, and clean canonical finalization.

This module creates places for missing values; it never supplies business data.
Every supplied value must cite an independent, non-U.S.-Bank source.
"""

from __future__ import annotations

import csv
import json
import os
import tempfile
from collections import Counter
from dataclasses import asdict, fields, replace
from decimal import Decimal, InvalidOperation
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from nport.ap_orders import aggregate_flows, parse_ap_orders
from nport.config import (
    _CONFIG_KEY_MAP, _FILING_KEY_MAP, _HOLDINGS_KEY_MAP,
    _OPTIONAL_FILING_KEYS, parse_config,
)
from nport.input_validation import validate_all
from nport.models import FilingData, FundConfig, Holding
from nport.filing_master import _month_ranges, _return_formula
from nport.master_sheet import (
    BLOOMBERG_SPECS, _bbg_spec_key, _bloomberg_formula,
    _normalize_coupon_kind, _normalize_date,
)
from nport.policy import apply_context, derive_context_from_config, report_date_for_period
from nport.preflight import Finding, _is_prohibited_source, sha256_file
from nport.schema import FIELD_SPECS, get_required_fields


INPUT_FILENAME = "filing_inputs.xlsx"
LEGACY_REVIEW_FILENAME = "human_review.xlsx"

FIELD_HEADERS = [
    "gapId", "targetFile", "recordKey", "fieldName", "currentValue",
    "proposedValue", "sourceId", "status", "comment",
]
SOURCE_HEADERS = [
    "sourceId", "dataset", "sourceType", "sourceSystem", "sourcePath",
    "sourceAsOf", "sha256", "recordCount", "comment",
]
BLOOMBERG_HEADERS = [
    "targetFile", "recordKey", "fieldName", "bbgid", "mnemonic", "value",
]
RECON_INPUT_HEADERS = [
    "checkId", "category", "actualBasis", "controlValue", "tolerance",
    "sourceId", "status", "comment",
]

FLOW_RECON_FIELDS = (
    "mon1Sales", "mon1Redemption", "mon1Reinvestment",
    "mon2Sales", "mon2Redemption", "mon2Reinvestment",
    "mon3Sales", "mon3Redemption", "mon3Reinvestment",
)

POLICY_FIELDS = {
    "fiscalYearEndMMDD": "fiscal_year_end_mmdd",
    "derivativesRegimePolicy": "derivatives_regime_policy",
    "liquidityRequired": "liquidity_required",
    "cashB2fRequired": "cash_b2f_required",
    "policyEffectiveFrom": "policy_effective_from",
    "policyEffectiveTo": "policy_effective_to",
    "policySourceRef": "policy_source_ref",
}

SYSTEM_DERIVED_FILING = {
    "submissionType", "repPdEnd", "repPdDate", "derivativesRegime",
}

SENSITIVE_HOLDING_GROUPS = {"debt", "deriv_common", "option", "ref_instrument", "swap", "liquidity"}

GAP_GUIDE_HEADERS = [
    "gapId", "category", "gap", "inputSheet", "rowToFind",
    "columnsToComplete", "sourceRow", "completionRule", "evidenceToRegister",
]

GAP_GUIDE = [
    ("G-001", "OPEN INPUT", "Position population", "CLI + HoldingFields",
     "Run prepare with --positions; then filter HoldingFields[gapId starts G-001]",
     "Holding exceptions: proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "The POSITIONS row is generated only so build can reload the file; no source evidence is required",
     "The position file loads, its imported row count ties, and no G-001 exception remains MISSING",
     "Supply the independent position export to --positions"),
    ("G-002", "OPEN INPUT", "Fund-accounting close values", "FundFields",
     "FundFields[gapId starts G-002]",
     "proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "sourceId is optional trace metadata and does not affect XML readiness",
     "Every applicable G-002 value is present and the NAV equation passes",
     "Enter the final internal accounting values in the named rows"),
    ("G-003", "OPEN INPUT / PARTLY AUTOMATED", "Create/redeem flows", "CLI + FundFields",
     "Pass --orders to prepare; then filter FundFields[gapId starts G-003]",
     "Only unresolved rows: proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "Automated order totals are prefilled; sourceId is optional",
     "Every applicable flow value is resolved",
     "Supply --orders and enter any remaining reinvestment/flow value in FundFields"),
    ("G-004", "OPEN INPUT", "Fund policy and static data", "FundFields",
     "FundFields[targetFile=fund_config.txt]",
     "proposedValue and status; policyEffectiveTo may remain blank when open-ended",
     "policySourceRef is not requested because it does not affect XML",
     "Every XML-impacting policy/static value is resolved and policy validation passes",
     "Enter the applicable fund-level value in the named row"),
    ("G-005", "OPTIONAL TRACE", "Source manifest", "Sources",
     "Optional; no reviewer row is required",
     "None for XML release",
     "Build may emit source_manifest.csv for diagnostics",
     "Never blocks XML generation",
     "No remediation required"),
    ("G-006", "ENFORCED INPUT BOUNDARY", "Prohibited comparison inputs", "CLI + optional Sources",
     "Any supplied path or populated Sources row naming U.S. Bank, EagleSTAR, prepared filings, or legacy masters",
     "Remove the prohibited path/row; do not copy replacement values from comparison outputs",
     "No source metadata is otherwise required",
     "No prohibited comparison output is used as an input",
     "Use an independent in-house input file"),
    ("G-007", "OPTIONAL TRACE", "Manual input provenance", "FundFields + HoldingFields",
     "Optional sourceId on any resolved row",
     "None for XML release; a NOT_APPLICABLE row still requires a factual comment",
     "sourceId is carried only to field_provenance.csv",
     "Never blocks XML generation",
     "No remediation required"),
    ("G-008", "CONDITIONAL OPEN INPUT", "Swap economics", "HoldingFields",
     "HoldingFields[gapId starts G-008] for each swap recordKey",
     "proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "sourceId is optional",
     "Every required swap field is resolved and derivative validation passes",
     "Enter each missing swap value in the named holding row"),
    ("G-009", "CONDITIONAL OPEN INPUT", "B.2.f / B.3 / B.5.c / B.9 / B.10", "FundFields",
     "FundFields[gapId starts G-009]",
     "proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "sourceId is optional",
     "Each applicable XML value is PROVIDED; each inapplicable row has a factual disposition",
     "Enter each missing fund-level value in the named row"),
    ("G-010", "CONDITIONAL OPEN INPUT", "C.7 liquidity", "HoldingFields",
     "HoldingFields[gapId starts G-010] for each recordKey",
     "proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "sourceId is optional",
     "Every applicable holding has liquidityClassificationJson and required circumstances",
     "Enter each liquidity value in the named holding row"),
    ("G-011", "CONDITIONAL OPEN INPUT", "Option terms and delta", "HoldingFields",
     "HoldingFields[gapId starts G-011] for each option recordKey",
     "proposedValue and status; for NOT_APPLICABLE also enter a factual comment",
     "sourceId is optional",
     "Every required option/reference field is resolved and option validation passes",
     "Enter each option/risk value in the named holding row"),
    ("G-012", "RECONCILIATION GATE", "Accuracy reconciliation", "ReconciliationInputs; results in reconciliation.csv",
     "ReconciliationInputs[checkId=POSITIONS_TO_GL/FLOW:<field>/DERIVATIVE_* when generated]",
     "controlValue, tolerance, status, and comment; never type the filed-side actual value",
     "sourceId is optional and does not affect the calculation",
     "Every check is resolved and the calculated difference is within tolerance; NAV and count also PASS",
     "Enter each control total and tolerance in ReconciliationInputs"),
    ("G-013", "OUT OF ACTIVE SCOPE", "Authenticated origin assurance", "No workbook entry",
     "Source metadata cannot authenticate origin",
     "None",
     "An authenticated connector would be a separate control",
     "Does not gate this workflow",
     "No remediation required"),
]


class ReviewBlocked(ValueError):
    """Compatibility exception raised when inputs are not ready for release."""

    def __init__(self, blockers: list[Finding]):
        self.blockers = blockers
        super().__init__(f"input workbook has {len(blockers)} open input/error item(s)")


def input_path(fund_dir: str | Path, period: str, *, existing: bool = False) -> Path:
    destination = Path(fund_dir) / "filings" / period / INPUT_FILENAME
    if existing and not destination.is_file():
        legacy = destination.with_name(LEGACY_REVIEW_FILENAME)
        if legacy.is_file():
            return legacy
    return destination


def review_path(fund_dir: str | Path, period: str) -> Path:
    """Compatibility alias for callers that still use the old function name."""
    return input_path(fund_dir, period, existing=True)


def _hash(path: Path) -> str:
    return sha256_file(path)


def _read_sheet(path: Path, sheet: str, *, data_only: bool = True) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    wb = load_workbook(path, data_only=data_only, read_only=True)
    if sheet not in wb.sheetnames:
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        if not raw or all(value is None for value in raw):
            continue
        out.append({
            headers[i]: "" if raw[i] is None else str(raw[i]).strip()
            for i in range(min(len(headers), len(raw))) if headers[i]
        })
    return out


def _field_key(row: dict[str, str]) -> tuple[str, str, str]:
    return row.get("targetFile", ""), row.get("recordKey", ""), row.get("fieldName", "")


def _merge_preserved(
    generated: list[dict[str, str]], existing: list[dict[str, str]],
    key_fn, *, legacy: bool = False,
) -> list[dict[str, str]]:
    prior = {key_fn(row): row for row in existing}
    editable = {
        "proposedValue", "sourceId", "sourceAsOf", "status", "comment",
        "dataset", "sourceType", "sourceSystem", "sourcePath", "recordCount",
        "controlValue", "tolerance",
    }
    out = []
    for row in generated:
        merged = dict(row)
        old = prior.get(key_fn(row))
        if old:
            for key in editable:
                if old.get(key, ""):
                    value = old[key]
                    if key == "status":
                        normalized = value.strip().upper()
                        if legacy and normalized not in {"SYSTEM_DERIVED", "SYSTEM_CONTROL"}:
                            # Preserve the old value for reference, but never carry an old
                            # acceptance decision into the independent workflow.
                            value = "MISSING"
                        else:
                            value = {
                                "APPROVED": "PROVIDED",
                                "NEEDS_REVIEW": "MISSING",
                                "PENDING": "MISSING",
                            }.get(normalized, value)
                    merged[key] = value
        out.append(merged)
    for key, old in prior.items():
        if key not in {key_fn(row) for row in generated}:
            out.append(old)
    return out


def _write_table(
    ws, headers: list[str], rows: list[dict[str, str]], *,
    editable_headers: set[str] | None = None,
) -> None:
    """Write a clean review table and visibly distinguish operator input columns."""
    editable_headers = editable_headers or set()
    thin = Side(style="thin", color="D4DDE6")
    ws.append(headers)
    for header, cell in zip(headers, ws[1]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            "solid", fgColor="356F78" if header in editable_headers else "28465C"
        )
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 85
    ws.row_dimensions[1].height = 30
    for column in ws.columns:
        letter = column[0].column_letter
        width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        ws.column_dimensions[letter].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def _add_list_validation(ws, header: str, values: list[str]) -> None:
    """Add an Excel dropdown to a named column without supplying a decision."""
    headers = {cell.value: cell.column for cell in ws[1]}
    column = headers[header]
    letter = ws.cell(1, column).column_letter
    validation = DataValidation(
        type="list", formula1='"' + ",".join(values) + '"', allow_blank=False,
    )
    validation.error = "Choose one of the controlled workflow statuses."
    validation.errorTitle = "Invalid status"
    validation.prompt = "Select the status that matches the supplied evidence."
    validation.promptTitle = "Controlled status"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{max(2, ws.max_row)}")


def _gap_row(
    gap_id: str, target: str, key: str, field: str, current: str = "",
    status: str = "MISSING", comment: str = "",
) -> dict[str, str]:
    return {
        "gapId": gap_id, "targetFile": target, "recordKey": key,
        "fieldName": field, "currentValue": current, "proposedValue": "",
        "sourceId": "", "status": status, "comment": comment,
    }


def _config_rows(config: FundConfig) -> list[dict[str, str]]:
    rows = []
    for external, attr in _CONFIG_KEY_MAP.items():
        if external in {"requiredSources", "policyApprovedBy", "policyApprovedAt", "policySourceRef"}:
            # Internal trace/approval metadata does not affect the XML payload.
            continue
        value = getattr(config, attr, "")
        is_policy = external in POLICY_FIELDS
        rows.append(_gap_row(
            f"G-004:{external}", "fund_config.txt", "FUND", external, value,
            "MISSING",
            ("Confirm from an independent internal policy/governance record."
             if is_policy else "Confirm from independent internal legal/entity/static-data evidence."),
        ))
    return rows


def _filing_rows() -> list[dict[str, str]]:
    rows = []
    for external, attr in _FILING_KEY_MAP.items():
        if external in SYSTEM_DERIVED_FILING:
            rows.append(_gap_row(
                f"SYSTEM:{external}", "filing_data.txt", "FUND", external,
                "", "SYSTEM_DERIVED", "Derived from period and approved fund policy.",
            ))
            continue
        current = "TEST" if external == "liveTestFlag" else ""
        status = "SYSTEM_CONTROL" if external == "liveTestFlag" else "MISSING"
        gap = "G-002"
        if external.startswith("mon"):
            gap = "G-003"
        elif external in {"cashNotReportedInCOrD", "curMetricsJson", "creditSprdRiskIgJson",
                          "creditSprdRiskNonigJson", "monthlyReturnCategoriesJson"} or \
                external.startswith(("deriv", "median", "backtesting")):
            gap = "G-009"
        rows.append(_gap_row(
            f"{gap}:{external}", "filing_data.txt", "FUND", external,
            current, status, "No default business value is supplied.",
        ))
    return rows


def _holding_ids(rows: list[dict[str, str]]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for index, row in enumerate(rows, 1):
        base = (row.get("holdingId") or row.get("ticker") or row.get("otherValue")
                or row.get("cusip") or f"row-{index}").strip()
        count = seen.get(base, 0)
        seen[base] = count + 1
        out.append(base if count == 0 else f"{base}-{count}")
    return out


def _raw_positions(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: (value or "").strip() for key, value in row.items()} for row in reader]
    return _holding_ids(rows), rows


def _bloomberg_rows(path: Path | None, ticker: str, period: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    fund_bbgid = f"{ticker.upper()} US Equity"
    for field, (start, end) in zip(("rtn1", "rtn2", "rtn3"), _month_ranges(period)):
        rows.append({
            "targetFile": "filing_data.txt", "recordKey": "FUND", "fieldName": field,
            "bbgid": fund_bbgid, "mnemonic": "CUST_TRR_RETURN_HOLDING_PER",
            "formulaFactory": lambda cell, s=start, e=end: _return_formula(cell, s, e),
        })
    if path is None:
        return rows
    ids, positions = _raw_positions(path)
    for hid, position in zip(ids, positions):
        spec_key = _bbg_spec_key(position)
        if spec_key is None:
            continue
        key_fn, fields = BLOOMBERG_SPECS[spec_key]
        bbgid = key_fn(position)
        if not bbgid.strip() or bbgid.strip().split()[0] in {"US", "Corp", "Govt", "Equity"}:
            continue
        for field, (mnemonic, kind) in fields.items():
            if position.get(field, "").strip():
                continue
            rows.append({
                "targetFile": "holdings.csv", "recordKey": hid, "fieldName": field,
                "bbgid": bbgid, "mnemonic": mnemonic,
                "formulaFactory": lambda cell, m=mnemonic, k=kind: _bloomberg_formula(m, cell, k),
            })
    return rows


def _holding_input_rows(
    path: Path | None, bloomberg_keys: set[tuple[str, str]],
) -> list[dict[str, str]]:
    if path is None:
        return []
    ids, positions = _raw_positions(path)
    spec_by_external = {spec.csv_header: spec for spec in FIELD_SPECS}
    rows: list[dict[str, str]] = []
    for hid, position in zip(ids, positions):
        deriv = position.get("derivCat", "")
        has_debt = bool(position.get("maturityDt")) or position.get("assetCat") in {
            "DBT", "ABS-MBS", "ABS-CBDO", "ABS-O",
        }
        required_internal = set(get_required_fields(deriv, has_debt))
        required_external = {
            next(spec.csv_header for spec in FIELD_SPECS if spec.name == name)
            for name in required_internal
        }
        for external, spec in spec_by_external.items():
            value = position.get(external, "")
            sensitive = spec.group in SENSITIVE_HOLDING_GROUPS and (
                value or external in required_external
            )
            bloomberg_field = (hid, external) in bloomberg_keys
            if not value and external not in required_external and not bloomberg_field:
                continue
            if value and not sensitive:
                continue
            gap = "G-001"
            if spec.group == "swap":
                gap = "G-008"
            elif spec.group == "option" or (spec.group == "ref_instrument" and deriv in {"OPT", "SWO", "WAR"}):
                gap = "G-011"
            elif spec.group == "liquidity":
                gap = "G-010"
            rows.append(_gap_row(
                f"{gap}:{hid}:{external}", "holdings.csv", hid, external, value,
                "MISSING",
                f"Holding field; requirement: {spec.required} {spec.condition}".strip(),
            ))
    return rows


def _reconciliation_input_rows(path: Path | None) -> list[dict[str, str]]:
    """Create stable control rows; values remain blank until independently supplied."""
    rows = [{
        "checkId": "POSITIONS_TO_GL", "category": "POSITION_GL",
        "actualBasis": "Sum holdings.csv valUSD",
        "controlValue": "", "tolerance": "", "sourceId": "", "status": "MISSING",
        "comment": "",
    }]
    for field in FLOW_RECON_FIELDS:
        rows.append({
            "checkId": f"FLOW:{field}", "category": "FLOW",
            "actualBasis": f"filing_data.txt {field}",
            "controlValue": "", "tolerance": "", "sourceId": "", "status": "MISSING",
            "comment": "",
        })
    if path and path.is_file():
        _ids, positions = _raw_positions(path)
        if any(row.get("derivCat", "") for row in positions):
            rows.extend((
                {
                    "checkId": "DERIVATIVE_MARKET_VALUE", "category": "DERIVATIVE",
                    "actualBasis": "Sum holdings.csv valUSD where derivCat is populated",
                    "controlValue": "", "tolerance": "", "sourceId": "", "status": "MISSING",
                    "comment": "",
                },
                {
                    "checkId": "DERIVATIVE_UNREALIZED", "category": "DERIVATIVE",
                    "actualBasis": "Sum holdings.csv unrealizedAppr where derivCat is populated",
                    "controlValue": "", "tolerance": "", "sourceId": "", "status": "MISSING",
                    "comment": "",
                },
            ))
    return rows


def prepare_inputs(
    fund_dir: str | Path, period: str, *, positions: str | Path | None = None,
    orders: str | Path | None = None,
) -> Path:
    """Create or refresh the fund-period input workbook without inventing data."""
    base = Path(fund_dir)
    config_path = base / "fund_config.txt"
    config = parse_config(config_path)
    existing = input_path(base, period, existing=True)
    legacy_input = existing.name == LEGACY_REVIEW_FILENAME
    destination = input_path(base, period)
    destination.parent.mkdir(parents=True, exist_ok=True)

    existing_fields = _read_sheet(existing, "FundFields", data_only=False)
    existing_holding = _read_sheet(existing, "HoldingFields", data_only=False)
    existing_sources = _read_sheet(existing, "Sources", data_only=False)
    existing_reconciliation = _read_sheet(existing, "ReconciliationInputs", data_only=False)

    position_path = Path(positions).resolve() if positions else None
    order_path = Path(orders).resolve() if orders else None
    bloomberg_rows = _bloomberg_rows(position_path, base.name, period)
    sources: list[dict[str, str]] = []
    for sid, dataset, stype, path in (
        ("POSITIONS", "internal_positions", "INDEPENDENT_POSITIONS", position_path),
        ("ORDERS", "internal_create_redeem", "CREATE_REDEEM", order_path),
    ):
        if path is None:
            continue
        record_count = ""
        if path.is_file() and path.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            with open(path, newline="", encoding="utf-8-sig") as handle:
                record_count = str(sum(1 for _ in csv.reader(handle, delimiter=delimiter)) - 1)
        sources.append({
            "sourceId": sid, "dataset": dataset, "sourceType": stype,
            "sourceSystem": "", "sourcePath": str(path), "sourceAsOf": "",
            "sha256": _hash(path) if path.is_file() else "",
            "recordCount": record_count, "comment": "",
        })
    if bloomberg_rows:
        sources.append({
            "sourceId": "BLOOMBERG", "dataset": "internal_bloomberg_reference",
            "sourceType": "BLOOMBERG_TERMINAL", "sourceSystem": "Bloomberg Excel Add-In",
            "sourcePath": "bloomberg://desktop", "sourceAsOf": report_date_for_period(period).isoformat(),
            "sha256": "", "recordCount": str(len(bloomberg_rows)),
            "comment": "Workbook formulas must be opened, calculated, and saved on a Bloomberg terminal.",
        })
    sources = _merge_preserved(sources, existing_sources, lambda row: row.get("sourceId", ""))
    for source in sources:
        source_path = Path(source.get("sourcePath", ""))
        if not source_path.is_file():
            continue
        actual_hash = _hash(source_path)
        recorded_hash = source.get("sha256", "").strip().lower()
        if not recorded_hash:
            source["sha256"] = actual_hash
        elif recorded_hash != actual_hash:
            source["sha256"] = actual_hash
            source["comment"] = "File changed; confirm the new input before building."
        if not source.get("recordCount", "") and source_path.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if source_path.suffix.lower() == ".tsv" else ","
            with open(source_path, newline="", encoding="utf-8-sig") as handle:
                source["recordCount"] = str(max(0, sum(1 for _ in csv.reader(handle, delimiter=delimiter)) - 1))

    generated_fund_rows = _config_rows(config) + _filing_rows()
    if order_path and order_path.is_file():
        flows = aggregate_flows(parse_ap_orders(order_path), period).get(base.name.upper(), {})
        for row in generated_fund_rows:
            field = row.get("fieldName", "")
            if field in flows and not field.endswith("Reinvestment"):
                row["currentValue"] = flows[field]
                row["sourceId"] = "ORDERS"
                row["status"] = "PROVIDED"
                row["comment"] = "Calculated from the supplied independent create/redeem orders."
    fund_rows = _merge_preserved(
        generated_fund_rows, existing_fields, _field_key, legacy=legacy_input,
    )
    bloomberg_keys = {
        (row["recordKey"], row["fieldName"])
        for row in bloomberg_rows if row["targetFile"] == "holdings.csv"
    }
    holding_rows = _merge_preserved(
        _holding_input_rows(position_path, bloomberg_keys), existing_holding, _field_key,
        legacy=legacy_input,
    )
    reconciliation_inputs = _merge_preserved(
        _reconciliation_input_rows(position_path), existing_reconciliation,
        lambda row: row.get("checkId", ""), legacy=legacy_input,
    )

    bloomberg_lookup = {
        (row["targetFile"], row["recordKey"], row["fieldName"]): index
        for index, row in enumerate(bloomberg_rows, 2)
    }
    for rows in (fund_rows, holding_rows):
        for row in rows:
            bbg_row = bloomberg_lookup.get(_field_key(row))
            if bbg_row and not row.get("proposedValue") and not row.get("sourceId"):
                value_cell = f"'Bloomberg'!F{bbg_row}"
                row["proposedValue"] = f"={value_cell}"
                row["sourceId"] = "BLOOMBERG"
                row["status"] = f'=IFERROR(IF({value_cell}="","MISSING","PROVIDED"),"MISSING")'
                row["comment"] = "Bloomberg formula; unresolved or error results remain missing."

    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("Summary")
    summary_rows = [
        ["FUND-PERIOD INPUT WORKFLOW", ""],
        ["Fund", base.name.upper()],
        ["Period", period],
        ["Current status", "NOT READY until the final build reports zero open inputs and validation errors"],
        ["Non-negotiable rule", "Do not use U.S. Bank, EagleSTAR, prepared filings, or values derived from them."],
        ["1 - Bloomberg", "Open this workbook on a Bloomberg terminal, wait for the Bloomberg sheet to calculate, and save."],
        ["2 - Filing fields", "Filter FundFields and HoldingFields to MISSING. Enter XML values or factual NOT_APPLICABLE decisions."],
        ["3 - Reconciliation", "Complete ReconciliationInputs controlValue, tolerance, status, and comment. sourceId is optional trace metadata."],
        ["4 - Build", f"Run: nport build {base.name.lower()} {period} --from-inputs"],
        ["Allowed statuses", "PROVIDED, NOT_APPLICABLE, MISSING, SYSTEM_DERIVED, or SYSTEM_CONTROL."],
        ["Sources sheet", "Optional provenance only. The reviewer does not complete it to make an XML field release-eligible."],
    ]
    for values in summary_rows:
        summary.append(values)
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 95
    summary.freeze_panes = "A2"
    summary.sheet_view.showGridLines = False
    summary.sheet_view.zoomScale = 90
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="28465C")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
    for row_cells in summary.iter_rows(min_row=2):
        row_cells[0].font = Font(bold=True, color="28465C")
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.auto_filter.ref = summary.dimensions
    bloomberg_ws = wb.create_sheet("Bloomberg")
    _write_table(
        bloomberg_ws, BLOOMBERG_HEADERS, bloomberg_rows,
    )
    for row_number, row in enumerate(bloomberg_rows, 2):
        bbgid_cell = f"$D{row_number}"
        bloomberg_ws.cell(row_number, 6, row["formulaFactory"](bbgid_cell))
    bloomberg_ws.sheet_properties.pageSetUpPr.fitToPage = True

    source_ws = wb.create_sheet("Sources")
    _write_table(
        source_ws, SOURCE_HEADERS, sources,
        editable_headers={"sourceId", "dataset", "sourceType", "sourceSystem", "sourcePath", "sourceAsOf", "comment"},
    )

    _write_table(
        wb.create_sheet("GapGuide"),
        GAP_GUIDE_HEADERS,
        [dict(zip(GAP_GUIDE_HEADERS, row)) for row in GAP_GUIDE],
    )
    field_inputs = {"proposedValue", "sourceId", "status", "comment"}
    fund_ws = wb.create_sheet("FundFields")
    _write_table(fund_ws, FIELD_HEADERS, fund_rows, editable_headers=field_inputs)
    _add_list_validation(
        fund_ws, "status",
        ["MISSING", "PROVIDED", "NOT_APPLICABLE", "SYSTEM_DERIVED", "SYSTEM_CONTROL"],
    )
    holding_ws = wb.create_sheet("HoldingFields")
    _write_table(holding_ws, FIELD_HEADERS, holding_rows, editable_headers=field_inputs)
    _add_list_validation(
        holding_ws, "status", ["MISSING", "PROVIDED", "NOT_APPLICABLE"],
    )
    recon_input_ws = wb.create_sheet("ReconciliationInputs")
    _write_table(
        recon_input_ws, RECON_INPUT_HEADERS, reconciliation_inputs,
        editable_headers={"controlValue", "tolerance", "sourceId", "status", "comment"},
    )
    _add_list_validation(
        recon_input_ws, "status", ["MISSING", "PROVIDED", "NOT_APPLICABLE"],
    )
    _write_table(wb.create_sheet("Reconciliation"),
                 ["check", "status", "actual", "expected", "detail"], [])

    fd, tmp = tempfile.mkstemp(dir=destination.parent, suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        Path(tmp).replace(destination)
    except Exception:
        Path(tmp).unlink(missing_ok=True)
        raise
    return destination


def prepare_review(
    fund_dir: str | Path, period: str, *, positions: str | Path | None = None,
    orders: str | Path | None = None,
) -> Path:
    """Compatibility alias for the 0.1.x review-oriented API."""
    return prepare_inputs(fund_dir, period, positions=positions, orders=orders)


def _resolved(row: dict[str, str]) -> str:
    return row.get("proposedValue", "").strip() or row.get("currentValue", "").strip()


def _provided(row: dict[str, str]) -> bool:
    return row.get("status", "").strip().upper() in {
        "PROVIDED", "APPROVED", "SYSTEM_DERIVED", "SYSTEM_CONTROL",
    }


def _source_map(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {row.get("sourceId", ""): row for row in rows if row.get("sourceId")}


def _fill_source_controls(sources: list[dict[str, str]]) -> None:
    """Calculate missing file controls in memory without rewriting the input workbook."""
    for source in sources:
        if source.get("sourceType", "").strip().upper() == "BLOOMBERG_TERMINAL":
            continue
        path = Path(source.get("sourcePath", ""))
        if not path.is_file():
            continue
        if not source.get("sha256", "").strip():
            source["sha256"] = _hash(path)
        if not source.get("recordCount", "").strip() and path.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            with open(path, newline="", encoding="utf-8-sig") as handle:
                source["recordCount"] = str(max(0, sum(1 for _ in csv.reader(handle, delimiter=delimiter)) - 1))


def _block(code: str, technical: str, plain: str) -> Finding:
    return Finding(code, "BLOCKER", technical, plain)


def _validate_source_boundary(sources: list[dict[str, str]]) -> list[Finding]:
    """Reject comparison outputs if explicitly registered; provenance is otherwise optional."""
    blockers: list[Finding] = []
    for index, source in enumerate(sources, start=2):
        values = [source.get(key, "") for key in ("dataset", "sourceType", "sourceSystem", "sourcePath")]
        if _is_prohibited_source(*values):
            sid = source.get("sourceId", "").strip()
            location = f"Sources row {index}" + (f" (sourceId={sid!r})" if sid else "")
            blockers.append(_block(
                "INPUT_SOURCE_PROHIBITED",
                f"{location} -> remove the comparison-output reference and rerun prepare with independent in-house inputs",
                "A comparison-only U.S. Bank/EagleSTAR/prepared-filing reference cannot be used as an input.",
            ))
    return blockers


def _apply_config_rows(config: FundConfig, rows: list[dict[str, str]]) -> FundConfig:
    values = {}
    for row in rows:
        external = row.get("fieldName", "")
        attr = _CONFIG_KEY_MAP.get(external)
        if attr and _provided(row):
            values[attr] = _resolved(row)
    return replace(config, **values)


def _filing_from_rows(rows: list[dict[str, str]], config: FundConfig, ticker: str, period: str) -> FilingData:
    values: dict[str, str] = {field.name: "" for field in fields(FilingData)}
    for row in rows:
        external = row.get("fieldName", "")
        attr = _FILING_KEY_MAP.get(external)
        if attr and _provided(row) and external not in SYSTEM_DERIVED_FILING:
            values[attr] = _resolved(row)
    context = derive_context_from_config(config, ticker, period, "fund_config.txt")
    return apply_context(FilingData(**values), context)


def _holding_objects(
    positions_path: Path, rows: list[dict[str, str]], default_source: str,
) -> tuple[list[Holding], list[dict[str, str]]]:
    ids, raw = _raw_positions(positions_path)
    overrides = {
        (row.get("recordKey", ""), row.get("fieldName", "")): _resolved(row)
        for row in rows if _provided(row)
    }
    external_fields = list(_HOLDINGS_KEY_MAP)
    provenance: list[dict[str, str]] = []
    holdings = []
    row_by_key = {(row.get("recordKey", ""), row.get("fieldName", "")): row for row in rows}
    for hid, source_row in zip(ids, raw):
        mapped: dict[str, str] = {}
        for external, attr in _HOLDINGS_KEY_MAP.items():
            value = overrides.get((hid, external), source_row.get(external, ""))
            if external == "couponKind":
                value = _normalize_coupon_kind(value)
            elif external in {"maturityDt", "expDt", "terminationDt"}:
                value = _normalize_date(value)
            mapped[attr] = value
            review_row = row_by_key.get((hid, external))
            provenance.append({
                "targetFile": "holdings.csv", "recordKey": hid, "fieldName": external,
                "value": value, "sourceId": (
                    review_row.get("sourceId", "") if review_row and _provided(review_row)
                    else default_source
                ),
                "method": (
                    "BLOOMBERG_FORMULA" if review_row and _provided(review_row)
                    and review_row.get("sourceId") == "BLOOMBERG"
                    else "MANUAL_INPUT" if review_row and _provided(review_row)
                    else "SOURCE_FILE"
                ),
            })
        holdings.append(Holding(**mapped))
    return holdings, provenance


def _field_blockers(rows: list[dict[str, str]], *, optional: set[str] | None = None) -> list[Finding]:
    optional = optional or set()
    blockers = []
    for row in rows:
        field = row.get("fieldName", "")
        status = row.get("status", "").upper()
        if status in {"SYSTEM_DERIVED", "SYSTEM_CONTROL"}:
            continue
        if field in optional and not _resolved(row):
            continue
        if status == "NOT_APPLICABLE":
            if not row.get("comment", ""):
                blockers.append(_block(
                    "INPUT_NA_EVIDENCE",
                    f"{('FundFields' if row.get('targetFile') != 'holdings.csv' else 'HoldingFields')}"
                    f"[recordKey={row.get('recordKey')!r}, fieldName={field!r}].comment -> enter the factual not-applicable reason",
                    f"{field} is marked not applicable but lacks a factual reason.",
                ))
            continue
        if (row.get("targetFile") == "fund_config.txt" and _provided(row)
                and not row.get("proposedValue", "").strip()):
            blockers.append(_block(
                "INPUT_CONFIG_VALUE",
                f"FundFields[targetFile='fund_config.txt', recordKey='FUND', fieldName={field!r}].proposedValue -> enter the independently supported value",
                f"{field} must be entered from its independent source; the existing config value is reference-only.",
            ))
            continue
        if not _provided(row) or not _resolved(row):
            sheet = "HoldingFields" if row.get("targetFile") == "holdings.csv" else "FundFields"
            blockers.append(_block(
                row.get("gapId", "INPUT_FIELD"),
                f"{sheet}[recordKey={row.get('recordKey')!r}, fieldName={field!r}] -> complete proposedValue and status, or a factual NOT_APPLICABLE comment",
                f"{field} is an open input and needs a supported value or not-applicable disposition.",
            ))
    return blockers


def _positions_source(sources: list[dict[str, str]]) -> dict[str, str] | None:
    return next((row for row in sources if row.get("dataset") == "internal_positions"), None)


def _strict_decimal(value: str, label: str) -> Decimal:
    text = str(value or "").replace(",", "").strip()
    if not text or text.upper() == "N/A":
        raise ValueError(f"{label} is not numeric")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{label} is not numeric: {value!r}") from exc


def _evaluate_reconciliation_inputs(
    rows: list[dict[str, str]], filing: FilingData | None, holdings: list[Holding],
) -> tuple[list[dict[str, str]], list[Finding]]:
    """Calculate G-012 checks from filed values and independent control inputs."""
    results: list[dict[str, str]] = []
    findings: list[Finding] = []
    for row in rows:
        check_id = row.get("checkId", "")
        category = row.get("category", "")
        location = f"ReconciliationInputs[checkId={check_id!r}]"
        status = row.get("status", "").strip().upper()
        if status == "NOT_APPLICABLE":
            external = check_id.split(":", 1)[1] if check_id.startswith("FLOW:") else ""
            filing_value = getattr(filing, _FILING_KEY_MAP[external]) if filing and external else ""
            actual_is_na = str(filing_value).strip().upper() == "N/A"
            if category != "FLOW" or not actual_is_na or not row.get("comment", ""):
                findings.append(_block(
                    "G-012_INPUT", f"{location} -> use NOT_APPLICABLE only when the filed flow is N/A, and complete comment",
                    f"Reconciliation control {check_id} has an unsupported NOT_APPLICABLE disposition.",
                ))
            results.append({
                "check": check_id, "status": "NOT_APPLICABLE", "actual": "",
                "expected": "", "detail": row.get("comment", ""),
            })
            continue
        required = ("controlValue", "tolerance", "comment")
        missing = [key for key in required if not row.get(key, "").strip()]
        if status != "PROVIDED" or missing:
            findings.append(_block(
                "G-012_INPUT",
                f"{location} -> complete controlValue, tolerance, comment, and status=PROVIDED",
                f"Reconciliation control {check_id} is an open input.",
            ))
            continue
        try:
            control = _strict_decimal(row["controlValue"], f"{check_id}.controlValue")
            tolerance = _strict_decimal(row["tolerance"], f"{check_id}.tolerance")
            if tolerance < 0:
                raise ValueError(f"{check_id}.tolerance must be non-negative")
        except ValueError as exc:
            findings.append(_block(
                "G-012_INPUT", f"{location} -> correct controlValue/tolerance: {exc}",
                f"Reconciliation control {check_id} contains an invalid number.",
            ))
            continue

        actual_values: list[tuple[str, Decimal]] = []
        try:
            if check_id == "POSITIONS_TO_GL":
                actual_values.append((
                    "Holdings total to GL control",
                    sum((_strict_decimal(h.val_usd or "0", "holding.valUSD") for h in holdings), Decimal("0")),
                ))
            elif check_id.startswith("FLOW:"):
                external = check_id.split(":", 1)[1]
                if filing is None:
                    continue
                filing_value = getattr(filing, _FILING_KEY_MAP[external])
                if str(filing_value).strip().upper() == "N/A":
                    findings.append(_block(
                        "G-012_INPUT", f"{location} -> set status=NOT_APPLICABLE with a factual comment",
                        f"{external} is N/A and its reconciliation row needs the same disposition.",
                    ))
                    continue
                actual_values.append((
                    f"Filed {external} to independent flow control",
                    _strict_decimal(filing_value, external),
                ))
            elif check_id == "DERIVATIVE_MARKET_VALUE":
                derivative_values = [h for h in holdings if h.deriv_cat]
                actual_values.append((
                    "Derivative market value to independent ledger",
                    sum((_strict_decimal(h.val_usd or "0", "derivative.valUSD") for h in derivative_values), Decimal("0")),
                ))
            elif check_id == "DERIVATIVE_UNREALIZED":
                derivative_values = [h for h in holdings if h.deriv_cat]
                actual_values.append((
                    "Derivative unrealized to independent ledger",
                    sum((_strict_decimal(h.unrealized_appr or "0", "derivative.unrealizedAppr") for h in derivative_values), Decimal("0")),
                ))
            else:
                findings.append(_block(
                    "G-012_INPUT", f"{location} -> remove the unsupported checkId and rerun prepare",
                    f"Reconciliation control {check_id!r} is not recognized.",
                ))
                continue
        except (KeyError, ValueError) as exc:
            findings.append(_block(
                "G-012_INPUT", f"{location} -> correct the upstream filed/holding value: {exc}",
                f"Reconciliation control {check_id} cannot calculate its filed value.",
            ))
            continue

        for label, actual in actual_values:
            difference = actual - control
            result_status = "PASS" if abs(difference) <= tolerance else "FAIL"
            results.append({
                "check": label, "status": result_status, "actual": f"{actual:.2f}",
                "expected": f"{control:.2f}",
                "detail": f"difference={difference:.2f}; tolerance={tolerance:.2f}",
            })
            if result_status == "FAIL":
                findings.append(_block(
                    "G-012_RECON",
                    f"{location} -> investigate the independent difference; correct the upstream filed input or the control value, then rebuild",
                    f"{label} exceeds the supplied tolerance.",
                ))
    return results, findings


def evaluate_inputs(fund_dir: str | Path, period: str) -> dict:
    """Return resolved models, open items, provenance, and reconciliation without writing."""
    base = Path(fund_dir)
    ticker = base.name.upper()
    workbook = input_path(base, period, existing=True)
    config = parse_config(base / "fund_config.txt")
    blockers: list[Finding] = []
    if not workbook.is_file():
        blockers.append(_block("INPUT_WORKBOOK", f"Run nport prepare to create {workbook}",
                               "Run nport prepare to create the fund-period input workbook."))
        return {"blockers": blockers}
    fund_rows = _read_sheet(workbook, "FundFields")
    holding_rows = _read_sheet(workbook, "HoldingFields")
    reconciliation_input_rows = _read_sheet(workbook, "ReconciliationInputs")
    sources = _read_sheet(workbook, "Sources")
    _fill_source_controls(sources)

    # Optional/conditional filing sections still need an explicit disposition.
    # Only an open-ended policy date and relative-VaR identifiers that are
    # inapplicable under the selected regime may remain blank.
    blockers.extend(_field_blockers(
        fund_rows, optional={"policyEffectiveTo", "nameDesignatedIndex", "indexIdentifier"}
    ))
    blockers.extend(_field_blockers(holding_rows, optional={"isin", "ticker"}))

    used_sources = {
        row.get("sourceId", "") for row in fund_rows + holding_rows + reconciliation_input_rows
        if (_provided(row) or row.get("status", "").upper() == "NOT_APPLICABLE")
        and row.get("sourceId", "")
    }
    position_source = _positions_source(sources)
    if position_source:
        used_sources.add(position_source.get("sourceId", ""))
    else:
        blockers.append(_block("G-001", f"Run nport prepare {ticker.lower()} {period} --positions <independent.csv>",
                               "An independent position population has not been supplied."))
    if not reconciliation_input_rows:
        blockers.append(_block(
            "G-012_INPUT",
            f"Run nport prepare {ticker.lower()} {period} with the independent positions file to create ReconciliationInputs",
            "The reconciliation input sheet is missing or empty.",
        ))
    elif position_source and Path(position_source.get("sourcePath", "")).is_file():
        try:
            expected_ids = {
                row["checkId"] for row in _reconciliation_input_rows(Path(position_source["sourcePath"]))
            }
        except (KeyError, OSError, TypeError, ValueError):
            expected_ids = set()
        if expected_ids:
            actual_ids = [row.get("checkId", "") for row in reconciliation_input_rows]
            counts = Counter(actual_ids)
            missing_ids = sorted(expected_ids - set(actual_ids))
            duplicate_ids = sorted(check_id for check_id, count in counts.items() if check_id and count > 1)
            if missing_ids or duplicate_ids:
                detail = []
                if missing_ids:
                    detail.append(f"missing checkId(s): {', '.join(missing_ids)}")
                if duplicate_ids:
                    detail.append(f"duplicate checkId(s): {', '.join(duplicate_ids)}")
                blockers.append(_block(
                    "G-012_INPUT",
                    f"Run nport prepare {ticker.lower()} {period} again to restore ReconciliationInputs ({'; '.join(detail)})",
                    "The reconciliation input rows do not match the required checks for this filing.",
                ))
    blockers.extend(_validate_source_boundary(sources))

    try:
        config = _apply_config_rows(config, fund_rows)
        source_by_id = _source_map(sources)
        used_datasets = sorted({
            source_by_id[sid].get("dataset", "") for sid in used_sources
            if sid in source_by_id and source_by_id[sid].get("dataset", "")
        })
        config = replace(config, required_sources=";".join(used_datasets))
        filing = _filing_from_rows(fund_rows, config, ticker, period)
    except ValueError as exc:
        blockers.append(_block(
            "G-004",
            f"FundFields[targetFile='fund_config.txt'] -> correct the policy/config row named by this error: {exc}",
            "The fund policy is incomplete or invalid.",
        ))
        filing = None

    holdings: list[Holding] = []
    provenance: list[dict[str, str]] = []
    if position_source and Path(position_source.get("sourcePath", "")).is_file():
        try:
            holdings, provenance = _holding_objects(
                Path(position_source["sourcePath"]), holding_rows,
                position_source.get("sourceId", "POSITIONS"),
            )
        except (OSError, TypeError, ValueError) as exc:
            blockers.append(_block("G-001", f"Replace the --positions file referenced by Sources[sourceId='POSITIONS'].sourcePath: {exc}",
                                   "The independent position file is invalid."))

    reconciliation: list[dict[str, str]] = []
    if filing and holdings:
        errors, warnings = validate_all(config, filing, holdings)
        for message in errors:
            blockers.append(_block(
                "INPUT_VALIDATION",
                f"FundFields or HoldingFields -> locate the field named in this validation message and correct its proposedValue/status: {message}",
                "A supplied canonical filing value is invalid.",
            ))
        try:
            equation = float(filing.tot_assets) - float(filing.tot_liabs) - float(filing.net_assets)
            status = "PASS" if abs(equation) <= 1.0 else "FAIL"
            reconciliation.append({
                "check": "NAV equation", "status": status, "actual": f"{equation:.2f}",
                "expected": "0.00", "detail": "totAssets - totLiabs - netAssets",
            })
            if status == "FAIL":
                blockers.append(_block("G-012_NAV", "FundFields[fieldName='totAssets'/'totLiabs'/'netAssets'] -> correct the independently supported value(s)",
                                       "Assets, liabilities, and net assets do not reconcile."))
        except ValueError:
            pass
        reconciliation.append({
            "check": "Position import", "status": "PASS", "actual": str(len(holdings)),
            "expected": "Imported rows", "detail": "Holdings parsed from the --positions input",
        })
        supplied_reconciliation, reconciliation_findings = _evaluate_reconciliation_inputs(
            reconciliation_input_rows, filing, holdings,
        )
        reconciliation.extend(supplied_reconciliation)
        blockers.extend(reconciliation_findings)
        for row in fund_rows:
            if _provided(row) and row.get("fieldName") in _CONFIG_KEY_MAP:
                provenance.append({
                    "targetFile": "fund_config.txt", "recordKey": "FUND",
                    "fieldName": row.get("fieldName", ""), "value": _resolved(row),
                    "sourceId": row.get("sourceId", ""),
                    "method": "BLOOMBERG_FORMULA" if row.get("sourceId") == "BLOOMBERG" else "MANUAL_INPUT",
                })
            elif _provided(row) and row.get("fieldName") in _FILING_KEY_MAP:
                provenance.append({
                    "targetFile": "filing_data.txt", "recordKey": "FUND",
                    "fieldName": row.get("fieldName", ""), "value": _resolved(row),
                    "sourceId": row.get("sourceId", ""),
                    "method": "BLOOMBERG_FORMULA" if row.get("sourceId") == "BLOOMBERG" else "MANUAL_INPUT",
                })

    return {
        "workbook": workbook, "config": config, "filing": filing, "holdings": holdings,
        "sources": sources, "usedSources": used_sources, "blockers": blockers,
        "provenance": provenance, "reconciliation": reconciliation,
        "reconciliationInputs": reconciliation_input_rows,
    }


def evaluate_review(fund_dir: str | Path, period: str) -> dict:
    """Compatibility alias for the 0.1.x review-oriented API."""
    return evaluate_inputs(fund_dir, period)


def _write_kv(path: Path, mapping: dict[str, str]) -> None:
    path.write_text("".join(f"{key}={value}\n" for key, value in mapping.items()), encoding="utf-8")


def _write_holdings(path: Path, holdings: list[Holding]) -> None:
    field_to_external = {value: key for key, value in _HOLDINGS_KEY_MAP.items()}
    headers = [field_to_external[field.name] for field in fields(Holding)]
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for holding in holdings:
            raw = asdict(holding)
            writer.writerow({field_to_external[key]: value for key, value in raw.items()})


def finalize_inputs(
    fund_dir: str | Path, period: str, *, output_root: str | Path = "data/builds",
    run_id: str | None = None,
) -> Path:
    """Write a clean, versioned canonical bundle only when inputs are ready."""
    result = evaluate_inputs(fund_dir, period)
    if result["blockers"]:
        raise ReviewBlocked(result["blockers"])
    ticker = Path(fund_dir).name.upper()
    run_id = run_id or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    destination = Path(output_root) / ticker.lower() / period / run_id
    destination.mkdir(parents=True, exist_ok=False)

    config: FundConfig = result["config"]
    filing: FilingData = result["filing"]
    config_inverse = {value: key for key, value in _CONFIG_KEY_MAP.items()}
    filing_inverse = {value: key for key, value in _FILING_KEY_MAP.items()}
    _write_kv(destination / "fund_config.txt", {
        config_inverse[field.name]: str(getattr(config, field.name))
        for field in fields(FundConfig) if field.name in config_inverse
    })
    _write_kv(destination / "filing_data.txt", {
        filing_inverse[field.name]: str(getattr(filing, field.name))
        for field in fields(FilingData) if field.name in filing_inverse
    })
    _write_holdings(destination / "holdings.csv", result["holdings"])

    with open(destination / "source_manifest.csv", "w", newline="", encoding="utf-8") as handle:
        headers = ["dataset", "source_type", "source_system", "source_path", "as_of",
                   "sha256", "record_count"]
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for source in result["sources"]:
            if source.get("sourceId", "") in result["usedSources"]:
                writer.writerow({
                    "dataset": source.get("dataset", ""),
                    "source_type": source.get("sourceType", ""),
                    "source_system": source.get("sourceSystem", ""),
                    "source_path": source.get("sourcePath", ""),
                    "as_of": source.get("sourceAsOf", ""),
                    "sha256": source.get("sha256", ""),
                    "record_count": source.get("recordCount", ""),
                })

    for filename, rows, headers in (
        ("field_provenance.csv", result["provenance"],
         ["targetFile", "recordKey", "fieldName", "value", "sourceId", "method"]),
        ("reconciliation.csv", result["reconciliation"],
         ["check", "status", "actual", "expected", "detail"]),
    ):
        with open(destination / filename, "w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

    receipt = {
        "fund": ticker, "period": period, "runId": run_id,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "inputWorkbook": str(result["workbook"]),
        "inputWorkbookSha256": _hash(result["workbook"]),
        "blockers": [],
        "outputs": {},
        "originControlLimitation": (
            "Local hashes detect changed files; local code cannot prove that a person did "
            "not manually copy and relabel a prohibited value."
        ),
    }
    for name in ("fund_config.txt", "filing_data.txt", "holdings.csv",
                 "source_manifest.csv", "field_provenance.csv", "reconciliation.csv"):
        receipt["outputs"][name] = _hash(destination / name)
    (destination / "input_receipt.json").write_text(
        json.dumps(receipt, indent=2) + "\n", encoding="utf-8"
    )
    return destination


def finalize_review(
    fund_dir: str | Path, period: str, *, output_root: str | Path = "data/builds",
    run_id: str | None = None,
) -> Path:
    """Compatibility alias for the 0.1.x review-oriented API."""
    return finalize_inputs(fund_dir, period, output_root=output_root, run_id=run_id)
