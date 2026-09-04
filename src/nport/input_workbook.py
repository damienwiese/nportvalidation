"""Versioned, multi-fund Excel input workbook for the supported pipeline.

The workbook is a data-entry surface only.  Importing one fund-period creates a
normal reviewed ``filing_inputs.xlsx`` and immutable CSV extracts, so every
downstream validation and release gate remains unchanged.
"""

from __future__ import annotations

import csv
import hashlib
import io
import os
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from nport.config import CONFIG_KEY_MAP, FILING_KEY_MAP, HOLDINGS_KEY_MAP, parse_config
from nport.constants import NPORT_SCHEMA_VERSION
from nport.pipeline import SYSTEM_DERIVED_FILING, prepare_inputs
from nport.policy import report_date_for_period
from nport.schema import DEBT_ASSET_CATEGORIES, FIELD_SPECS, get_required_fields
from nport.workbook_support import (
    BLOOMBERG_SPECS,
    bloomberg_formula,
    bloomberg_spec_key,
    month_ranges,
    return_formula,
)

INPUT_WORKBOOK_FORMAT_VERSION = "4"
INPUT_WORKBOOK_SOURCE_ID = "INPUT_WORKBOOK"

FUND_UNIVERSE_FILENAME = "fund_universe.csv"
CONFIG_SHEET_SUFFIX = "_Config"
FUND_DATA_SHEET_SUFFIX = "_FundData"
POSITIONS_SHEET_SUFFIX = "_Positions"
ORDERS_SHEET = "Orders"

CONFIG_EXCLUDED_FIELDS = frozenset({
    "requiredSources", "policyApprovedBy", "policyApprovedAt", "policySourceRef",
})
CONFIG_INPUT_FIELDS = tuple(
    field for field in CONFIG_KEY_MAP if field not in CONFIG_EXCLUDED_FIELDS
)
FILING_INPUT_FIELDS = tuple(
    field
    for field in FILING_KEY_MAP
    if field not in SYSTEM_DERIVED_FILING and field != "liveTestFlag"
)
POSITION_FIELDS = tuple(spec.csv_header for spec in FIELD_SPECS)
POSITION_BLOOMBERG_FIELDS = frozenset(
    field
    for _key_function, fields in BLOOMBERG_SPECS.values()
    for field in fields
)

CONFIG_HEADERS = ("fieldName", "value")
FUND_DATA_HEADERS = ("Period", "fieldName", "value")
POSITION_HEADERS = ("Period", "holdingId", *POSITION_FIELDS)
ORDER_HEADERS = (
    "Fund", "Period", "Ticker", "Side", "Trade Date", "Notional", "Status",
)

_NAVY = "17365D"
_BLUE = "0000FF"
_WHITE = "FFFFFF"
_PALE_BLUE = "DCE6F1"
_PALE_YELLOW = "FFF2CC"
_GRAY = "E7E6E6"
_DARK_GRAY = "666666"
_THIN_GRAY = Side(style="thin", color="D9E1F2")

_LISTS = {
    "Funds": [],
    "Periods": [],
    "YN": ["Y", "N"],
    "OrderSides": ["CREATE", "REDEEM"],
    "OrderStatuses": ["ACCEPTED", "CANCELLED"],
    "Units": ["NS", "PA", "NC", "OU"],
    "PayoffProfiles": ["Long", "Short", "N/A"],
    "FairValueLevels": ["1", "2", "3", "N/A"],
    "AssetCategories": [
        "STIV", "RA", "EC", "EP", "DBT", "DCO", "DCR", "DE", "DFE",
        "DIR", "DO", "SN", "LON", "ABS-MBS", "ABS-APCP", "ABS-CBDO",
        "ABS-O", "COMM", "RE", "OTHER",
    ],
    "IssuerCategories": ["CORP", "UST", "USGA", "USGSE", "MUN", "NUSS", "PF", "RF", "OTHER"],
    "DerivativeCategories": ["FWD", "FUT", "SWP", "OPT", "SWO", "WAR", "OTH"],
    "CouponKinds": ["Fixed", "Floating", "Variable", "None"],
    "PutCall": ["Put", "Call"],
    "WrittenPurchased": ["Written", "Purchased"],
    "FixedFloating": ["Fixed", "Floating", "Other"],
    "ReferenceTypes": ["indexBasket", "otherRefInst"],
    "TenorUnits": ["Day", "Month", "Year"],
    "DerivativeRegimes": ["NONE", "LIMITED", "VAR_RELATIVE", "VAR_ABSOLUTE"],
}


def _periods(periods: list[str] | tuple[str, ...]) -> tuple[str, ...]:
    normalized = tuple(dict.fromkeys(str(period).strip() for period in periods))
    if not normalized:
        raise ValueError("at least one reporting period is required")
    for period in normalized:
        report_date_for_period(period)
    return normalized


def _sheet_name(fund: str, suffix: str) -> str:
    name = f"{fund}{suffix}"
    if len(name) > 31:
        raise ValueError(f"fund worksheet name exceeds Excel's 31-character limit: {name}")
    return name


def _expected_sheet_names(funds: list[str]) -> list[str]:
    return [
        name
        for fund in funds
        for name in (
            _sheet_name(fund, CONFIG_SHEET_SUFFIX),
            _sheet_name(fund, FUND_DATA_SHEET_SUFFIX),
            _sheet_name(fund, POSITIONS_SHEET_SUFFIX),
        )
    ] + [ORDERS_SHEET]


def _fund_entries(funds_dir: Path) -> list[tuple[str, object | None, str]]:
    config_paths = {
        path.parent.name.upper(): path
        for path in sorted(funds_dir.glob("*/fund_config.txt"))
    }
    universe_path = funds_dir / FUND_UNIVERSE_FILENAME
    registered: list[tuple[str, str]] = []
    if universe_path.is_file():
        with universe_path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames != ["fund", "category"]:
                raise ValueError(f"{universe_path}: expected headers fund,category")
            for row_number, row in enumerate(reader, 2):
                fund = str(row.get("fund") or "").strip().upper()
                category = str(row.get("category") or "").strip()
                if not fund or not fund.isascii() or not fund.isalnum():
                    raise ValueError(f"{universe_path}:{row_number}: invalid fund ticker {fund!r}")
                if category not in {
                    "Thematic", "2x Leveraged", "Structured Buffer", "Fixed Income",
                }:
                    raise ValueError(
                        f"{universe_path}:{row_number}: invalid category {category!r}"
                    )
                registered.append((fund, category))
    else:
        registered = [(fund, "Thematic") for fund in sorted(config_paths)]
    funds = [fund for fund, _category in registered]
    if not funds:
        raise ValueError(f"no funds found under {funds_dir}")
    duplicates = sorted({fund for fund in funds if funds.count(fund) > 1})
    if duplicates:
        raise ValueError(f"{universe_path}: duplicate funds: {','.join(duplicates)}")
    unregistered_configs = sorted(set(config_paths) - set(funds))
    if unregistered_configs:
        raise ValueError(
            f"active fund configs missing from {universe_path}: "
            + ",".join(unregistered_configs)
        )
    return [
        (
            fund,
            parse_config(config_paths[fund]) if fund in config_paths else None,
            category,
        )
        for fund, category in registered
    ]


def _append_rows(sheet, headers: tuple[str, ...], rows: list[dict[str, object]]) -> None:
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _style_table(sheet, *, freeze: str = "A2", editable_headers: set[str] | None = None) -> None:
    editable_headers = editable_headers or set()
    has_editable_columns = bool(editable_headers)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=_NAVY))
    header_by_column = {cell.column: str(cell.value or "") for cell in sheet[1]}
    for column, header in header_by_column.items():
        letter = sheet.cell(1, column).column_letter
        width = 14
        if header in {"fieldName", "value"}:
            width = 28
        elif header in {"Fund", "Period", "Status", "holdingId"}:
            width = 18
        sheet.column_dimensions[letter].width = width
        if header in editable_headers and sheet.max_row >= 2:
            rng = sheet[f"{letter}2:{letter}{sheet.max_row}"]
            for cell_tuple in rng:
                cell = cell_tuple[0]
                cell.font = Font(name="Aptos", size=10, color=_BLUE)
                cell.fill = PatternFill("solid", fgColor=_PALE_YELLOW)
        elif has_editable_columns and sheet.max_row >= 2:
            rng = sheet[f"{letter}2:{letter}{sheet.max_row}"]
            for cell_tuple in rng:
                cell = cell_tuple[0]
                cell.font = Font(name="Aptos", size=10, color=_DARK_GRAY)
                cell.fill = PatternFill("solid", fgColor=_GRAY)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=_THIN_GRAY)
    sheet.auto_filter.ref = sheet.dimensions


def _style_bloomberg_cells(sheet, coordinates: list[str]) -> None:
    for coordinate in coordinates:
        cell = sheet[coordinate]
        cell.font = Font(name="Aptos", size=10, color=_BLUE)
        cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
        cell.comment = Comment(
            "Generated Bloomberg formula. Enter the lookup fields, open on a "
            "Bloomberg-enabled Excel workstation, calculate, and save.",
            "N-PORT",
        )


def _excel_ref(headers: dict[str, int], field: str, row_number: int) -> str:
    return f"${get_column_letter(headers[field])}{row_number}"


def _position_bloomberg_formula(
    row_number: int, field: str, headers: dict[str, int],
) -> str | None:
    if field not in POSITION_BLOOMBERG_FIELDS:
        return None
    asset = _excel_ref(headers, "assetCat", row_number)
    issuer = _excel_ref(headers, "issuerCat", row_number)
    ticker = _excel_ref(headers, "ticker", row_number)
    cusip = _excel_ref(headers, "cusip", row_number)
    derivative = _excel_ref(headers, "derivCat", row_number)
    reference_cusip = _excel_ref(headers, "refCusip", row_number)
    selectors = {
        "EC": (f'AND({asset}="EC",{ticker}<>"")', f'{ticker}&" US Equity"'),
        "STIV": (f'AND({asset}="STIV",{ticker}<>"")', f'{ticker}&" US Equity"'),
        "DBT_UST": (
            f'AND({asset}="DBT",{issuer}="UST",{cusip}<>"")',
            f'{cusip}&" Govt"',
        ),
        "DBT_CORP": (
            f'AND({asset}="DBT",{issuer}<>"",{issuer}<>"UST",{cusip}<>"")',
            f'{cusip}&" Corp"',
        ),
        "SWP_REF": (
            f'AND({derivative}="SWP",{reference_cusip}<>"")',
            f'{reference_cusip}&" Equity"',
        ),
    }
    branches: list[tuple[str, str]] = []
    for spec_key, (_key_function, fields) in BLOOMBERG_SPECS.items():
        spec = fields.get(field)
        if spec is None:
            continue
        mnemonic, kind = spec
        condition, key_expression = selectors[spec_key]
        branches.append((
            condition,
            bloomberg_formula(mnemonic, key_expression, kind).removeprefix("="),
        ))
    expression = '""'
    for condition, result in reversed(branches):
        expression = f"IF({condition},{result},{expression})"
    return f'=IFERROR({expression},"")'


def _fund_bloomberg_formula(
    sheet, row_number: int, field: str, headers: dict[str, int], fund: str,
) -> str | None:
    if field not in {"rtn1", "rtn2", "rtn3"}:
        return None
    period = str(sheet.cell(row_number, headers["Period"]).value or "")
    try:
        month_index = ("rtn1", "rtn2", "rtn3").index(field)
        start, end = month_ranges(period)[month_index]
    except (ValueError, IndexError):
        return None
    result = return_formula(f'"{fund} US Equity"', start, end).removeprefix("=")
    return f'=IFERROR({result},"")'


def _seed_bloomberg_formulas(
    fund: str, fund_sheet, positions_sheet,
) -> tuple[list[str], list[str]]:
    fund_headers = {str(cell.value): cell.column for cell in fund_sheet[1]}
    fund_coordinates = []
    for row_number in range(2, fund_sheet.max_row + 1):
        field = str(fund_sheet.cell(row_number, fund_headers["fieldName"]).value or "")
        formula = _fund_bloomberg_formula(
            fund_sheet, row_number, field, fund_headers, fund,
        )
        if formula is None:
            continue
        cell = fund_sheet.cell(row_number, fund_headers["value"], formula)
        fund_coordinates.append(cell.coordinate)

    position_headers = {str(cell.value): cell.column for cell in positions_sheet[1]}
    position_coordinates = []
    for row_number in range(2, positions_sheet.max_row + 1):
        for field in sorted(POSITION_BLOOMBERG_FIELDS):
            formula = _position_bloomberg_formula(row_number, field, position_headers)
            if formula is None:
                continue
            cell = positions_sheet.cell(row_number, position_headers[field], formula)
            position_coordinates.append(cell.coordinate)
    return fund_coordinates, position_coordinates


def _expected_fund_formula(sheet, row_number: int, header: str, headers: dict[str, int]):
    if header != "value":
        return None
    field = str(sheet.cell(row_number, headers["fieldName"]).value or "")
    if not sheet.title.endswith(FUND_DATA_SHEET_SUFFIX):
        return None
    fund = sheet.title.removesuffix(FUND_DATA_SHEET_SUFFIX)
    return _fund_bloomberg_formula(sheet, row_number, field, headers, fund)


def _expected_position_formula(
    _sheet, row_number: int, header: str, headers: dict[str, int],
):
    return _position_bloomberg_formula(row_number, header, headers)


def _add_list_validation(sheet, header: str, list_name: str, max_row: int = 5000) -> None:
    headers = {str(cell.value): cell.column for cell in sheet[1]}
    column = headers[header]
    values = _LISTS[list_name]
    serialized = ",".join(values)
    if len(serialized) > 255:
        raise ValueError(f"{list_name} validation list exceeds Excel's inline limit")
    validation = DataValidation(
        type="list",
        formula1=f'"{serialized}"',
        allow_blank=True,
    )
    validation.error = f"Choose a value from the {list_name} list."
    validation.errorTitle = "Invalid value"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    target_letter = sheet.cell(1, column).column_letter
    validation.add(f"{target_letter}2:{target_letter}{max(max_row, sheet.max_row)}")


def _config_rows(config: object | None) -> list[dict[str, object]]:
    rows = []
    for external in CONFIG_INPUT_FIELDS:
        current = str(getattr(config, CONFIG_KEY_MAP[external], "")) if config else ""
        proposed = "" if external == "ccc" or current == "XXXXXXXX" else current
        rows.append({
            "fieldName": external,
            "value": proposed,
        })
    return rows


def _fund_data_rows(periods: tuple[str, ...]) -> list[dict[str, object]]:
    rows = []
    for period in periods:
        for field in FILING_INPUT_FIELDS:
            rows.append({
                "Period": period,
                "fieldName": field,
                "value": "",
            })
    return rows


def _position_starter_rows(
    periods: tuple[str, ...], headers: tuple[str, ...],
) -> list[dict[str, object]]:
    return [
        {**{header: "" for header in headers}, "Period": period}
        for period in periods
    ]


def _order_starter_rows(
    funds: list[str], periods: tuple[str, ...], headers: tuple[str, ...],
) -> list[dict[str, object]]:
    return [
        {**{header: "" for header in headers}, "Fund": fund, "Period": period}
        for fund in funds for period in periods
    ]


def create_input_workbook_template(
    funds_dir: str | Path, periods: list[str] | tuple[str, ...], output: str | Path,
) -> Path:
    """Create one fund-partitioned data-entry workbook for the full universe."""
    periods = _periods(periods)
    entries = _fund_entries(Path(funds_dir))
    funds = [fund for fund, _config, _category in entries]
    _LISTS["Funds"] = funds
    _LISTS["Periods"] = list(periods)
    output = Path(output)
    if output.suffix.lower() != ".xlsx":
        raise ValueError("input workbook output must use the .xlsx extension")
    if output.exists():
        raise FileExistsError(f"refusing to overwrite input workbook: {output}")

    workbook = Workbook()
    workbook.remove(workbook.active)

    for fund, config, _category in entries:
        config_sheet = workbook.create_sheet(_sheet_name(fund, CONFIG_SHEET_SUFFIX))
        config_sheet.sheet_properties.tabColor = "70AD47"
        _append_rows(config_sheet, CONFIG_HEADERS, _config_rows(config))
        _style_table(config_sheet, editable_headers={"value"})

        fund_sheet = workbook.create_sheet(_sheet_name(fund, FUND_DATA_SHEET_SUFFIX))
        fund_sheet.sheet_properties.tabColor = "5B9BD5"
        _append_rows(fund_sheet, FUND_DATA_HEADERS, _fund_data_rows(periods))
        _style_table(fund_sheet, editable_headers={"value"})

        positions_sheet = workbook.create_sheet(_sheet_name(fund, POSITIONS_SHEET_SUFFIX))
        positions_sheet.sheet_properties.tabColor = "ED7D31"
        _append_rows(
            positions_sheet, POSITION_HEADERS,
            _position_starter_rows(periods, POSITION_HEADERS),
        )
        fund_bloomberg_cells, position_bloomberg_cells = _seed_bloomberg_formulas(
            fund, fund_sheet, positions_sheet,
        )
        _style_bloomberg_cells(fund_sheet, fund_bloomberg_cells)
        _style_table(positions_sheet, freeze="C2", editable_headers=set(POSITION_HEADERS))
        _style_bloomberg_cells(positions_sheet, position_bloomberg_cells)
        for column in range(3, positions_sheet.max_column + 1):
            letter = positions_sheet.cell(1, column).column_letter
            positions_sheet.column_dimensions[letter].width = 16

        for sheet in (fund_sheet, positions_sheet):
            _add_list_validation(sheet, "Period", "Periods")
        for header, list_name in (
            ("units", "Units"), ("payoffProfile", "PayoffProfiles"),
            ("assetCat", "AssetCategories"), ("issuerCat", "IssuerCategories"),
            ("fairValLevel", "FairValueLevels"), ("derivCat", "DerivativeCategories"),
            ("couponKind", "CouponKinds"), ("putOrCall", "PutCall"),
            ("writtenOrPur", "WrittenPurchased"), ("recFixedOrFloating", "FixedFloating"),
            ("pmntFixedOrFloating", "FixedFloating"), ("refInstType", "ReferenceTypes"),
            ("recRateTenor", "TenorUnits"), ("recResetDt", "TenorUnits"),
            ("pmntRateTenor", "TenorUnits"), ("pmntResetDt", "TenorUnits"),
        ):
            _add_list_validation(positions_sheet, header, list_name)
        for header in (
            "isRestrictedSec", "isCashCollateral", "isNonCashCollateral", "isLoanByFund",
            "isDefault", "areIntrstPmntsInArrs", "isPaidKind", "swapFlag",
        ):
            _add_list_validation(positions_sheet, header, "YN")

    orders_sheet = workbook.create_sheet(ORDERS_SHEET)
    orders_sheet.sheet_properties.tabColor = "A5A5A5"
    order_rows = _order_starter_rows(funds, periods, ORDER_HEADERS)
    for row in order_rows:
        row["Ticker"] = row["Fund"]
    _append_rows(orders_sheet, ORDER_HEADERS, order_rows)
    _style_table(orders_sheet, editable_headers=set(ORDER_HEADERS))

    for name, value in (
        ("nportFormatVersion", INPUT_WORKBOOK_FORMAT_VERSION),
        ("nportSchemaVersion", NPORT_SCHEMA_VERSION),
        ("nportPeriods", ";".join(periods)),
        ("nportFunds", ";".join(funds)),
        (
            "nportFundCategoryCounts",
            "Thematic=30;2x Leveraged=125;Structured Buffer=36;Fixed Income=6",
        ),
    ):
        workbook.custom_doc_props.append(StringProperty(name=name, value=value))

    if len(",".join(funds)) <= 255:
        _add_list_validation(orders_sheet, "Fund", "Funds")
    _add_list_validation(orders_sheet, "Period", "Periods")
    _add_list_validation(orders_sheet, "Side", "OrderSides")
    _add_list_validation(orders_sheet, "Status", "OrderStatuses")

    workbook.calculation.calcMode = "auto"
    workbook.calculation.calcOnSave = True
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=output.parent, suffix=".xlsx", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, output)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    return output


def _headers(sheet, expected: tuple[str, ...]) -> dict[str, int]:
    values = [str(cell.value or "").strip() for cell in sheet[1]]
    duplicates = sorted({value for value in values if value and values.count(value) > 1})
    missing = sorted(set(expected) - set(values))
    unknown = sorted(set(values) - set(expected))
    if duplicates or missing or unknown:
        details = []
        if duplicates:
            details.append("duplicate=" + ",".join(duplicates))
        if missing:
            details.append("missing=" + ",".join(missing))
        if unknown:
            details.append("unknown=" + ",".join(unknown))
        raise ValueError(f"{sheet.title}: invalid header ({'; '.join(details)})")
    return {value: index + 1 for index, value in enumerate(values)}


def _cell_text(value: object, location: str) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        if value.startswith("="):
            raise ValueError(f"{location}: formulas are not accepted as input values")
        return value.strip()
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (int, float, Decimal)):
        return format(Decimal(str(value)), "f")
    return str(value).strip()


def _normalized_formula(value: str) -> str:
    return (
        value.replace("_xll.", "")
        .replace("_xludf.", "")
        .replace("@BDP(", "BDP(")
    )


def _unresolved_bloomberg_value(value: object) -> bool:
    if not isinstance(value, str):
        return False
    normalized = value.strip().upper()
    return normalized.startswith((
        "#N/A", "#VALUE!", "#NAME?", "#REF!", "#NUM!", "#DIV/0!",
        "#NULL!", "#GETTING_DATA", "REQUESTING DATA",
    ))


def _sheet_rows(
    sheet, expected: tuple[str, ...], *, cached_sheet=None,
    formula_factory=None, formula_cells: set[tuple[int, str]] | None = None,
) -> list[dict[str, str]]:
    headers = _headers(sheet, expected)
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        row = {}
        for header, column in headers.items():
            cell = sheet.cell(row_number, column)
            location = f"{sheet.title}!{cell.coordinate}"
            raw = cell.value
            if isinstance(raw, str) and raw.startswith("="):
                expected_formula = (
                    formula_factory(sheet, row_number, header, headers)
                    if formula_factory is not None else None
                )
                if expected_formula is None or _normalized_formula(raw) != expected_formula:
                    raise ValueError(f"{location}: formulas are not accepted as input values")
                if formula_cells is not None:
                    formula_cells.add((row_number, header))
                cached_cell = cached_sheet.cell(row_number, column)
                raw = cached_cell.value
                if cached_cell.data_type == "e" or _unresolved_bloomberg_value(raw):
                    raw = ""
            row[header] = _cell_text(raw, location)
        if any(row.values()):
            row["_rowNumber"] = str(row_number)
            rows.append(row)
    return rows


def _selected(rows: list[dict[str, str]], fund: str, period: str | None = None) -> list[dict[str, str]]:
    return [
        row for row in rows
        if row.get("Fund", "").upper() == fund
        and (period is None or row.get("Period", "") == period)
    ]


def _payload_rows(
    rows: list[dict[str, str]], payload_fields: tuple[str, ...], *, sheet_name: str,
    dimension_fields: tuple[str, ...] = ("Fund", "Period"),
) -> list[dict[str, str]]:
    out = []
    for row in rows:
        if any(row.get(field, "") for field in payload_fields):
            missing = [field for field in dimension_fields if not row.get(field)]
            if missing:
                raise ValueError(
                    f"{sheet_name}: populated row is missing " + " or ".join(missing)
                )
            out.append(row)
    return out


def _write_immutable_csv(path: Path, headers: tuple[str, ...], rows: list[dict[str, str]]) -> None:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    content = buffer.getvalue().encode("utf-8")
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        if path.read_bytes() != content:
            raise FileExistsError(
                f"normalized input artifact already exists with different content: {path}"
            )
        return
    with tempfile.NamedTemporaryFile(dir=path.parent, suffix=".tmp", delete=False) as handle:
        handle.write(content)
        temporary = Path(handle.name)
    try:
        os.replace(temporary, path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def _apply_review_rows(
    sheet, master_rows: list[dict[str, str]], *, key_fields: tuple[str, ...],
) -> None:
    headers = {str(cell.value): cell.column for cell in sheet[1]}
    target = {}
    for row_number in range(2, sheet.max_row + 1):
        key = tuple(str(sheet.cell(row_number, headers[field]).value or "") for field in key_fields)
        target[key] = row_number
    seen = set()
    for master in master_rows:
        key = tuple(master[field] for field in key_fields)
        if key in seen:
            raise ValueError(f"duplicate review row key: {key}")
        seen.add(key)
        row_number = target.get(key)
        if row_number is None:
            if master.get("status") not in {"", "MISSING"}:
                raise ValueError(f"review row does not exist in generated workbook: {key}")
            continue
        # MISSING is the untouched template state.  It must not erase values
        # deterministically populated by prepare_inputs (notably order flows).
        if master.get("status") in {"", "MISSING"}:
            continue
        for field in ("proposedValue", "sourceId", "status", "comment"):
            value = master.get(field, "")
            if field == "sourceId" and master.get("status") == "PROVIDED" and not value:
                value = INPUT_WORKBOOK_SOURCE_ID
            sheet.cell(row_number, headers[field], value)


def _to_review_rows(
    rows: list[dict[str, str]], *, bloomberg_cells: set[tuple[int, str]] | None = None,
) -> list[dict[str, str]]:
    """Translate input-only rows into the downstream review contract."""
    bloomberg_cells = bloomberg_cells or set()
    review_rows = []
    for row in rows:
        value = row.get("value", "")
        is_bloomberg = (int(row["_rowNumber"]), "value") in bloomberg_cells
        review_rows.append({
            **row,
            "proposedValue": value,
            "sourceId": "BLOOMBERG" if is_bloomberg and value else INPUT_WORKBOOK_SOURCE_ID,
            "status": "PROVIDED" if value else "MISSING",
            "comment": (
                "Calculated by the input-workbook Bloomberg formula."
                if is_bloomberg and value else "Entered in the input workbook."
                if value else ""
            ),
        })
    return review_rows


def _position_required_fields(row: dict[str, str], liquidity_required: bool) -> set[str]:
    required = set(get_required_fields(
        deriv_cat=row.get("derivCat", ""),
        has_debt=(
            bool(row.get("maturityDt")) or row.get("assetCat", "") in DEBT_ASSET_CATEGORIES
        ),
        ref_inst_type=row.get("refInstType", ""),
        rec_fixed_or_floating=row.get("recFixedOrFloating", ""),
        pmnt_fixed_or_floating=row.get("pmntFixedOrFloating", ""),
        currency=row.get("curCd", ""),
    ))
    if row.get("issuerCat") == "OTHER":
        required.add("issuer_conditional_desc")
    if row.get("assetCat") == "OTHER":
        required.add("asset_conditional_desc")
    if liquidity_required:
        required.add("liquidity_classification_json")
    return required


def _apply_positions(
    sheet, positions: list[dict[str, str]], liquidity_required: bool,
    bloomberg_fields: set[tuple[str, str]],
) -> None:
    headers = {str(cell.value): cell.column for cell in sheet[1]}
    by_id = {row["holdingId"]: row for row in positions}
    if len(by_id) != len(positions):
        raise ValueError("Positions: holdingId values must be unique within a fund-period")
    field_to_attr = HOLDINGS_KEY_MAP
    for row_number in range(2, sheet.max_row + 1):
        holding_id = str(sheet.cell(row_number, headers["recordKey"]).value or "")
        field = str(sheet.cell(row_number, headers["fieldName"]).value or "")
        source = by_id.get(holding_id)
        if source is None:
            continue
        value = source.get(field, "")
        required = _position_required_fields(source, liquidity_required)
        source_id = (
            "BLOOMBERG" if (holding_id, field) in bloomberg_fields
            else INPUT_WORKBOOK_SOURCE_ID
        )
        if value:
            sheet.cell(row_number, headers["proposedValue"], value)
            sheet.cell(row_number, headers["sourceId"], source_id)
            sheet.cell(row_number, headers["status"], "PROVIDED")
            sheet.cell(
                row_number, headers["comment"],
                "Calculated by the input-workbook Bloomberg formula."
                if source_id == "BLOOMBERG" else "Entered in the input workbook.",
            )
        elif (holding_id, field) in bloomberg_fields:
            # Retain the standard filing_inputs Bloomberg formula when the
            # input workbook has not cached a resolved value yet.
            continue
        elif field not in {"isin", "ticker"} and field_to_attr[field] not in required:
            sheet.cell(row_number, headers["status"], "NOT_APPLICABLE")
            sheet.cell(row_number, headers["comment"], "Not applicable from the selected holding classification.")


def _metadata(workbook) -> dict[str, str]:
    properties = {item.name: str(item.value) for item in workbook.custom_doc_props}
    values = {
        "formatVersion": properties.get("nportFormatVersion", ""),
        "nportSchemaVersion": properties.get("nportSchemaVersion", ""),
        "periods": properties.get("nportPeriods", ""),
        "funds": properties.get("nportFunds", ""),
    }
    if values.get("formatVersion") != INPUT_WORKBOOK_FORMAT_VERSION:
        raise ValueError("unsupported input workbook format version")
    if values.get("nportSchemaVersion") != NPORT_SCHEMA_VERSION:
        raise ValueError(
            f"workbook schema v{values.get('nportSchemaVersion')} does not match v{NPORT_SCHEMA_VERSION}"
        )
    funds = values.get("funds", "").split(";")
    if (
        not all(funds)
        or len(funds) != len(set(funds))
        or any(not fund.isascii() or not fund.isalnum() for fund in funds)
    ):
        raise ValueError("input workbook has invalid fund metadata")
    expected_sheets = _expected_sheet_names(funds)
    if workbook.sheetnames != expected_sheets:
        missing = [name for name in expected_sheets if name not in workbook.sheetnames]
        extra = [name for name in workbook.sheetnames if name not in expected_sheets]
        details = []
        if missing:
            details.append("missing=" + ",".join(missing[:10]))
        if extra:
            details.append("extra=" + ",".join(extra[:10]))
        if not missing and not extra:
            details.append("worksheet order changed")
        raise ValueError(
            "input workbook violates the fund-partitioned sheet contract ("
            + "; ".join(details) + ")"
        )
    return values


def _require_field_contract(
    rows: list[dict[str, str]], expected: tuple[str, ...], *, sheet_name: str,
) -> None:
    fields = [row.get("fieldName", "") for row in rows]
    duplicates = sorted({field for field in fields if fields.count(field) > 1})
    missing = sorted(set(expected) - set(fields))
    unknown = sorted(set(fields) - set(expected))
    if duplicates or missing or unknown:
        details = []
        if duplicates:
            details.append("duplicate=" + ",".join(duplicates))
        if missing:
            details.append("missing=" + ",".join(missing))
        if unknown:
            details.append("unknown=" + ",".join(unknown))
        raise ValueError(f"{sheet_name}: invalid field contract ({'; '.join(details)})")


def _require_registered_dimensions(
    rows: list[dict[str, str]], *, sheet_name: str,
    funds: set[str], periods: set[str] | None = None,
) -> None:
    for row_number, row in enumerate(rows, 2):
        fund = row.get("Fund", "").upper()
        if fund not in funds:
            raise ValueError(
                f"{sheet_name}!row {row_number}: Fund {row.get('Fund')!r} is not registered"
            )
        if periods is not None and row.get("Period", "") not in periods:
            raise ValueError(
                f"{sheet_name}!row {row_number}: Period {row.get('Period')!r} is not registered"
            )


def _require_registered_periods(
    rows: list[dict[str, str]], *, sheet_name: str, periods: set[str],
) -> None:
    for row_number, row in enumerate(rows, 2):
        if row.get("Period", "") not in periods:
            raise ValueError(
                f"{sheet_name}!row {row_number}: "
                f"Period {row.get('Period')!r} is not registered"
            )


def prepare_from_input_workbook(
    fund_dir: str | Path, fund: str, period: str, workbook_path: str | Path,
) -> Path:
    """Import one fund-period from a multi-fund workbook into the normal review path."""
    fund = fund.strip().upper()
    period = period.strip()
    report_date = report_date_for_period(period)
    workbook_path = Path(workbook_path).resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"input workbook not found: {workbook_path}")
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()

    source = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        cached = load_workbook(workbook_path, data_only=True, read_only=False)
    except Exception:
        source.close()
        raise
    try:
        metadata = _metadata(source)
        registered_funds = metadata.get("funds", "").split(";")
        registered_periods = metadata.get("periods", "").split(";")
        if not all(registered_periods) or len(registered_periods) != len(set(registered_periods)):
            raise ValueError("input workbook has invalid period metadata")
        if fund not in registered_funds:
            raise ValueError(f"fund {fund} is not registered in the input workbook")
        if period not in registered_periods:
            raise ValueError(f"period {period} is not registered in the input workbook")

        valid_funds = set(registered_funds)
        valid_periods = set(registered_periods)
        config_sheet_name = _sheet_name(fund, CONFIG_SHEET_SUFFIX)
        fund_data_sheet_name = _sheet_name(fund, FUND_DATA_SHEET_SUFFIX)
        positions_sheet_name = _sheet_name(fund, POSITIONS_SHEET_SUFFIX)
        fund_formula_cells: set[tuple[int, str]] = set()
        position_formula_cells: set[tuple[int, str]] = set()
        all_config_rows = _sheet_rows(source[config_sheet_name], CONFIG_HEADERS)
        all_fund_rows = _sheet_rows(
            source[fund_data_sheet_name], FUND_DATA_HEADERS,
            cached_sheet=cached[fund_data_sheet_name],
            formula_factory=_expected_fund_formula,
            formula_cells=fund_formula_cells,
        )
        _require_registered_periods(
            all_fund_rows, sheet_name=fund_data_sheet_name, periods=valid_periods,
        )
        config_rows = _to_review_rows(all_config_rows)
        fund_rows = _to_review_rows(
            [row for row in all_fund_rows if row.get("Period") == period],
            bloomberg_cells=fund_formula_cells,
        )
        _require_field_contract(
            config_rows, CONFIG_INPUT_FIELDS, sheet_name=config_sheet_name,
        )
        _require_field_contract(
            fund_rows, FILING_INPUT_FIELDS, sheet_name=fund_data_sheet_name,
        )
        raw_positions = _payload_rows(
            _sheet_rows(
                source[positions_sheet_name], POSITION_HEADERS,
                cached_sheet=cached[positions_sheet_name],
                formula_factory=_expected_position_formula,
                formula_cells=position_formula_cells,
            ),
            ("holdingId", *POSITION_FIELDS), sheet_name=positions_sheet_name,
            dimension_fields=("Period",),
        )
        _require_registered_periods(
            raw_positions, sheet_name=positions_sheet_name, periods=valid_periods,
        )
        positions = [row for row in raw_positions if row.get("Period") == period]
        if not positions:
            raise ValueError(
                f"{positions_sheet_name}: no populated rows for {fund} {period}"
            )
        if any(not row.get("holdingId") for row in positions):
            raise ValueError(
                f"{positions_sheet_name}: every {fund} {period} row requires holdingId"
            )
        allowed_bloomberg_by_row: dict[int, set[str]] = {}
        for row in positions:
            row_number = int(row["_rowNumber"])
            spec_key = bloomberg_spec_key(row)
            allowed = set(BLOOMBERG_SPECS[spec_key][1]) if spec_key else set()
            allowed_bloomberg_by_row[row_number] = allowed
            for field in POSITION_BLOOMBERG_FIELDS - allowed:
                if (row_number, field) in position_formula_cells:
                    row[field] = ""
        position_bloomberg_fields = {
            (row["holdingId"], field)
            for row in positions
            for field in allowed_bloomberg_by_row[int(row["_rowNumber"])]
            if (int(row["_rowNumber"]), field) in position_formula_cells
        }

        raw_orders = _payload_rows(
            _sheet_rows(source[ORDERS_SHEET], ORDER_HEADERS),
            ("Side", "Trade Date", "Notional", "Status"), sheet_name=ORDERS_SHEET,
        )
        _require_registered_dimensions(
            raw_orders, sheet_name=ORDERS_SHEET,
            funds=valid_funds, periods=valid_periods,
        )
        orders = _selected(raw_orders, fund, period)
        for row in orders:
            row["Ticker"] = row.get("Ticker") or fund
            if row["Ticker"].upper() != fund:
                raise ValueError(f"Orders: Ticker {row['Ticker']!r} does not match Fund {fund}")
            raw_date = row["Trade Date"]
            try:
                parsed = date.fromisoformat(raw_date)
            except ValueError:
                parsed = None
            if parsed is not None:
                row["Trade Date"] = f"{parsed.month}/{parsed.day}/{parsed.year}"

    finally:
        source.close()
        cached.close()

    normalized = (
        workbook_path.parent / "_normalized" / f"nport-v{NPORT_SCHEMA_VERSION}"
        / digest / fund.lower() / period
    )
    position_path = normalized / "positions.csv"
    position_rows = [
        {
            header: (
                "" if (int(row["_rowNumber"]), header) in position_formula_cells
                else row.get(header, "")
            )
            for header in ("holdingId", *POSITION_FIELDS)
        }
        for row in positions
    ]
    _write_immutable_csv(position_path, ("holdingId", *POSITION_FIELDS), position_rows)

    order_path = None
    if orders:
        order_path = normalized / "orders.csv"
        _write_immutable_csv(
            order_path, ("Ticker", "Side", "Trade Date", "Notional", "Status"), orders,
        )

    output = prepare_inputs(
        fund_dir, period, positions=position_path, orders=order_path,
    )
    reviewed = load_workbook(output)
    try:
        _apply_review_rows(
            reviewed["FundFields"], config_rows,
            key_fields=("fieldName",),
        )
        _apply_review_rows(
            reviewed["FundFields"], fund_rows,
            key_fields=("fieldName",),
        )
        config_values = {
            row["fieldName"]: row.get("proposedValue", "")
            for row in config_rows if row.get("status") == "PROVIDED"
        }
        liquidity_required = config_values.get("liquidityRequired", "").upper() in {"Y", "TRUE"}
        _apply_positions(
            reviewed["HoldingFields"], positions, liquidity_required,
            position_bloomberg_fields,
        )

        sources_sheet = reviewed["Sources"]
        existing_source_ids = {
            str(sources_sheet.cell(row, 1).value or "")
            for row in range(2, sources_sheet.max_row + 1)
        }
        if INPUT_WORKBOOK_SOURCE_ID in existing_source_ids:
            raise ValueError(f"reserved source ID already exists: {INPUT_WORKBOOK_SOURCE_ID}")
        sources_sheet.append([
            INPUT_WORKBOOK_SOURCE_ID, "input_workbook", "INTERNAL_WORKBOOK",
            "N-PORT Excel input workbook", str(workbook_path), report_date.isoformat(),
            digest, str(len(config_rows) + len(fund_rows) + len(positions) + len(orders)),
            f"Workbook format {INPUT_WORKBOOK_FORMAT_VERSION}; schema v{NPORT_SCHEMA_VERSION}",
        ])

        with tempfile.NamedTemporaryFile(dir=output.parent, suffix=".xlsx", delete=False) as handle:
            temporary = Path(handle.name)
        try:
            reviewed.save(temporary)
            os.replace(temporary, output)
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
    finally:
        reviewed.close()
    return output
