"""Centralized human-review workbook schema.

The workbook is generated after custodian, EagleSTAR, create/redeem, and Bloomberg
automation. It contains the residual values a person must complete before splitting
the data to each fund.

``data/humanreview/<period>_review.xlsx``, one sheet per category. The pipeline surfaces
two things here for a person to handle:

* **gaps** — a required field with no reliable feed (a security Bloomberg has no country
  for, a swap with no leg economics, a fund with no designated index), and
* **sus values** — low-confidence values that are ideal for a human to confirm.

The operator fills/edits the value cells; ``split``/``build`` read them back as a traceable
``REVIEWED`` source (recorded in the provenance manifest). An unfilled REQUIRED cell stays
blank → tagged ``GAP`` → blocks LIVE (see the reconciliation/LIVE gate).

The reference categories (designated index, swap counterparties, option underlying index)
are the hardcoded maps that USED to live in ``custodian.py`` / ``filing_master.py``; they are
seeded ONCE from the legacy literals below and thereafter owned by the workbook — no
code-embedded guessing. Per-item categories (swap legs, invCountry, isin, sus) are generated
from the current holdings each run, pre-filled with what's known, blank where a human must act.

Rebuild preserves operator edits: an existing workbook's filled cells survive; only new
rows/categories are added (same keep-manual contract as the master workbooks). Read at
split/build needs no Bloomberg terminal — the workbook is plain xlsx.
"""
import os
import tempfile
import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Legacy literals, relocated here as ONE-TIME migration seeds ────────────────
# These were hardcoded maps in custodian.py / filing_master.py. They seed a fresh
# workbook; once the operator has reviewed/saved, the workbook is the source of truth.

SEED_SWAP_COUNTERPARTIES = {
    # code -> (legalName, lei)  — GLEIF-verified
    "CANT": ("Cantor Fitzgerald & Co.", "5493004J7H4GCPG6OB62"),
    "CLST": ("Clear Street LLC", "549300KNQS43Y7TO3X67"),
    "CS": ("Clear Street LLC", "549300KNQS43Y7TO3X67"),
    "MREX": ("Marex Capital Markets Inc.", "5493006BWPDUCYG6EQ34"),
}

SEED_OPTION_INDEX = {
    # option underlying -> (indexName, indexIdentifier)
    "SPY": ("S&P 500 Index", "SPX"),
    "QQQ": ("NASDAQ 100 Index", "NDX"),
    "IWM": ("Russell 2000 Index", "RTY"),
    "EEM": ("MSCI Emerging Markets Index", "MXEF"),
    "EFA": ("MSCI EAFE Index", "MXEA"),
}

# The default total-return-swap leg economics that custodian.build_swap_entry used to
# hardcode for EVERY swap. Used to pre-fill the swap_legs sheet (recDesc is per-swap, so
# it's filled from the reference issuer at generation time, not seeded here).
SEED_SWAP_LEG_TEMPLATE = {
    "recFixedOrFloating": "Other",
    "pmntFloatingRtIndex": "USD-SOFR",
    "pmntFloatingRtSpread": "",          # from the swap confirm — operator fills
    "pmntRateTenor": "Month",
    "pmntRateUnit": "3",
}

# Designated broad-based index per fund (ticker -> (indexName, indexIdentifier)).
# Relocated from filing_master._DESIGNATED_INDEX (sourced from each fund's 497K).
_SEED_INDEX_NAME = {
    "SPX": "S&P 500 Index", "SPXT": "S&P 500 Total Return Index",
    "NDX": "Nasdaq-100 Index", "RTY": "Russell 2000 Index",
    "MXWD": "MSCI ACWI Index", "MXEF": "MSCI Emerging Markets Index", "MXEA": "MSCI EAFE Index",
    "CFIIBL3P": "FTSE US Treasury Bill 3-12 Months Index",
    "SBMMTB3": "FTSE 3-Month US Treasury Bill Index",
    "CFIIH52C": "FTSE US High-Yield Market 0-5 Years 2% Capped Index",
    "CFIIBD37": "FTSE US Treasury 3-7 Years Index",
    "SBUSC15U": "FTSE US Broad Investment-Grade Corporate Bond 1-5 Years Index",
    "SBUST13U": "FTSE US Treasury 1-3 Years Index",
}
_SEED_BENCHMARK_FUNDS = {
    "SPX": ("AV BAY BLCK BREW BZZ CBOT CJUN CMAG CQTM CTJN DIPR EUV EUVX EYES GASZ GLAM GNMX "
            "HJUN HMAY HULL JOUL JUNC KYC LATR NYNY ODDZ OWN PTNT STYL USX VBX VOOX WATS WNDR "
            "WR XA XAGI XBIX XCOM XHOA XIWC XKRE XLBX XLEX XLFX XLKX XLPX XLUX XLVX XLYX XPAV "
            "XSEM XVO XVUG YUNG "
            "ACLZ ACMM CAMC CARX CRUC KEYX LASC LRNX MNSX MSIX ONTX SIMX TPLX UMCX"),
    "SPXT": "FDRS GPTZ CMAY MAYC",
    "NDX": "QJN QMY QQJN QQMY",
    "RTY": "SCJN SCMY",
    "MXWD": "BRZX CCPX EMXX KRWX TAJX WEBX WX XW XTAI",
    "MXEF": "EMJN EMMY",
    "MXEA": "IDJN IDMY",
    "CFIIBL3P": "CBIL", "SBMMTB3": "CGOV", "CFIIH52C": "CHYG",
    "CFIIBD37": "CIEI", "SBUSC15U": "CIVG", "SBUST13U": "CUST",
}
SEED_DESIGNATED_INDEX = {
    t: (_SEED_INDEX_NAME[bench], bench)
    for bench, funds in _SEED_BENCHMARK_FUNDS.items()
    for t in funds.split()
}


# ── Sheet schema ───────────────────────────────────────────────────────────────
# Each sheet: key columns (identify the row, never edited) + value columns (operator
# fills/edits) + a trailing ``sourceNote``. ``required`` value columns that stay blank are
# GAPs that block LIVE.

@dataclass
class SheetSpec:
    name: str
    keys: list[str]
    values: list[str]
    required: list[str]          # subset of values whose blank = GAP
    editable: bool = True


_SHEETS = [
    # Fund-level policy and conditional XML values are deliberately in this workbook.
    # The reviewer does not edit derived dates/submission type; build derives those from
    # the policy fields below.
    SheetSpec("fund_policy", ["fund"],
              ["fiscalYearEndMMDD", "derivativesRegimePolicy", "liquidityRequired",
               "cashB2fRequired", "policyEffectiveFrom", "policyEffectiveTo"],
              ["fiscalYearEndMMDD", "derivativesRegimePolicy", "liquidityRequired",
               "cashB2fRequired", "policyEffectiveFrom"]),
    SheetSpec("fund_risk", ["fund"],
              ["cashNotReportedInCOrD", "monthlyReturnCategoriesJson",
               "derivExposurePct", "derivCurrencyExposurePct",
               "derivInterestRateExposurePct", "derivDaysInExcess",
               "medianDailyVarPct", "medianVarRatioPct", "backtestingExceptions"], []),
    SheetSpec("designated_index", ["ticker"], ["indexName", "indexIdentifier"], []),
    SheetSpec("swap_counterparties", ["code"], ["legalName", "lei"], ["legalName", "lei"]),
    SheetSpec("option_index", ["underlying"], ["indexName", "indexIdentifier"],
              ["indexName", "indexIdentifier"]),
    # pmntFloatingRtSpread IS required: input_validation demands a number for it whenever
    # pmntFixedOrFloating is Floating (which every TRS is), so leaving it optional here let
    # a blank reach the build and fail it. A no-spread swap is an explicit "0", not a blank.
    SheetSpec("swap_legs", ["fund", "swapTicker"],
              ["recFixedOrFloating", "recDesc", "pmntFloatingRtIndex",
               "pmntFloatingRtSpread", "pmntRateTenor", "pmntRateUnit"],
              ["recFixedOrFloating", "recDesc", "pmntFloatingRtIndex",
               "pmntFloatingRtSpread", "pmntRateTenor", "pmntRateUnit"]),
    SheetSpec("option_delta", ["fund", "ticker"], ["delta"], ["delta"]),
    SheetSpec("holding_liquidity", ["fund", "ticker", "cusip", "name"],
              ["liquidityClassificationJson", "liquidityCircumstancesJson"], []),
    SheetSpec("invCountry", ["fund", "ticker", "cusip", "name"], ["invCountry"], ["invCountry"]),
    # isin is optional enrichment — a missing ISIN is filed as other=CUSIP, so it is NOT a
    # blocking gap (required=[]).
    SheetSpec("isin", ["fund", "ticker", "cusip", "name"], ["isin"], []),
    SheetSpec("sus_review", ["fund", "field"], ["value", "reason", "confirmed"], []),
    # These sheets expose operational failures in the same review stage. They are
    # intentionally read-only: a source disagreement is resolved in the source system,
    # not by overwriting the filed value in Excel.
    SheetSpec("reconciliation",
              ["check", "fund", "source_a", "value_a", "source_b", "value_b",
               "diff", "flag"], [], [], editable=False),
    SheetSpec("pipeline_status", ["gate", "status", "detail", "resolution"], [], [],
              editable=False),
]
_SHEET_BY_NAME = {s.name: s for s in _SHEETS}


@dataclass
class Review:
    """Reviewed values read back from the workbook, keyed for direct lookup."""
    sheets: dict[str, list[dict[str, str]]] = field(default_factory=dict)

    def _rows(self, sheet: str) -> list[dict[str, str]]:
        return self.sheets.get(sheet, [])

    def designated_index(self) -> dict[str, tuple[str, str]]:
        return {r["ticker"]: (r["indexName"], r["indexIdentifier"])
                for r in self._rows("designated_index")
                if r.get("ticker") and r.get("indexName") and r.get("indexIdentifier")}

    def swap_counterparties(self) -> dict[str, tuple[str, str]]:
        return {r["code"].upper(): (r["legalName"], r["lei"])
                for r in self._rows("swap_counterparties")
                if r.get("code") and r.get("legalName")}

    def option_index(self) -> dict[str, tuple[str, str]]:
        return {r["underlying"]: (r["indexName"], r["indexIdentifier"])
                for r in self._rows("option_index")
                if r.get("underlying") and r.get("indexName")}

    def swap_legs(self) -> dict[tuple[str, str], dict[str, str]]:
        return {(r["fund"], r["swapTicker"]): r
                for r in self._rows("swap_legs") if r.get("swapTicker")}

    def inv_country(self) -> dict[tuple[str, str], str]:
        return {(r["fund"], r["cusip"] or r["ticker"]): r["invCountry"]
                for r in self._rows("invCountry") if r.get("invCountry")}

    def isin(self) -> dict[tuple[str, str], str]:
        return {(r["fund"], r["cusip"] or r["ticker"]): r["isin"]
                for r in self._rows("isin") if r.get("isin")}

    def fund_policy(self) -> dict[str, dict[str, str]]:
        return {r["fund"].upper(): r for r in self._rows("fund_policy") if r.get("fund")}

    def fund_risk(self) -> dict[str, dict[str, str]]:
        return {r["fund"].upper(): r for r in self._rows("fund_risk") if r.get("fund")}

    def option_delta(self) -> dict[tuple[str, str], str]:
        return {(r["fund"], r["ticker"]): r["delta"]
                for r in self._rows("option_delta") if r.get("delta")}

    def holding_liquidity(self) -> dict[tuple[str, str], dict[str, str]]:
        return {(r["fund"], r.get("cusip") or r.get("ticker")): r
                for r in self._rows("holding_liquidity")
                if r.get("fund") and (r.get("cusip") or r.get("ticker"))}


# ── I/O ────────────────────────────────────────────────────────────────────────


def _row_key(spec: SheetSpec, row: dict[str, str]) -> tuple:
    return tuple((row.get(k) or "").strip() for k in spec.keys)


def read_review(path: Path) -> Review:
    """Read the workbook into a ``Review``; an absent file yields an empty review."""
    path = Path(path)
    if not path.is_file():
        return Review()
    wb = load_workbook(path, data_only=True)
    sheets: dict[str, list[dict[str, str]]] = {}
    for spec in _SHEETS:
        if spec.name not in wb.sheetnames:
            continue
        ws = wb[spec.name]
        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            continue
        header = [str(h).strip() if h is not None else "" for h in raw[0]]
        rows = []
        for r in raw[1:]:
            if r is None or all(c is None for c in r):
                continue
            rec = {header[i]: ("" if r[i] is None else str(r[i]))
                   for i in range(min(len(header), len(r)))}
            rows.append(rec)
        sheets[spec.name] = rows
    return Review(sheets=sheets)


def _required_for_row(spec: SheetSpec, row: dict[str, str],
                      policies: dict[str, dict[str, str]]) -> list[str]:
    """Return required cells after applying explicit fund policy; infer nothing."""
    required = list(spec.required)
    fund = (row.get("fund") or "").upper()
    policy = policies.get(fund, {})
    regime = (policy.get("derivativesRegimePolicy") or "").strip().upper()
    if spec.name == "fund_risk":
        if (policy.get("cashB2fRequired") or "").strip().upper() == "Y":
            required.append("cashNotReportedInCOrD")
        if regime == "LIMITED":
            required += ["derivExposurePct", "derivCurrencyExposurePct",
                         "derivInterestRateExposurePct", "derivDaysInExcess"]
        if regime in {"VAR_ABSOLUTE", "VAR_RELATIVE"}:
            required += ["medianDailyVarPct", "backtestingExceptions"]
        if regime == "VAR_RELATIVE":
            required.append("medianVarRatioPct")
    elif spec.name == "holding_liquidity":
        if (policy.get("liquidityRequired") or "").strip().upper() == "Y":
            required.append("liquidityClassificationJson")
    return required


def _style_sheet(ws, spec: SheetSpec, cols: list[str], required_by_row: list[list[str]]) -> None:
    """Make input columns unmistakable and keep large review sheets usable."""
    navy, teal, pale_blue = "243B53", "287D8E", "E8F1F5"
    grey, red, green = "E7ECF1", "FDE7E4", "E7F4EC"
    thin = Side(style="thin", color="CDD6DF")
    editable_cols = set(spec.values) if spec.editable else set()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.row_dimensions[1].height = 32
    ws.sheet_properties.tabColor = teal if spec.editable else "7B8794"
    for col_index, header in enumerate(cols, 1):
        cell = ws.cell(1, col_index)
        cell.fill = PatternFill("solid", fgColor=teal if header in editable_cols else navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = max(12, min(42, max(len(str(ws.cell(r, col_index).value or ""))
                                    for r in range(1, ws.max_row + 1)) + 2))
        ws.column_dimensions[cell.column_letter].width = width
    for row_index in range(2, ws.max_row + 1):
        required = set(required_by_row[row_index - 2]) if row_index - 2 < len(required_by_row) else set()
        for col_index, header in enumerate(cols, 1):
            cell = ws.cell(row_index, col_index)
            cell.number_format = "@"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
            if header in editable_cols:
                if header in required:
                    cell.fill = PatternFill(
                        "solid", fgColor=red if not str(cell.value or "").strip() else green
                    )
                else:
                    cell.fill = PatternFill("solid", fgColor=pale_blue)
            elif not spec.editable:
                cell.fill = PatternFill("solid", fgColor=grey)


def _add_controlled_values(wb) -> None:
    """Add dropdowns for the small policy vocabularies; no business value is supplied."""
    if "fund_policy" not in wb.sheetnames:
        return
    ws = wb["fund_policy"]
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    controls = {
        "derivativesRegimePolicy": ["NONE", "LIMITED", "VAR_ABSOLUTE", "VAR_RELATIVE"],
        "liquidityRequired": ["Y", "N"],
        "cashB2fRequired": ["Y", "N"],
    }
    for header, values in controls.items():
        col = headers[header]
        validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"')
        validation.error = "Choose a listed value; the pipeline does not infer policy."
        validation.showErrorMessage = True
        ws.add_data_validation(validation)
        validation.add(f"{col}2:{col}{max(2, ws.max_row)}")


def build_review_workbook(path: Path, generated: dict[str, list[dict[str, str]]]) -> dict[str, int]:
    """Write/refresh the review workbook, preserving operator edits.

    ``generated`` maps a sheet name → the rows the current run wants present (keys + any
    known pre-fill values + sourceNote). Reference sheets not in ``generated`` are seeded
    from the legacy literals. For every sheet, an existing row (matched by its key columns)
    keeps the operator's filled value cells; new rows are added; rows no longer generated
    are dropped (except reference sheets, which persist). Returns ``{sheet: gap_count}``.
    """
    path = Path(path)
    existing = read_review(path).sheets if path.is_file() else {}
    generated = generated or {}

    out_by_sheet: dict[str, list[dict[str, str]]] = {}
    for spec in _SHEETS:
        prior = {_row_key(spec, r): r for r in existing.get(spec.name, [])}
        wanted = generated.get(spec.name)
        if wanted is None:
            wanted = _seed_rows(spec.name)            # reference sheet → seed/persist
            if not wanted and prior:                  # nothing seeded but operator has rows
                wanted = list(prior.values())
        out_rows: list[dict[str, str]] = []
        for row in wanted:
            key = _row_key(spec, row)
            merged = dict(row)
            if key in prior:                          # keep operator's filled values
                for v in spec.values:
                    if (prior[key].get(v) or "").strip():
                        merged[v] = prior[key][v]
            out_rows.append(merged)
        out_by_sheet[spec.name] = out_rows

    policies = {
        (row.get("fund") or "").upper(): row
        for row in out_by_sheet.get("fund_policy", []) if row.get("fund")
    }
    wb = Workbook()
    wb.remove(wb.active)
    gaps: dict[str, int] = {}
    for spec in _SHEETS:
        cols = spec.keys + spec.values + ["instruction"]
        out_rows = out_by_sheet.get(spec.name, [])
        required_by_row = [_required_for_row(spec, row, policies) for row in out_rows]
        gaps[spec.name] = sum(
            1 for row, required in zip(out_rows, required_by_row)
            if any(not (row.get(value) or "").strip() for value in required)
        )
        ws = wb.create_sheet(spec.name)
        ws.append(cols)
        for row in out_rows:
            ws.append([
                row.get("instruction", row.get("sourceNote", "")) if c == "instruction"
                else row.get(c, "") for c in cols
            ])
        _style_sheet(ws, spec, cols, required_by_row)
    _add_controlled_values(wb)

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        Path(tmp).replace(path)
    except Exception:
        Path(tmp).unlink(missing_ok=True)
        raise
    return gaps


def generate_from_master_rows(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    """Build the per-item review rows from the Bloomberg-populated master rows.

    Called at split (AFTER the operator has populated the workbook on the Bloomberg machine),
    so it surfaces only the residual gaps that custodian + Bloomberg + EagleSTAR could not
    fill. Reference sheets (designated_index, swap_counterparties, option_index) are seeded
    separately by ``build_review_workbook``.

    * ``swap_legs`` — one row per swap, pre-filled with the TRS template + a recDesc derived
      from the reference issuer (so it's reviewable, not a hidden hardcode).
    * ``invCountry`` — every holding still missing a country (a real gap that blocks LIVE).
    * ``isin``       — securities still missing an ISIN (optional enrichment; not a hard gap).
    """
    swap_legs, option_delta, liquidity, inv_country, isin_rows = [], [], [], [], []
    for r in rows:
        acct = (r.get("Account") or "").strip()
        dc = (r.get("derivCat") or "").strip()
        asset = (r.get("assetCat") or "").strip()
        if dc == "SWP":
            issuer = (r.get("refIssuerName") or r.get("title") or "").strip()
            leg = {"fund": acct, "swapTicker": (r.get("ticker") or "").strip(),
                   "recDesc": f"Total return of {issuer}" if issuer else "",
                   "sourceNote": "seed: TRS template — confirm against swap confirm"}
            leg.update(SEED_SWAP_LEG_TEMPLATE)
            swap_legs.append(leg)
        if dc == "OPT":
            option_delta.append({
                "fund": acct, "ticker": (r.get("ticker") or "").strip(),
                "delta": (r.get("delta") or "").strip(),
                "sourceNote": "report-date option delta; enter only when automation is blank",
            })
        if acct and (asset or dc):
            liquidity.append({
                "fund": acct, "ticker": (r.get("ticker") or "").strip(),
                "cusip": (r.get("cusip") or "").strip(), "name": (r.get("name") or "").strip(),
                "liquidityClassificationJson": (r.get("liquidityClassificationJson") or "").strip(),
                "liquidityCircumstancesJson": (r.get("liquidityCircumstancesJson") or "").strip(),
                "sourceNote": "required only when fund_policy.liquidityRequired=Y",
            })
        if asset and dc not in ("SWP", "OPT") and not (r.get("invCountry") or "").strip():
            inv_country.append({
                "fund": acct, "ticker": (r.get("ticker") or "").strip(),
                "cusip": (r.get("cusip") or "").strip(), "name": (r.get("name") or "").strip(),
                "invCountry": "", "sourceNote": "no country from Bloomberg — enter ISO-2"})
        if asset in ("EC", "DBT") and not (r.get("isin") or "").strip():
            isin_rows.append({
                "fund": acct, "ticker": (r.get("ticker") or "").strip(),
                "cusip": (r.get("cusip") or "").strip(), "name": (r.get("name") or "").strip(),
                "isin": "", "sourceNote": "no ISIN from Bloomberg (optional; CUSIP is filed)"})
    return {"swap_legs": swap_legs, "option_delta": option_delta,
            "holding_liquidity": liquidity, "invCountry": inv_country, "isin": isin_rows}


def _read_kv(path: Path) -> dict[str, str]:
    if not path.is_file():
        return {}
    values: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        values[key.strip()] = value.strip()
    return values


_POLICY_COLUMNS = [
    "fiscalYearEndMMDD", "derivativesRegimePolicy", "liquidityRequired",
    "cashB2fRequired", "policyEffectiveFrom", "policyEffectiveTo",
]
_RISK_COLUMNS = [
    "cashNotReportedInCOrD", "monthlyReturnCategoriesJson",
    "derivExposurePct", "derivCurrencyExposurePct", "derivInterestRateExposurePct",
    "derivDaysInExcess", "medianDailyVarPct", "medianVarRatioPct", "backtestingExceptions",
]


def generate_fund_rows(accounts: list[str], filing_rows: list[dict[str, str]],
                       funds_dir: Path) -> dict[str, list[dict[str, str]]]:
    """Generate the one-row-per-fund policy and conditional XML input tables."""
    filing_by_fund = {
        (row.get("Account") or "").strip().upper(): row for row in filing_rows
        if (row.get("Account") or "").strip()
    }
    policy_rows, risk_rows = [], []
    for fund in sorted({account.strip().upper() for account in accounts if account.strip()}):
        config = _read_kv(Path(funds_dir) / fund.lower() / "fund_config.txt")
        filing = filing_by_fund.get(fund, {})
        policy_rows.append({
            "fund": fund, **{column: config.get(column, "") for column in _POLICY_COLUMNS},
            "sourceNote": "enter approved internal fund policy; filing dates are derived",
        })
        risk_rows.append({
            "fund": fund, **{column: filing.get(column, "") for column in _RISK_COLUMNS},
            "sourceNote": "conditional XML values; required cells depend on fund_policy",
        })
    return {"fund_policy": policy_rows, "fund_risk": risk_rows}


def generate_reconciliation_rows(path: Path) -> list[dict[str, str]]:
    """Expose only REVIEW exceptions; values remain read-only in the workbook."""
    path = Path(path)
    if not path.is_file():
        return []
    with open(path, newline="", encoding="utf-8-sig") as handle:
        return [
            {**row, "sourceNote": "resolve in the named source input, then rerun masters"}
            for row in csv.DictReader(handle)
            if (row.get("flag") or "").strip().upper() == "REVIEW"
        ]


def generate_pipeline_status(period: str, accounts: list[str], reconciliation_rows: list[dict[str, str]],
                             filing_master_present: bool) -> list[dict[str, str]]:
    """Provide one visible list of non-editable gates alongside editable gaps."""
    review_count = len(reconciliation_rows)
    return [
        {"gate": "fund population", "status": "READY" if accounts else "BLOCKED",
         "detail": f"{len(set(accounts))} fund(s) in the security master",
         "resolution": "rerun masters with the correct custodian population"},
        {"gate": "filing master", "status": "READY" if filing_master_present else "BLOCKED",
         "detail": "filing_master.xlsx is present" if filing_master_present else "filing_master.xlsx is missing",
         "resolution": "rerun masters after correcting the received inputs"},
        {"gate": "reconciliation", "status": "READY" if review_count == 0 else "BLOCKED",
         "detail": f"{review_count} REVIEW row(s) for {period}",
         "resolution": "correct/replace the mismatched source input and rerun masters"},
    ]


def apply_fund_policy_to_configs(funds_dir: Path, review: Review) -> int:
    """Persist populated fund_policy cells to each fund config without inventing blanks."""
    updated = 0
    for fund, row in review.fund_policy().items():
        path = Path(funds_dir) / fund.lower() / "fund_config.txt"
        if not path.is_file():
            continue
        supplied = {key: (row.get(key) or "").strip() for key in _POLICY_COLUMNS
                    if (row.get(key) or "").strip()}
        if not supplied:
            continue
        lines = path.read_text(encoding="utf-8-sig").splitlines()
        seen: set[str] = set()
        out: list[str] = []
        for line in lines:
            stripped = line.strip()
            key = stripped.split("=", 1)[0].strip() if "=" in stripped and not stripped.startswith("#") else ""
            if key in supplied:
                out.append(f"{key}={supplied[key]}")
                seen.add(key)
            else:
                out.append(line)
        if any(key not in seen for key in supplied):
            out += ["", "# Human-review policy inputs"]
            out += [f"{key}={supplied[key]}" for key in _POLICY_COLUMNS if key in supplied and key not in seen]
        text = "\n".join(out).rstrip() + "\n"
        if text != path.read_text(encoding="utf-8-sig"):
            fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".txt")
            os.close(fd)
            try:
                Path(tmp).write_text(text, encoding="utf-8")
                Path(tmp).replace(path)
            except Exception:
                Path(tmp).unlink(missing_ok=True)
                raise
            updated += 1
    return updated


def _seed_rows(sheet: str) -> list[dict[str, str]]:
    """Seed rows for a reference sheet from the legacy literals (empty for per-item sheets)."""
    if sheet == "designated_index":
        return [{"ticker": t, "indexName": n, "indexIdentifier": i, "sourceNote": "seed: 497K"}
                for t, (n, i) in sorted(SEED_DESIGNATED_INDEX.items())]
    if sheet == "swap_counterparties":
        return [{"code": c, "legalName": n, "lei": lei, "sourceNote": "seed: GLEIF"}
                for c, (n, lei) in sorted(SEED_SWAP_COUNTERPARTIES.items())]
    if sheet == "option_index":
        return [{"underlying": u, "indexName": n, "indexIdentifier": i, "sourceNote": "seed: legacy map"}
                for u, (n, i) in sorted(SEED_OPTION_INDEX.items())]
    return []
